"""
Reporte consultable/exportable de creadores (listado administrativo).
No es un dashboard de rendimiento: usa creadores + detalle + estados existentes.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from psycopg2.extras import RealDictCursor

from DataBase import get_connection_context
from main_auth import manager_id_para_filtro, obtener_usuario_actual
from creadores_catalogo import CREADOR_ESTADO_NOMBRE_ACTIVO

logger = logging.getLogger("uvicorn.error")
router = APIRouter()

CREADOR_ESTADO_NOMBRE_INACTIVO = "Inactivo"
MAX_EXPORT_ROWS = 5000


def _iso_date(val: Any) -> Optional[str]:
    if val is None:
        return None
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def normalizar_estado_filtro(estado: Optional[str]) -> Optional[str]:
    if estado is None:
        return None
    key = str(estado).strip().lower()
    if key in ("", "todos", "all"):
        return None
    if key in ("activo", "activos"):
        return CREADOR_ESTADO_NOMBRE_ACTIVO
    if key in ("inactivo", "inactivos"):
        return CREADOR_ESTADO_NOMBRE_INACTIVO
    raise HTTPException(
        status_code=400,
        detail="estado inválido. Use: activo, inactivo o todos.",
    )


def construir_where(
    estado_nombre: Optional[str],
    manager_id: Optional[int],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    search: Optional[str],
) -> Tuple[str, List[Any]]:
    clauses = ["COALESCE(ce.activo, true) = true"]
    params: List[Any] = []

    if estado_nombre:
        clauses.append("ce.nombre = %s")
        params.append(estado_nombre)

    if manager_id is not None:
        clauses.append("cd.manager_id = %s")
        params.append(manager_id)

    if fecha_desde:
        clauses.append("cd.fecha_incorporacion >= %s")
        params.append(fecha_desde)

    if fecha_hasta:
        clauses.append("cd.fecha_incorporacion <= %s")
        params.append(fecha_hasta)

    termino = (search or "").strip()
    if termino:
        like = f"%{termino}%"
        clauses.append(
            """
            (
                COALESCE(c.nombre, '') ILIKE %s
                OR COALESCE(c.usuario_tiktok, '') ILIKE %s
                OR COALESCE(c.telefono, '') ILIKE %s
            )
            """
        )
        params.extend([like, like, like])

    return " AND ".join(clauses), params


SQL_FROM = """
FROM creadores c
INNER JOIN creadores_estados ce ON ce.id = c.estado_id
LEFT JOIN creadores_detalle cd ON cd.creador_id = c.id
LEFT JOIN administradores a ON a.id = cd.manager_id
"""

SQL_SELECT = f"""
SELECT
    c.id,
    COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(c.usuario_tiktok), ''), 'Sin nombre') AS nombre,
    COALESCE(NULLIF(TRIM(c.usuario_tiktok), ''), '') AS usuario_tiktok,
    COALESCE(NULLIF(TRIM(c.telefono), ''), '') AS telefono,
    cd.fecha_incorporacion,
    ce.nombre AS estado,
    cd.manager_id,
    COALESCE(NULLIF(TRIM(a.nombre_completo), ''), NULLIF(TRIM(a.username), ''), '') AS manager_nombre
{SQL_FROM}
"""


def _resolver_manager_id(usuario: dict, manager_id: Optional[int]) -> Optional[int]:
    forzado = manager_id_para_filtro(usuario)
    if forzado is not None:
        return forzado
    return manager_id


def _item_desde_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row.get("id"),
        "nombre": row.get("nombre") or "Sin nombre",
        "usuario_tiktok": row.get("usuario_tiktok") or "",
        "telefono": row.get("telefono") or "",
        "fecha_incorporacion": _iso_date(row.get("fecha_incorporacion")),
        "estado": row.get("estado") or "",
        "manager_id": row.get("manager_id"),
        "manager_nombre": row.get("manager_nombre") or "",
    }


def consultar_reporte(
    *,
    estado_nombre: Optional[str],
    manager_id: Optional[int],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    search: Optional[str],
    page: int,
    page_size: int,
    limit_override: Optional[int] = None,
) -> Dict[str, Any]:
    where_sql, params = construir_where(
        estado_nombre, manager_id, fecha_desde, fecha_hasta, search
    )
    offset = (page - 1) * page_size
    limit = limit_override if limit_override is not None else page_size

    with get_connection_context() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                f"SELECT COUNT(*) AS total {SQL_FROM} WHERE {where_sql}",
                tuple(params),
            )
            total = int((cur.fetchone() or {}).get("total") or 0)

            cur.execute(
                f"""
                {SQL_SELECT}
                WHERE {where_sql}
                ORDER BY COALESCE(NULLIF(TRIM(c.nombre), ''), c.usuario_tiktok, '') ASC,
                         c.id ASC
                LIMIT %s OFFSET %s
                """,
                tuple(params + [limit, offset]),
            )
            items = [_item_desde_row(dict(row)) for row in cur.fetchall()]

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def _generar_excel(items: List[Dict[str, Any]], filtros: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Creadores"

    headers = ["Creador", "Usuario TikTok", "Teléfono", "Manager", "Fecha de ingreso", "Estado"]
    header_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(bold=True, color="3730A3")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for item in items:
        ws.append(
            [
                item.get("nombre") or "",
                item.get("usuario_tiktok") or "",
                item.get("telefono") or "",
                item.get("manager_nombre") or "",
                (item.get("fecha_incorporacion") or "")[:10],
                item.get("estado") or "",
            ]
        )

    ws.append([])
    ws.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    partes = []
    if filtros.get("estado"):
        partes.append(f"Estado: {filtros['estado']}")
    if filtros.get("manager_nombre"):
        partes.append(f"Manager: {filtros['manager_nombre']}")
    if filtros.get("fecha_desde") or filtros.get("fecha_hasta"):
        partes.append(
            f"Ingreso: {filtros.get('fecha_desde') or '—'} a {filtros.get('fecha_hasta') or '—'}"
        )
    if filtros.get("search"):
        partes.append(f"Búsqueda: {filtros['search']}")
    if partes:
        ws.append(["Filtros: " + " · ".join(partes)])

    for col in ws.columns:
        maxlen = 12
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                maxlen = max(maxlen, min(len(str(cell.value)), 40))
        ws.column_dimensions[letter].width = maxlen + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generar_pdf(items: List[Dict[str, Any]], filtros: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise HTTPException(
            status_code=501,
            detail="Exportar PDF requiere la dependencia reportlab en el servidor.",
        ) from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Reporte de Creadores",
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Creadores", styles["Title"]))
    story.append(
        Paragraph(
            f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        )
    )
    partes = []
    if filtros.get("estado"):
        partes.append(f"Estado: {filtros['estado']}")
    if filtros.get("manager_nombre"):
        partes.append(f"Manager: {filtros['manager_nombre']}")
    if filtros.get("fecha_desde") or filtros.get("fecha_hasta"):
        partes.append(
            f"Ingreso: {filtros.get('fecha_desde') or '—'} a {filtros.get('fecha_hasta') or '—'}"
        )
    if filtros.get("search"):
        partes.append(f"Búsqueda: {filtros['search']}")
    if partes:
        story.append(Paragraph("Filtros: " + " · ".join(partes), styles["Normal"]))
    story.append(Spacer(1, 8))

    data = [["Creador", "Usuario TikTok", "Teléfono", "Manager", "Fecha ingreso", "Estado"]]
    for item in items:
        data.append(
            [
                item.get("nombre") or "—",
                item.get("usuario_tiktok") or "—",
                item.get("telefono") or "—",
                item.get("manager_nombre") or "—",
                (item.get("fecha_incorporacion") or "—")[:10],
                item.get("estado") or "—",
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#3730A3")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _filtros_etiqueta(
    estado_nombre: Optional[str],
    manager_id: Optional[int],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    search: Optional[str],
    items: List[Dict[str, Any]],
) -> Dict[str, Any]:
    manager_nombre = ""
    if manager_id is not None:
        for item in items:
            if item.get("manager_id") == manager_id and item.get("manager_nombre"):
                manager_nombre = item["manager_nombre"]
                break
        if not manager_nombre:
            try:
                with get_connection_context() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            SELECT COALESCE(NULLIF(TRIM(nombre_completo), ''), username)
                            FROM administradores
                            WHERE id = %s
                            """,
                            (manager_id,),
                        )
                        row = cur.fetchone()
                        if row:
                            manager_nombre = row[0] or ""
            except Exception:
                logger.exception("No se pudo resolver nombre de manager para export")
    return {
        "estado": estado_nombre,
        "manager_nombre": manager_nombre,
        "fecha_desde": _iso_date(fecha_desde),
        "fecha_hasta": _iso_date(fecha_hasta),
        "search": (search or "").strip(),
    }


@router.get("/api/reportes/creadores")
def listar_reporte_creadores(
    estado: Optional[str] = Query(None),
    manager_id: Optional[int] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    usuario: dict = Depends(obtener_usuario_actual),
):
    estado_nombre = normalizar_estado_filtro(estado)
    mid = _resolver_manager_id(usuario, manager_id)
    try:
        return consultar_reporte(
            estado_nombre=estado_nombre,
            manager_id=mid,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            search=search,
            page=page,
            page_size=page_size,
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error listando reporte de creadores")
        raise HTTPException(status_code=500, detail="Error al consultar el reporte de creadores")


def _datos_exportacion(
    usuario: dict,
    estado: Optional[str],
    manager_id: Optional[int],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    search: Optional[str],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    estado_nombre = normalizar_estado_filtro(estado)
    mid = _resolver_manager_id(usuario, manager_id)
    data = consultar_reporte(
        estado_nombre=estado_nombre,
        manager_id=mid,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        search=search,
        page=1,
        page_size=MAX_EXPORT_ROWS,
        limit_override=MAX_EXPORT_ROWS,
    )
    items = data["items"]
    filtros = _filtros_etiqueta(
        estado_nombre, mid, fecha_desde, fecha_hasta, search, items
    )
    return items, filtros


@router.get("/api/reportes/creadores/export.xlsx")
def exportar_reporte_creadores_excel(
    estado: Optional[str] = Query(None),
    manager_id: Optional[int] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    usuario: dict = Depends(obtener_usuario_actual),
):
    try:
        items, filtros = _datos_exportacion(
            usuario, estado, manager_id, fecha_desde, fecha_hasta, search
        )
        contenido = _generar_excel(items, filtros)
        return Response(
            content=contenido,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="reporte_creadores.xlsx"'
            },
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error exportando Excel de creadores")
        raise HTTPException(status_code=500, detail="Error al exportar Excel")


@router.get("/api/reportes/creadores/export.pdf")
def exportar_reporte_creadores_pdf(
    estado: Optional[str] = Query(None),
    manager_id: Optional[int] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    usuario: dict = Depends(obtener_usuario_actual),
):
    try:
        items, filtros = _datos_exportacion(
            usuario, estado, manager_id, fecha_desde, fecha_hasta, search
        )
        contenido = _generar_pdf(items, filtros)
        return Response(
            content=contenido,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="reporte_creadores.pdf"'},
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error exportando PDF de creadores")
        raise HTTPException(status_code=500, detail="Error al exportar PDF")

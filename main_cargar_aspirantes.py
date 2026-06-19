import os
import json
import uuid
import re
import logging

import gspread
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials

from fastapi import APIRouter, UploadFile, File, Form, Body, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field, field_validator

from DataBase import get_connection_context, limpiar_telefono, safe_int
from utils_aspirantes import registrar_cambio_estado  # <-- AJUSTAR IMPORT REAL

logger = logging.getLogger("uvicorn.error")
router = APIRouter()

ESTADO_INICIAL_ID = 1


class AspiranteManualCreate(BaseModel):
    nickname: str
    usuario: str
    siguiendo: int = Field(default=0, ge=0)
    seguidores: int = Field(default=0, ge=0)
    likes: int = Field(default=0, ge=0)
    biografia: str | None = None

    @field_validator("usuario")
    @classmethod
    def normalizar_usuario(cls, v: str) -> str:
        usuario = (v or "").strip()
        if usuario.startswith("@"):
            usuario = usuario[1:].strip()
        if not usuario:
            raise ValueError("usuario es obligatorio")
        return usuario

    @field_validator("nickname")
    @classmethod
    def validar_nickname(cls, v: str) -> str:
        nickname = (v or "").strip()
        if not nickname:
            raise ValueError("nickname es obligatorio")
        return nickname

    @field_validator("biografia")
    @classmethod
    def normalizar_biografia(cls, v: str | None) -> str | None:
        if v is None:
            return None
        texto = str(v).strip()
        return texto or None


# =========================================================
# Helpers
# =========================================================

def _keep(s: str) -> str:
    """Conserva el texto tal cual, recortando solo espacios al inicio/fin."""
    return (s or "").strip()


def get_text(val):
    """Texto limpio (str) o cadena vacía si viene None."""
    return "" if val is None else str(val).strip()


def to_int_relaxed(val, default=0):
    """
    Convierte '1,151' o '47 237' -> 1151 / 47237.
    - Elimina espacios y comas.
    - Si no se puede convertir, devuelve default.
    """
    if val is None:
        return default

    s = str(val).strip()
    if not s:
        return default

    s = s.replace(" ", "").replace(",", "")

    try:
        return int(s)
    except Exception:
        return default


def parse_days_to_int(val, default=0):
    """
    Convierte '24d' -> 24, '1d' -> 1, '0d' -> 0, '15' -> 15.
    """
    if val is None:
        return default

    s = str(val).strip().lower()
    if not s:
        return default

    if s.endswith("d"):
        s = s[:-1]

    try:
        return int(s)
    except ValueError:
        return to_int_relaxed(s, default=default)


def parse_duration_to_hours_int(text, default=0):
    """
    Convierte valores como:
    '3d 2h 54m 10s' -> horas enteras redondeadas
    """
    if text is None:
        return default

    s = str(text).strip().lower()
    if not s:
        return default

    s = s.replace("min", "m").replace("mins", "m").replace("sec", "s").replace("secs", "s")

    days = hours = mins = secs = 0

    md = re.search(r"(\d+)\s*d", s)
    mh = re.search(r"(\d+)\s*h", s)
    mm = re.search(r"(\d+)\s*m(?![a-z])", s)
    ms = re.search(r"(\d+)\s*s", s)

    if md:
        days = int(md.group(1))
    if mh:
        hours = int(mh.group(1))
    if mm:
        mins = int(mm.group(1))
    if ms:
        secs = int(ms.group(1))

    if not (md or mh or mm or ms):
        try:
            return int(round(float(s)))
        except ValueError:
            return default

    total_hours = days * 24 + hours + mins / 60.0 + secs / 3600.0
    return int(round(total_hours))


def _parse_fecha_solicitud(s: str | None):
    """
    Intenta parsear una fecha en varios formatos comunes.
    Devuelve datetime o None si no se puede parsear.
    """
    if not s:
        return None

    s = str(s).strip()

    formatos = [
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M",
        "%d-%m-%Y %H:%M",
        "%m-%d-%Y %H:%M",
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    try:
        s2 = " ".join(s.split())
        for fmt in formatos:
            try:
                return datetime.strptime(s2, fmt)
            except ValueError:
                continue
    except Exception:
        pass

    return None


# =========================================================
# Parser TXT
# =========================================================

def parsear_bloques_desde_txt(ruta_txt: str | Path) -> dict:
    texto = Path(ruta_txt).read_text(encoding="utf-8", errors="ignore")
    lineas = [ln.rstrip("\n\r") for ln in texto.splitlines()]
    i, n = 0, len(lineas)
    out = {}

    _cnt_total = 0
    _cnt_con_fecha = 0
    _cnt_con_canal = 0
    _ejemplos = []

    while i < n:
        linea = lineas[i].strip()
        i += 1

        if not linea:
            continue

        usuario = linea

        while i < n and lineas[i].strip() == "":
            i += 1
        if i < n and lineas[i].strip().lower() in ("nombre", "name"):
            i += 1

        while i < n and lineas[i].strip() == "":
            i += 1
        nombre = _keep(lineas[i]) if i < n else ""
        i += 1

        while i < n and lineas[i].strip() == "":
            i += 1
        metrica_line = lineas[i].strip() if i < n else ""
        i += 1

        m = re.match(r"([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)", metrica_line)
        if m:
            s_seg, s_vid, s_like, s_dur, s_dias = m.groups()
        else:
            partes = metrica_line.split("\t") if metrica_line else []
            while len(partes) < 5:
                partes += [""]
            s_seg, s_vid, s_like, s_dur, s_dias = partes[:5]

        seguidores = _keep(s_seg)
        videos = _keep(s_vid)
        likes = _keep(s_like)
        duracion = _keep(s_dur)
        dias_validos = _keep(s_dias)

        while i < n and lineas[i].strip() == "":
            i += 1
        caducidad = _keep(lineas[i]) if i < n else ""
        i += 1

        while i < n and lineas[i].strip() == "":
            i += 1
        agente = _keep(lineas[i]) if i < n else ""
        i += 1

        while i < n and lineas[i].strip() == "":
            i += 1

        fecha_sol = ""
        canal = ""

        if i < n:
            ult = lineas[i].strip()
            i += 1

            partes = ult.split("\t")
            if len(partes) >= 2:
                canal = _keep(partes[0])
                fecha_sol = _keep(partes[1])
            else:
                mm = re.search(r"\b\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}\b", ult)
                if mm:
                    fecha_sol = mm.group(0)

                canal_match = re.match(r"^\s*([A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+)", ult)
                if canal_match:
                    canal = _keep(canal_match.group(1))

        out[usuario] = {
            "usuario": usuario,
            "nombre": nombre,
            "seguidores": seguidores,
            "videos": videos,
            "likes": likes,
            "Duracion_Emisiones": duracion,
            "Dias_Emisiones": dias_validos,
            "caducidad_solicitud": caducidad,
            "agente_recluta": agente,
            "fecha_solicitud": fecha_sol,
            "canal": canal,
        }

        _cnt_total += 1
        if canal:
            _cnt_con_canal += 1
        if fecha_sol:
            _cnt_con_fecha += 1
            if len(_ejemplos) < 3:
                _ejemplos.append((usuario, canal, fecha_sol))

        while i < n and lineas[i].strip() == "":
            i += 1

    logger.info(
        f"TXT parseado: total={_cnt_total}, con_canal={_cnt_con_canal}, "
        f"con_fecha={_cnt_con_fecha}, ejemplos={_ejemplos}"
    )
    return out


# =========================================================
# Google Sheets
# =========================================================

def get_gspread_client():
    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    cred_json = os.getenv("GOOGLE_CREDENTIALS_JSON_VIEJA")
    cred_dict = json.loads(cred_json)
    creds = Credentials.from_service_account_info(cred_dict, scopes=scope)
    return gspread.authorize(creds)


def obtener_aspirantes_desde_hoja(str_key, nombre_hoja, ruta_txt):
    try:
        info_por_usuario = parsear_bloques_desde_txt(ruta_txt)

        gc = get_gspread_client()
        spreadsheet = gc.open_by_key(str_key)
        worksheet = spreadsheet.worksheet(nombre_hoja)

        col_B_desde_4 = worksheet.col_values(2)[3:]
        col_B_no_vacias = [c for c in col_B_desde_4 if str(c).strip() != ""]
        ultima_fila = 3 + len(col_B_no_vacias)

        rango = f"A4:X{ultima_fila}"
        filas = worksheet.get(rango)

        if not filas and col_B_desde_4:
            rango_fallback = f"A4:X{3 + len(col_B_desde_4)}"
            filas = worksheet.get(rango_fallback)

        aspirantes = []

        for i, fila in enumerate(filas):
            if len(fila) < 24:
                fila += [""] * (24 - len(fila))

            usuario = fila[1].strip()
            if not usuario:
                continue

            aspirante = {
                "usuario": usuario,
                "telefono": fila[2].strip().replace(" ", "").replace("+", ""),
                "disponibilidad": fila[3].strip(),
                "motivo_no_apto": fila[4].strip().upper(),
                "perfil": fila[5].strip(),
                "contacto": fila[8].strip(),
                "respuesta_creador": fila[9].strip(),
                "entrevista": fila[11].strip(),
                "tipo_solicitud": fila[15].strip(),
                "email": fila[16].strip(),
                "nickname": fila[17].strip(),
                "razon_no_contacto": fila[18].strip().upper(),
                "seguidores": "",
                "videos": "",
                "likes": "",
                "Duracion_Emisiones": "",
                "Dias_Emisiones": "",
                "fila_excel": i + 4,
            }

            datos_txt = info_por_usuario.get(usuario, {})
            if datos_txt:
                aspirante["seguidores"] = datos_txt.get("seguidores", "")
                aspirante["videos"] = datos_txt.get("videos", "")
                aspirante["likes"] = datos_txt.get("likes", "")
                aspirante["Duracion_Emisiones"] = datos_txt.get("Duracion_Emisiones", "")
                aspirante["Dias_Emisiones"] = datos_txt.get("Dias_Emisiones", "")
                aspirante["nombre"] = datos_txt.get("nombre", "")
                aspirante["caducidad_solicitud"] = datos_txt.get("caducidad_solicitud", "")
                aspirante["agente_recluta"] = datos_txt.get("agente_recluta", "")
                aspirante["fecha_solicitud"] = datos_txt.get("fecha_solicitud", "")
                aspirante["canal"] = datos_txt.get("canal", "")

            aspirantes.append(aspirante)

        return aspirantes

    except Exception as e:
        logger.error(f"❌ Error leyendo hoja de cálculo: {e}", exc_info=True)
        return []


# =========================================================
# Guardado principal
# =========================================================

def guardar_aspirantes(
    aspirantes,
    nombre_archivo=None,
    hoja_excel=None,
    lote_carga=None,
    procesado_por=None,
    observaciones=None
):
    """
    Guarda aspirantes y asegura estado inicial = 1.
    - Nuevos aspirantes: se insertan con estado_id = 1
    - Aspirantes existentes: se actualizan (incl. tiene_solicitud = TRUE) y luego se normaliza estado con registrar_cambio_estado()
    """
    conn = get_connection_context()
    cur = conn.cursor()

    resultados = []
    filas_fallidas = []

    for c in aspirantes:
        try:
            usuario = get_text(c.get("usuario"))
            nickname = get_text(c.get("nickname"))
            email = get_text(c.get("email"))
            telefono = limpiar_telefono(get_text(c.get("telefono")))
            disponibilidad = get_text(c.get("disponibilidad"))
            perfil = get_text(c.get("perfil"))
            motivo_no_apto = get_text(c.get("motivo_no_apto"))
            contacto = get_text(c.get("contacto"))
            respuesta_creador = get_text(c.get("respuesta_creador"))
            entrevista = get_text(c.get("entrevista"))
            tipo_solicitud = get_text(c.get("tipo_solicitud"))
            razon_no_contacto = get_text(c.get("razon_no_contacto"))

            fecha_sol_txt = c.get("fecha_solicitud") or c.get("fecha_solcitud")
            fecha_solicitud_dt = _parse_fecha_solicitud(fecha_sol_txt)

            if fecha_sol_txt:
                logger.info(f"[{usuario}] fecha_solicitud TXT='{fecha_sol_txt}' -> parsed={fecha_solicitud_dt}")
            else:
                logger.info(f"[{usuario}] SIN fecha_solicitud en el aspirante")

            seguidores = to_int_relaxed(c.get("seguidores"), default=0)
            cantidad_videos = to_int_relaxed(c.get("videos"), default=0)
            likes_totales = to_int_relaxed(c.get("likes"), default=0)
            duracion_emisiones = parse_duration_to_hours_int(c.get("Duracion_Emisiones"), default=0)
            dias_emisiones = parse_days_to_int(c.get("Dias_Emisiones"), default=0)

            fila_excel = c.get("fila_excel")
            apto = not bool(motivo_no_apto)

            aspirante_id = None
            aspirante_existia = False

            # =====================================================
            # ASPIRANTES
            # =====================================================
            cur.execute(
                """
                SELECT id
                FROM aspirantes
                WHERE usuario = %s
                """,
                (usuario,)
            )
            aspirante_row = cur.fetchone()

            if aspirante_row:
                aspirante_id = aspirante_row[0]
                aspirante_existia = True

                if fecha_solicitud_dt:
                    cur.execute(
                        """
                        UPDATE aspirantes
                        SET nickname = %s,
                            email = %s,
                            telefono = %s,
                            fecha_solicitud = %s,
                            tiene_solicitud = TRUE,
                            actualizado_en = NOW()
                        WHERE id = %s
                        """,
                        (
                            get_text(c.get("nombre")),
                            email,
                            telefono,
                            fecha_solicitud_dt,
                            aspirante_id
                        )
                    )
                else:
                    cur.execute(
                        """
                        UPDATE aspirantes
                        SET nickname = %s,
                            email = %s,
                            telefono = %s,
                            tiene_solicitud = TRUE,
                            actualizado_en = NOW()
                        WHERE id = %s
                        """,
                        (
                            get_text(c.get("nombre")),
                            email,
                            telefono,
                            aspirante_id
                        )
                    )

            else:
                if fecha_solicitud_dt:
                    cur.execute(
                        """
                        INSERT INTO aspirantes (
                            usuario,
                            nickname,
                            email,
                            telefono,
                            fecha_solicitud,
                            estado_id,
                            activo,
                            creado_en,
                            actualizado_en
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, TRUE, NOW(), NOW())
                        RETURNING id
                        """,
                        (
                            usuario,
                            get_text(c.get("nombre")),
                            email,
                            telefono,
                            fecha_solicitud_dt,
                            ESTADO_INICIAL_ID
                        )
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO aspirantes (
                            usuario,
                            nickname,
                            email,
                            telefono,
                            estado_id,
                            activo,
                            creado_en,
                            actualizado_en
                        )
                        VALUES (%s, %s, %s, %s, %s, TRUE, NOW(), NOW())
                        RETURNING id
                        """,
                        (
                            usuario,
                            get_text(c.get("nombre")),
                            email,
                            telefono,
                            ESTADO_INICIAL_ID
                        )
                    )

                aspirante_id = cur.fetchone()[0]

            # =====================================================
            # ASPIRANTES_PERFIL
            # =====================================================
            cur.execute(
                """
                SELECT id
                FROM aspirantes_perfil
                WHERE aspirante_id = %s
                """,
                (aspirante_id,)
            )
            perfil_row = cur.fetchone()

            if perfil_row:
                cur.execute(
                    """
                    UPDATE aspirantes_perfil
                    SET usuario = %s,
                        seguidores = %s,
                        videos = %s,
                        likes = %s,
                        duracion_emisiones = %s,
                        dias_emisiones = %s,
                        nombre = %s,
                        actualizado_en = NOW()
                    WHERE aspirante_id = %s
                    """,
                    (
                        usuario,
                        seguidores,
                        cantidad_videos,
                        likes_totales,
                        duracion_emisiones,
                        dias_emisiones,
                        get_text(c.get("nombre")),
                        aspirante_id
                    )
                )
            else:
                cur.execute(
                    """
                    INSERT INTO aspirantes_perfil (
                        usuario,
                        aspirante_id,
                        seguidores,
                        videos,
                        likes,
                        duracion_emisiones,
                        dias_emisiones,
                        nombre,
                        creado_en,
                        actualizado_en
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """,
                    (
                        usuario,
                        aspirante_id,
                        seguidores,
                        cantidad_videos,
                        likes_totales,
                        duracion_emisiones,
                        dias_emisiones,
                        get_text(c.get("nombre"))
                    )
                )

            # =====================================================
            # ASPIRANTES_CARGUE
            # =====================================================
            cur.execute(
                """
                SELECT id
                FROM aspirantes_cargue
                WHERE usuario = %s
                  AND hoja_excel = %s
                """,
                (usuario, hoja_excel)
            )
            cargue_row = cur.fetchone()

            if cargue_row:
                cargue_id = cargue_row[0]

                cur.execute(
                    """
                    UPDATE aspirantes_cargue
                    SET nickname = %s,
                        email = %s,
                        telefono = %s,
                        disponibilidad = %s,
                        perfil = %s,
                        motivo_no_apto = %s,
                        contacto = %s,
                        respuesta_creador = %s,
                        entrevista = %s,
                        tipo_solicitud = %s,
                        razon_no_contacto = %s,
                        seguidores = %s,
                        cantidad_videos = %s,
                        likes_totales = %s,
                        duracion_emisiones = %s,
                        dias_emisiones = %s,
                        nombre_archivo = %s,
                        fila_excel = %s,
                        lote_carga = %s,
                        estado = %s,
                        procesado = %s,
                        procesado_por = %s,
                        aspirante_id = %s,
                        apto = %s,
                        observaciones = %s,
                        actualizado_en = NOW()
                    WHERE id = %s
                    """,
                    (
                        nickname,
                        email,
                        telefono,
                        disponibilidad,
                        perfil,
                        motivo_no_apto,
                        contacto,
                        respuesta_creador,
                        entrevista,
                        tipo_solicitud,
                        razon_no_contacto,
                        seguidores,
                        cantidad_videos,
                        likes_totales,
                        duracion_emisiones,
                        dias_emisiones,
                        nombre_archivo,
                        fila_excel,
                        lote_carga,
                        "Procesando",
                        False,
                        procesado_por,
                        aspirante_id,
                        apto,
                        observaciones,
                        cargue_id
                    )
                )
            else:
                cur.execute(
                    """
                    INSERT INTO aspirantes_cargue (
                        usuario,
                        nickname,
                        email,
                        telefono,
                        disponibilidad,
                        perfil,
                        motivo_no_apto,
                        contacto,
                        respuesta_creador,
                        entrevista,
                        tipo_solicitud,
                        razon_no_contacto,
                        seguidores,
                        cantidad_videos,
                        likes_totales,
                        duracion_emisiones,
                        dias_emisiones,
                        nombre_archivo,
                        hoja_excel,
                        fila_excel,
                        lote_carga,
                        estado,
                        procesado,
                        procesado_por,
                        aspirante_id,
                        apto,
                        observaciones,
                        activo,
                        creado_en,
                        actualizado_en
                    )
                    VALUES (
                        %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, TRUE, NOW(), NOW()
                    )
                    """,
                    (
                        usuario,
                        nickname,
                        email,
                        telefono,
                        disponibilidad,
                        perfil,
                        motivo_no_apto,
                        contacto,
                        respuesta_creador,
                        entrevista,
                        tipo_solicitud,
                        razon_no_contacto,
                        seguidores,
                        cantidad_videos,
                        likes_totales,
                        duracion_emisiones,
                        dias_emisiones,
                        nombre_archivo,
                        hoja_excel,
                        fila_excel,
                        lote_carga,
                        "Procesando",
                        False,
                        procesado_por,
                        aspirante_id,
                        apto,
                        observaciones
                    )
                )

            # =====================================================
            # CAMBIO DE ESTADO PARA EXISTENTES
            # =====================================================
            if aspirante_existia:
                conn.commit()

                try:
                    registrar_cambio_estado(
                        aspirante_id=aspirante_id,
                        nuevo_estado_id=ESTADO_INICIAL_ID,
                        usuario_id=procesado_por,
                        origen_cambio="cargue_masivo",
                        observacion="Aspirante actualizado desde cargue masivo y normalizado a estado inicial"
                    )
                except Exception as e_estado:
                    logger.error(
                        f"❌ Error actualizando estado del aspirante {aspirante_id}: {e_estado}",
                        exc_info=True
                    )

            resultados.append({
                "fila": fila_excel,
                "usuario": usuario,
                "aspirante_id": aspirante_id,
                "estado_id": ESTADO_INICIAL_ID
            })

        except Exception as e:
            logger.error(
                f"Error al guardar aspirante (fila={c.get('fila_excel')}, usuario={c.get('usuario')}): {e}",
                exc_info=True
            )
            filas_fallidas.append({
                "fila": c.get("fila_excel"),
                "usuario": c.get("usuario"),
                "error": str(e),
            })

    conn.commit()
    cur.close()
    conn.close()

    print(f"✅ Contactos procesados. Filas exitosas: {len(resultados)}. Filas fallidas: {len(filas_fallidas)}")
    return {
        "status": "ok",
        "exitosos": resultados,
        "fallidos": filas_fallidas
    }


# =========================================================
# Endpoints
# =========================================================

@router.post("/upload_txt")
async def upload_txt(file: UploadFile = File(...)):
    filename = file.filename or "datos.txt"

    if not filename.lower().endswith(".txt"):
        return {"status": "error", "mensaje": "El archivo debe ser .txt"}

    base_dir = Path("uploads/txt")
    base_dir.mkdir(parents=True, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}_{filename}"
    destino = base_dir / unique_name

    contenido = await file.read()
    destino.write_bytes(contenido)

    return {"status": "ok", "ruta_txt": str(destino)}


@router.get("/listar_hojas")
def listar_hojas(str_key: str):
    """
    Devuelve la lista de hojas disponibles en el documento Google Sheets.
    """
    try:
        gc = get_gspread_client()
        spreadsheet = gc.open_by_key(str_key)
        hojas = [ws.title for ws in spreadsheet.worksheets()]
        return {"status": "ok", "hojas": hojas}
    except Exception as e:
        logger.error(f"❌ Error en listar_hojas con str_key={str_key}: {e}", exc_info=True)
        return {"status": "error", "mensaje": f"Error al listar hojas: {str(e)}"}


@router.post("/cargar_aspirantes")
async def cargar_aspirantes_desde_workspace(
    nombre_hoja: str = Form(...),
    str_key: str = Form(...),
    txt_file: UploadFile = File(None),
    ruta_txt: str = Form(None),
):
    try:
        logger.info(f"📥 Iniciando carga: hoja={nombre_hoja}, str_key={str_key}")

        if txt_file is not None:
            if not (txt_file.filename or "").lower().endswith(".txt"):
                return {"status": "error", "mensaje": "El archivo TXT debe tener extensión .txt"}

            tmp_dir = Path("/tmp/txt")
            tmp_dir.mkdir(parents=True, exist_ok=True)

            safe_name = f"{uuid.uuid4().hex}_{os.path.basename(txt_file.filename)}"
            txt_path = tmp_dir / safe_name
            txt_path.write_bytes(await txt_file.read())

            ruta_txt_resuelta = str(txt_path)
            logger.info(f"🗂️ TXT guardado en {ruta_txt_resuelta}")

        else:
            if not ruta_txt:
                return {
                    "status": "error",
                    "mensaje": "Falta el TXT: envía txt_file (archivo) o ruta_txt (servidor)."
                }

            if not Path(ruta_txt).exists():
                return {
                    "status": "error",
                    "mensaje": f"Ruta TXT no encontrada en el servidor: {ruta_txt}"
                }

            ruta_txt_resuelta = ruta_txt

        aspirantes = obtener_aspirantes_desde_hoja(str_key, nombre_hoja, ruta_txt_resuelta)

        if not aspirantes:
            logger.warning(f"⚠️ No se encontraron aspirantes en la hoja {nombre_hoja}")
            return {
                "status": "error",
                "mensaje": "No se encontraron aspirantes en la hoja"
            }

        guardar_aspirantes(
            aspirantes,
            nombre_archivo=None,
            hoja_excel=nombre_hoja
        )

        logger.info(f"✅ {len(aspirantes)} aspirantes cargados y guardados correctamente")
        return {
            "status": "ok",
            "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"
        }

    except Exception as e:
        logger.error(f"❌ Error en cargar_aspirantes (workspace): {e}", exc_info=True)
        return {
            "status": "error",
            "mensaje": f"Error al cargar aspirantes: {str(e)}"
        }


@router.post("/cargar_aspirantes_local")
async def cargar_aspirantes_desde_archivo(
    file: UploadFile = File(...),
    txt_file: UploadFile = File(...),
    nombre_hoja: str | None = Form(None),
):
    """
    Carga aspirantes desde un Excel local y un TXT con métricas/nickname.
    """
    xlsx_path = None
    txt_path = None
    wb = None

    try:
        # =========================
        # Guardar TXT
        # =========================
        info_por_usuario = {}

        txt_name = (txt_file.filename or "").lower()
        if not txt_name.endswith(".txt"):
            return {"status": "error", "mensaje": "El archivo TXT debe tener extensión .txt"}

        tmp_dir = Path("/tmp")
        tmp_dir.mkdir(parents=True, exist_ok=True)

        safe_txt = f"txt_{uuid.uuid4().hex}_{os.path.basename(txt_file.filename)}"
        txt_path = tmp_dir / safe_txt
        txt_path.write_bytes(await txt_file.read())

        info_por_usuario = parsear_bloques_desde_txt(str(txt_path))
        logger.info(f"✅ TXT parseado, usuarios encontrados: {len(info_por_usuario)}")

        # =========================
        # Guardar Excel
        # =========================
        fname = (file.filename or "").lower()
        if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
            return {"status": "error", "mensaje": "El archivo Excel debe ser .xlsx o .xlsm"}

        safe_name = f"local_{uuid.uuid4().hex}_{os.path.basename(file.filename)}"
        xlsx_path = tmp_dir / safe_name
        xlsx_path.write_bytes(await file.read())

        # =========================
        # Abrir Excel y seleccionar hoja
        # =========================
        wb = load_workbook(filename=str(xlsx_path), data_only=True)

        if nombre_hoja:
            if nombre_hoja not in wb.sheetnames:
                return {"status": "error", "mensaje": f"La hoja '{nombre_hoja}' no existe en el Excel"}
            ws = wb[nombre_hoja]
        else:
            ws = wb[wb.sheetnames[0]]

        logger.info(f"📄 Hoja usada (local): {ws.title}")

        # =========================
        # Detectar última fila
        # =========================
        col_B = 2
        start_row = 4
        consecutive_empties = 0
        consecutive_empties_limit = 8
        ultima_fila = start_row - 1

        r = start_row
        while True:
            val = ws.cell(row=r, column=col_B).value
            s = ("" if val is None else str(val).strip())

            if s == "":
                consecutive_empties += 1
                if consecutive_empties >= consecutive_empties_limit:
                    break
            else:
                consecutive_empties = 0
                ultima_fila = r

            r += 1

        if ultima_fila < start_row:
            logger.warning("⚠️ No se encontraron usuarios en la columna B (fila 4 en adelante)")
            return {"status": "error", "mensaje": "No se encontraron aspirantes en el archivo"}

        logger.info(f"📊 Rango calculado (local): A{start_row}:X{ultima_fila}")

        # =========================
        # Construir aspirantes
        # =========================
        aspirantes = []

        for i in range(start_row, ultima_fila + 1):
            fila_vals = []
            for c in range(1, 24 + 1):
                v = ws.cell(row=i, column=c).value
                fila_vals.append("" if v is None else str(v).strip())

            usuario = fila_vals[1]
            if not usuario:
                continue

            asp = {
                "usuario": usuario,
                "telefono": fila_vals[2].replace(" ", "").replace("+", ""),
                "disponibilidad": fila_vals[3],
                "motivo_no_apto": fila_vals[4].upper(),
                "perfil": fila_vals[5],
                "contacto": fila_vals[8],
                "respuesta_creador": fila_vals[9],
                "entrevista": fila_vals[11],
                "tipo_solicitud": fila_vals[15],
                "email": fila_vals[16],
                "nickname": fila_vals[17],
                "razon_no_contacto": fila_vals[18].upper(),
                "seguidores": "",
                "videos": "",
                "likes": "",
                "Duracion_Emisiones": "",
                "Dias_Emisiones": "",
                "fila_excel": i,
            }

            dtx = info_por_usuario.get(usuario, {})
            if dtx:
                asp["seguidores"] = dtx.get("seguidores", "")
                asp["videos"] = dtx.get("videos", "")
                asp["likes"] = dtx.get("likes", "")
                asp["Duracion_Emisiones"] = dtx.get("Duracion_Emisiones", "")
                asp["Dias_Emisiones"] = dtx.get("Dias_Emisiones", "")
                asp["nombre"] = dtx.get("nombre", "")
                asp["caducidad_solicitud"] = dtx.get("caducidad_solicitud", "")
                asp["agente_recluta"] = dtx.get("agente_recluta", "")
                asp["fecha_solicitud"] = dtx.get("fecha_solicitud", "")
                asp["canal"] = dtx.get("canal", "")

            aspirantes.append(asp)

        if not aspirantes:
            return {"status": "error", "mensaje": "No se construyeron aspirantes desde el archivo"}

        guardar_aspirantes(
            aspirantes,
            nombre_archivo=xlsx_path.name,
            hoja_excel=ws.title
        )

        logger.info(f"✅ {len(aspirantes)} aspirantes (local) cargados y guardados correctamente")
        return {
            "status": "ok",
            "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"
        }

    except Exception as e:
        logger.error(f"❌ Error en cargar_aspirantes_local: {e}", exc_info=True)
        return {
            "status": "error",
            "mensaje": f"Error al cargar archivo local: {str(e)}"
        }

    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass

        try:
            if xlsx_path and Path(xlsx_path).exists():
                os.remove(xlsx_path)
        except Exception:
            pass

        try:
            if txt_path and Path(txt_path).exists():
                os.remove(txt_path)
        except Exception:
            pass


# =========================================================
# Alta manual de aspirante (perfil público TikTok)
# =========================================================
#
# Ejemplo payload (Swagger / Postman):
# {
#   "nickname": "Karito Villamil 💜",
#   "usuario": "karitovillamil",
#   "siguiendo": 480,
#   "seguidores": 960,
#   "likes": 1757,
#   "biografia": "Team DesKarito"
# }


@router.post("/aspirantes/manual")
def crear_aspirante_manual(data: AspiranteManualCreate):
    usuario = data.usuario
    nickname = data.nickname
    biografia = data.biografia

    if biografia is not None and len(biografia) > 200:
        raise HTTPException(
            status_code=400,
            detail="La biografía no puede superar 200 caracteres (varchar(200)).",
        )

    logger.info(f"📝 Iniciando creación manual de aspirante usuario={usuario}")

    try:
        with get_connection_context() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT id, tiene_solicitud
                    FROM aspirantes
                    WHERE usuario = %s
                    LIMIT 1
                    """,
                    (usuario,),
                )
                existente = cur.fetchone()
                if existente:
                    aspirante_id = existente[0]
                    cur.execute(
                        """
                        UPDATE aspirantes
                        SET tiene_solicitud = TRUE,
                            actualizado_en = NOW()
                        WHERE id = %s
                        """,
                        (aspirante_id,),
                    )
                    logger.info(
                        f"✅ Usuario ya existía; tiene_solicitud actualizado id={aspirante_id} usuario={usuario}"
                    )
                    return {
                        "status": "ok",
                        "mensaje": "Aspirante ya existía; tiene_solicitud actualizado a true",
                        "aspirante": {
                            "id": aspirante_id,
                            "usuario": usuario,
                            "nickname": nickname,
                            "tiene_solicitud": True,
                            "fecha_solicitud": None,
                            "actualizado": True,
                        },
                    }

                cur.execute(
                    """
                    INSERT INTO aspirantes (
                        usuario,
                        nickname,
                        estado_id,
                        verificado,
                        activo,
                        encuesta_terminada,
                        tiene_solicitud,
                        fecha_solicitud,
                        creado_en,
                        actualizado_en
                    )
                    VALUES (
                        %s, %s, %s, FALSE, TRUE, FALSE, FALSE, NULL, NOW(), NOW()
                    )
                    RETURNING id
                    """,
                    (usuario, nickname, ESTADO_INICIAL_ID),
                )
                aspirante_id = cur.fetchone()[0]

                cur.execute(
                    """
                    INSERT INTO aspirantes_perfil (
                        usuario,
                        aspirante_id,
                        nombre,
                        seguidores,
                        siguiendo,
                        likes,
                        biografia,
                        videos,
                        estado,
                        creado_en,
                        actualizado_en,
                        fecha_actualizacion
                    )
                    VALUES (
                        %s, %s, %s, %s, %s, %s, %s, 0, 'activo', NOW(), NOW(), NOW()
                    )
                    """,
                    (
                        usuario,
                        aspirante_id,
                        nickname,
                        data.seguidores,
                        data.siguiendo,
                        data.likes,
                        biografia,
                    ),
                )

        logger.info(
            f"✅ Aspirante manual creado id={aspirante_id} usuario={usuario}"
        )

        return {
            "status": "ok",
            "mensaje": "Aspirante manual creado correctamente",
            "aspirante": {
                "id": aspirante_id,
                "usuario": usuario,
                "nickname": nickname,
                "tiene_solicitud": False,
                "fecha_solicitud": None,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            f"❌ Error creando aspirante manual usuario={usuario}: {e}",
            exc_info=True,
        )
        raise HTTPException(
            status_code=500,
            detail=f"Error al crear aspirante manual: {str(e)}",
        )



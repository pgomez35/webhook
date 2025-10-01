from fastapi import APIRouter, Body, UploadFile, File
import os
import json
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd

import logging
from pathlib import Path
import uuid

router = APIRouter()


import re
from pathlib import Path

# --- Parser del archivo de texto (respeta los valores tal cual) ---
def _keep(s: str) -> str:
    return (s or "").strip()

import re
from pathlib import Path

def _keep(s: str) -> str:
    """Conserva el texto tal cual, recortando solo espacios al inicio/fin."""
    return (s or "").strip()

def parsear_bloques_desde_txt(ruta_txt: str | Path) -> dict:
    texto = Path(ruta_txt).read_text(encoding="utf-8", errors="ignore")
    lineas = [ln.rstrip("\n\r") for ln in texto.splitlines()]
    i, n = 0, len(lineas)
    out = {}
    _cnt_total = _cnt_con_fecha = _cnt_con_canal = 0
    _ejemplos = []

    while i < n:
        # Posible inicio de bloque: línea con el usuario
        linea = lineas[i].strip()
        i += 1
        if not linea:
            continue

        usuario = linea  # <-- aquí empieza el bloque del usuario

        # "Nombre" o "Name"
        while i < n and lineas[i].strip() == "":
            i += 1
        if i < n and lineas[i].strip().lower() in ("nombre", "name"):
            i += 1

        # Nombre real
        while i < n and lineas[i].strip() == "":
            i += 1
        nombre = _keep(lineas[i]) if i < n else ""
        i += 1

        # Métricas por tabs: seguidores, videos, likes, duración, días
        while i < n and lineas[i].strip() == "":
            i += 1
        metrica_line = lineas[i].strip() if i < n else ""
        i += 1

        m = re.match(r"([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)", metrica_line)
        if m:
            s_seg, s_vid, s_like, s_dur, s_dias = m.groups()
        else:
            partes = metrica_line.split("\t") if metrica_line else []
            while len(partes) < 5:
                partes += [""]
            s_seg, s_vid, s_like, s_dur, s_dias = partes[:5]

        seguidores = _keep(s_seg)    # texto tal cual
        videos     = _keep(s_vid)    # texto tal cual
        likes      = _keep(s_like)   # texto tal cual (ej. "47 237")
        duracion   = _keep(s_dur)
        dias_validos = _keep(s_dias)

        # Caducidad
        while i < n and lineas[i].strip() == "":
            i += 1
        caducidad = _keep(lineas[i]) if i < n else ""
        i += 1

        # Email agente recluta
        while i < n and lineas[i].strip() == "":
            i += 1
        agente = _keep(lineas[i]) if i < n else ""
        i += 1

        # Etiqueta de canal (LIVE / Código QR / QR code) + fecha
        while i < n and lineas[i].strip() == "":
            i += 1
        fecha_sol = ""
        canal = ""
        if i < n:
            ult = lineas[i].strip()
            i += 1

            # Caso típico: "<CANAL>\t<FECHA>"
            partes = ult.split("\t")
            if len(partes) >= 2:
                canal = _keep(partes[0])  # conserva exactamente lo que venga (LIVE, Código QR, QR code)
                fecha_sol = _keep(partes[1])
            else:
                # Intento flexible: extraer fecha dd/mm/yyyy hh:mm:ss en cualquier parte de la línea
                mm = re.search(r"\b\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}\b", ult)
                if mm:
                    fecha_sol = mm.group(0)
                # Extraer canal si la línea inicia con LIVE / Código QR / QR code (u otra etiqueta)
                # Guardamos el token inicial, que suele ser el canal.
                canal_match = re.match(r"^\s*([A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+)", ult)
                if canal_match:
                    canal = _keep(canal_match.group(1))

        out[usuario] = {
            "usuario": usuario,
            "nombre": nombre,
            "seguidores": seguidores,
            "videos": videos,
            "likes": likes,
            "Duracion_Emisiones": duracion,     # se mantiene la misma clave que usas en tu código
            "Dias_Emisiones": dias_validos,     # idem
            "caducidad_solicitud": caducidad,
            "agente_recluta": agente,
            "fecha_solicitud": fecha_sol,        # se respeta la grafía pedida
            "canal": canal,                     # NUEVO: LIVE / Código QR / QR code (texto tal cual)
        }

        _cnt_total += 1
        if canal: _cnt_con_canal += 1
        if fecha_sol:
            _cnt_con_fecha += 1
            if len(_ejemplos) < 3:
                _ejemplos.append((usuario, canal, fecha_sol))

        # Saltar blancos entre bloques
        while i < n and lineas[i].strip() == "":
            i += 1

    # log final
    logger.info(f"TXT parseado: total={_cnt_total}, con_canal={_cnt_con_canal}, con_fecha={_cnt_con_fecha}, ejemplos={_ejemplos}")
    return out


def get_gspread_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    cred_json = os.getenv("GOOGLE_CREDENTIALS_JSON_VIEJA")
    cred_dict = json.loads(cred_json)
    creds = Credentials.from_service_account_info(cred_dict, scopes=scope)
    return gspread.authorize(creds)


logger = logging.getLogger("uvicorn.error")

@router.post("/upload_txt")  # ruta será /api/upload_txt
async def upload_txt(file: UploadFile = File(...)):
    filename = file.filename or "datos.txt"
    if not filename.lower().endswith(".txt"):
        return {"status": "error", "mensaje": "El archivo debe ser .txt"}

    base_dir = Path("uploads/txt")
    base_dir.mkdir(parents=True, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}_{filename}"
    destino = base_dir / unique_name

    contenido = await file.read()
    destino.write_bytes(contenido)

    return {"status": "ok", "ruta_txt": str(destino)}


@router.get("/listar_hojas")
def listar_hojas(str_key: str):
    """
    Devuelve la lista de hojas disponibles en el documento Google Sheets.
    """
    try:
        gc = get_gspread_client()
        spreadsheet = gc.open_by_key(str_key)
        hojas = [ws.title for ws in spreadsheet.worksheets()]
        return {"status": "ok", "hojas": hojas}
    except Exception as e:
        logger.error(f"❌ Error en listar_hojas con str_key={str_key}: {e}", exc_info=True)
        return {"status": "error", "mensaje": f"Error al listar hojas: {str(e)}"}

@router.post("/cargar_aspirantes")
def cargar_aspirantes_desde_workspace(
    nombre_hoja: str = Body(..., embed=True),
    str_key: str = Body(..., embed=True),
    ruta_txt: str = Body(..., embed=True)
):
    try:
        logger.info(f"📥 Iniciando carga de aspirantes: hoja={nombre_hoja}, str_key={str_key}")
        aspirantes = obtener_aspirantes_desde_hoja(str_key, nombre_hoja, ruta_txt)
        if not aspirantes:
            logger.warning(f"⚠️ No se encontraron aspirantes en la hoja {nombre_hoja}")
            return {
                "status": "error",
                "mensaje": "No se encontraron aspirantes en la hoja"
            }
        guardar_aspirantes(aspirantes)
        logger.info(f"✅ {len(aspirantes)} aspirantes cargados y guardados correctamente")
        return {
            "status": "ok",
            "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"
        }
    except Exception as e:
        logger.error(
            f"❌ Error en cargar_aspirantes con hoja={nombre_hoja}, str_key={str_key}: {e}",
            exc_info=True
        )
        return {
            "status": "error",
            "mensaje": f"Error al cargar aspirantes: {str(e)}"
        }


# @router.post("/cargar_aspirantes_local")
# async def cargar_aspirantes_desde_archivo(file: UploadFile):
#     """
#     Carga aspirantes desde un archivo Excel local (.xlsx).
#     El usuario debe subir el archivo desde el frontend.
#     """
#     try:
#         filepath = f"/tmp/{file.filename}"
#         with open(filepath, "wb") as f:
#             f.write(await file.read())
#
#         df = pd.read_excel(filepath)
#         aspirantes = df.to_dict(orient="records")
#         guardar_aspirantes(aspirantes)
#         return {"status": "ok", "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"}
#     except Exception as e:
#         return {"status": "error", "mensaje": f"Error al cargar archivo: {str(e)}"}

from fastapi import UploadFile, File, Form
from openpyxl import load_workbook
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

from fastapi import UploadFile, File, Form
from pathlib import Path
from openpyxl import load_workbook
import uuid, os, logging

logger = logging.getLogger(__name__)

@router.post("/cargar_aspirantes_local")
async def cargar_aspirantes_desde_archivo(
    file: UploadFile = File(...),
    ruta_txt: str = Form(...),
    nombre_hoja: str = Form(None),
):
    """
    Carga aspirantes desde un archivo Excel local (.xlsx/.xlsm) subido por el usuario,
    mezclando datos con el TXT (parseado por usuario).
    Estructura esperada (igual a workspace):
      - Encabezados en fila 3
      - Datos desde fila 4
      - Usuarios en columna B
      - Rango A..X (24 columnas)
    """
    xlsx_path = None
    wb = None
    try:
        # 1) Guardar el Excel subido en /tmp (Render permite escribir allí)
        fname = (file.filename or "").lower()
        if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
            return {"status": "error", "mensaje": "El archivo debe ser .xlsx o .xlsm"}

        tmp_dir = Path("/tmp")
        tmp_dir.mkdir(parents=True, exist_ok=True)
        safe_name = f"local_{uuid.uuid4().hex}_{os.path.basename(file.filename)}"
        xlsx_path = tmp_dir / safe_name
        xlsx_path.write_bytes(await file.read())

        # 2) Parsear TXT (igual que workspace)
        info_por_usuario = parsear_bloques_desde_txt(ruta_txt)
        logger.info(f"✅ TXT parseado, usuarios encontrados: {len(info_por_usuario)}")

        # 3) Abrir Excel y seleccionar hoja
        wb = load_workbook(filename=str(xlsx_path), data_only=True)

        if nombre_hoja:
            if nombre_hoja not in wb.sheetnames:
                return {"status": "error", "mensaje": f"La hoja '{nombre_hoja}' no existe en el Excel"}
            ws = wb[nombre_hoja]
        else:
            # Fuerza la primera hoja del libro (más determinista que wb.active)
            ws = wb[wb.sheetnames[0]]
        logger.info(f"📄 Hoja usada (local): {ws.title}")

        # 4) Detectar última fila usando Columna B (usuarios)
        #    Fila 3 = encabezados; datos desde fila 4
        #    Permitimos huecos: cortamos tras 'consecutive_empties_limit' vacíos seguidos.
        col_B = 2
        start_row = 4
        consecutive_empties = 0
        consecutive_empties_limit = 8  # si hay 8 filas seguidas vacías en B, asumimos fin
        ultima_fila = start_row - 1

        r = start_row
        while True:
            val = ws.cell(row=r, column=col_B).value
            s = ("" if val is None else str(val).strip())
            if s == "":
                consecutive_empties += 1
                if consecutive_empties >= consecutive_empties_limit:
                    break
            else:
                consecutive_empties = 0
                ultima_fila = r
            r += 1

        if ultima_fila < start_row:
            logger.warning("⚠️ No se encontraron usuarios en la columna B (fila 4 en adelante)")
            return {"status": "error", "mensaje": "No se encontraron aspirantes en el archivo"}

        logger.info(f"📊 Rango calculado (local): A{start_row}:X{ultima_fila}")

        # 5) Leer filas A..X y construir aspirantes (igual mapeo que workspace)
        aspirantes = []
        for i in range(start_row, ultima_fila + 1):
            # Normalizar 24 columnas A..X (1..24)
            fila_vals = []
            for c in range(1, 24 + 1):
                v = ws.cell(row=i, column=c).value
                fila_vals.append("" if v is None else str(v).strip())

            usuario = fila_vals[1]  # Columna B (índice 1)
            if not usuario:
                continue

            aspirante = {
                "usuario": usuario,
                "telefono": fila_vals[2].replace(" ", "").replace("+", ""),  # C
                "disponibilidad": fila_vals[3],                               # D
                "motivo_no_apto": fila_vals[4].upper(),                       # E
                "perfil": fila_vals[5],                                       # F
                "contacto": fila_vals[8],                                     # I
                "respuesta_creador": fila_vals[9],                            # J
                "entrevista": fila_vals[11],                                  # L
                "tipo_solicitud": fila_vals[15],                              # P
                "email": fila_vals[16],                                       # Q
                "nickname": fila_vals[17],                                    # R
                "razon_no_contacto": fila_vals[18].upper(),                   # S

                # Métricas vienen del TXT (no del Excel)
                "seguidores": "",
                "videos": "",
                "likes": "",
                "Duracion_Emisiones": "",
                "Dias_Emisiones": "",

                "fila_excel": i,  # fila real en el archivo
            }

            # 6) Mezclar con datos del TXT por usuario
            dtx = info_por_usuario.get(usuario, {})
            if dtx:
                aspirante["seguidores"] = dtx.get("seguidores", "")
                aspirante["videos"] = dtx.get("videos", "")
                aspirante["likes"] = dtx.get("likes", "")
                aspirante["Duracion_Emisiones"] = dtx.get("Duracion_Emisiones", "")
                aspirante["Dias_Emisiones"] = dtx.get("Dias_Emisiones", "")
                aspirante["nombre"] = dtx.get("nombre", "")
                aspirante["caducidad_solicitud"] = dtx.get("caducidad_solicitud", "")
                aspirante["agente_recluta"] = dtx.get("agente_recluta", "")
                aspirante["fecha_solicitud"] = dtx.get("fecha_solicitud", "")
                aspirante["canal"] = dtx.get("canal", "")

            aspirantes.append(aspirante)

        if not aspirantes:
            return {"status": "error", "mensaje": "No se construyeron aspirantes desde el archivo"}

        # 7) Guardar en DB
        guardar_aspirantes(aspirantes, nombre_archivo=xlsx_path.name, hoja_excel=ws.title)
        logger.info(f"✅ {len(aspirantes)} aspirantes (local) cargados y guardados correctamente")
        return {
            "status": "ok",
            "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"
        }

    except Exception as e:
        logger.error(f"❌ Error en cargar_aspirantes_local: {e}", exc_info=True)
        return {"status": "error", "mensaje": f"Error al cargar archivo local: {str(e)}"}

    finally:
        # Limpieza de recursos
        try:
            if wb:
                wb.close()
        except Exception:
            pass
        try:
            if xlsx_path and Path(xlsx_path).exists():
                os.remove(xlsx_path)
        except Exception:
            pass

# @router.post("/cargar_aspirantes_local")
# async def cargar_aspirantes_desde_archivo(
#     file: UploadFile = File(...),
#     ruta_txt: str = Form(...),
#     nombre_hoja: str = Form(None),
# ):
#     """
#     Carga aspirantes desde un archivo Excel local (.xlsx) subido por el usuario,
#     mezclando datos con el TXT (parseado por usuario).
#     Estructura de columnas igual a la de Google Sheets:
#       - Encabezados en fila 3
#       - Datos desde fila 4
#       - Usuarios en columna B
#       - Rangos A..X (24 columnas)
#     """
#     try:
#         # 1) Guardar el Excel subido
#         if not file.filename.lower().endswith(".xlsx"):
#             return {"status": "error", "mensaje": "El archivo debe ser .xlsx"}
#
#         tmp_dir = Path("/tmp")
#         tmp_dir.mkdir(parents=True, exist_ok=True)
#         xlsx_path = tmp_dir / f"local_{file.filename}"
#         xlsx_path.write_bytes(await file.read())
#
#         # 2) Parsear TXT (igual que workspace)
#         info_por_usuario = parsear_bloques_desde_txt(ruta_txt)
#         logger.info(f"✅ TXT parseado, usuarios encontrados: {len(info_por_usuario)}")
#
#         # 3) Abrir Excel y seleccionar hoja
#         wb = load_workbook(filename=str(xlsx_path), data_only=True)
#         ws = wb[nombre_hoja] if nombre_hoja and nombre_hoja in wb.sheetnames else wb.active
#         logger.info(f"📄 Hoja usada (local): {ws.title}")
#
#         # 4) Detectar última fila usando Columna B (usuarios)
#         #    Encabezados en fila 3; datos desde fila 4
#         #    Tomamos B4 hacia abajo hasta la última con valor (trim != "")
#         col_B_desde_4 = []
#         row_idx = 4
#         while True:
#             cell_val = ws.cell(row=row_idx, column=2).value  # B = 2
#             if cell_val is None or str(cell_val).strip() == "":
#                 # detenemos cuando encontramos primera vacía consecutiva y no hay más datos en B
#                 # (si tu archivo tiene huecos en B con datos en otras columnas, puedes mejorar este corte)
#                 break
#             col_B_desde_4.append(str(cell_val).strip())
#             row_idx += 1
#
#         if not col_B_desde_4:
#             logger.warning("⚠️ No se encontraron usuarios en la columna B (fila 4 en adelante)")
#             return {"status": "error", "mensaje": "No se encontraron aspirantes en el archivo"}
#
#         ultima_fila = 3 + len(col_B_desde_4)  # fila 3 + cantidad de filas con usuario
#         logger.info(f"📊 Rango calculado (local): A4:X{ultima_fila}")
#
#         # 5) Leer filas A..X y construir aspirantes (igual mapeo que workspace)
#         aspirantes = []
#         for i in range(4, ultima_fila + 1):
#             # Normalizar 24 columnas A..X (1..24)
#             fila_vals = []
#             for c in range(1, 24 + 1):
#                 v = ws.cell(row=i, column=c).value
#                 fila_vals.append("" if v is None else str(v).strip())
#
#             # Usuario en columna B (idx 1)
#             usuario = fila_vals[1]
#             if not usuario:
#                 continue
#
#             aspirante = {
#                 "usuario": usuario,
#                 "telefono": fila_vals[2].replace(" ", "").replace("+", ""),  # C
#                 "disponibilidad": fila_vals[3],                               # D
#                 "motivo_no_apto": fila_vals[4].upper(),                       # E
#                 "perfil": fila_vals[5],                                       # F
#                 "contacto": fila_vals[8],                                     # I
#                 "respuesta_creador": fila_vals[9],                            # J
#                 "entrevista": fila_vals[11],                                  # L
#                 "tipo_solicitud": fila_vals[15],                              # P
#                 "email": fila_vals[16],                                       # Q
#                 "nickname": fila_vals[17],                                    # R
#                 "razon_no_contacto": fila_vals[18].upper(),                   # S
#
#                 # Métricas vienen del TXT (no del Excel)
#                 "seguidores": "",
#                 "videos": "",
#                 "likes": "",
#                 "Duracion_Emisiones": "",
#                 "Dias_Emisiones": "",
#
#                 "fila_excel": i,  # fila real en el archivo (coincide con Google Sheets)
#             }
#
#             # 6) Mezclar con datos del TXT por usuario
#             dtx = info_por_usuario.get(usuario, {})
#             if dtx:
#                 aspirante["seguidores"] = dtx.get("seguidores", "")
#                 aspirante["videos"] = dtx.get("videos", "")
#                 aspirante["likes"] = dtx.get("likes", "")
#                 aspirante["Duracion_Emisiones"] = dtx.get("Duracion_Emisiones", "")
#                 aspirante["Dias_Emisiones"] = dtx.get("Dias_Emisiones", "")
#                 aspirante["nombre"] = dtx.get("nombre", "")
#                 aspirante["caducidad_solicitud"] = dtx.get("caducidad_solicitud", "")
#                 aspirante["agente_recluta"] = dtx.get("agente_recluta", "")
#                 aspirante["fecha_solicitud"] = dtx.get("fecha_solcitud", "") or dtx.get("fecha_solicitud", "")
#                 aspirante["canal"] = dtx.get("canal", "")
#
#             aspirantes.append(aspirante)
#
#         if not aspirantes:
#             return {"status": "error", "mensaje": "No se construyeron aspirantes desde el archivo"}
#
#         # 7) Guardar en DB
#         guardar_aspirantes(aspirantes, nombre_archivo=xlsx_path.name, hoja_excel=ws.title)
#         logger.info(f"✅ {len(aspirantes)} aspirantes (local) cargados y guardados correctamente")
#         return {
#             "status": "ok",
#             "mensaje": f"{len(aspirantes)} aspirantes cargados y guardados correctamente"
#         }
#
#     except Exception as e:
#         logger.error(f"❌ Error en cargar_aspirantes_local: {e}", exc_info=True)
#         return {"status": "error", "mensaje": f"Error al cargar archivo local: {str(e)}"}

from DataBase import get_connection,limpiar_telefono,safe_int


import re

def get_text(val):
    """Texto limpio (str) o cadena vacía si viene None."""
    return "" if val is None else str(val).strip()

def to_int_relaxed(val, default=0):
    """
    Convierte '1,151' o '47 237' -> 1151 / 47237.
    - Elimina espacios y comas.
    - Si no se puede convertir, devuelve 'default' (por defecto 0).
    """
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    s = s.replace(" ", "").replace(",", "")
    return int(s) if s.isdigit() else default

def parse_days_to_int(val, default=0):
    """
    Convierte '24d' -> 24, '1d' -> 1, '0d' -> 0, '15' -> 15.
    Si no se puede convertir, devuelve 'default' (por defecto 0).
    """
    if val is None:
        return default
    s = str(val).strip().lower()
    if not s:
        return default
    if s.endswith("d"):
        s = s[:-1]
    try:
        return int(s)
    except ValueError:
        # último intento: limpiar como número relajado
        return to_int_relaxed(s, default=default)

def parse_duration_to_hours_int(text, default=0):

    if text is None:
        return default
    s = str(text).strip().lower()
    if not s:
        return default

    # Unificar 'min' -> 'm'
    s = s.replace("min", "m")

    days = hours = mins = secs = 0
    md = re.search(r"(\d+)\s*d", s)
    mh = re.search(r"(\d+)\s*h", s)
    mm = re.search(r"(\d+)\s*m(?![a-z])", s)  # evita 'ms'
    ms = re.search(r"(\d+)\s*s", s)

    if md: days = int(md.group(1))
    if mh: hours = int(mh.group(1))
    if mm: mins = int(mm.group(1))
    if ms: secs = int(ms.group(1))

    # Si no hay ningún match pero es número simple, interpretarlo como horas
    if not (md or mh or mm or ms):
        try:
            return int(round(float(s)))
        except ValueError:
            return default

    total_hours = days * 24 + hours + mins / 60.0 + secs / 3600.0
    return int(round(total_hours))


# -------------------- guardar_aspirantes (ajustada) --------------------
def obtener_aspirantes_desde_hoja(str_key, nombre_hoja, ruta_txt):
    try:
        # 1) Parseo TXT (como ya lo tienes)
        info_por_usuario = parsear_bloques_desde_txt(ruta_txt)

        # 2) Conectar y abrir hoja
        gc = get_gspread_client()
        spreadsheet = gc.open_by_key(str_key)
        worksheet = spreadsheet.worksheet(nombre_hoja)

        # ---- Cálculo de última fila basado en Columna B (usuarios) ----
        # B4 en adelante son datos; B3 es encabezado.
        col_B_desde_4 = worksheet.col_values(2)[3:]  # índice base 0 => [3:] salta 3 filas (hasta B3)
        # Filas no vacías (trim)
        col_B_no_vacias = [c for c in col_B_desde_4 if str(c).strip() != ""]
        ultima_fila = 3 + len(col_B_no_vacias)  # fila 3 + cantidad de filas con usuario

        # Si por algún motivo viene un hueco en B pero hay otras columnas con datos,
        # usamos un fallback: leer A4:X(3 + len(col_B_desde_4)) y luego filtrar vacíos.
        rango = f"A4:X{ultima_fila}"
        filas = worksheet.get(rango)

        # Fallback si el rango vino vacío pero había datos en B
        if not filas and col_B_desde_4:
            rango_fallback = f"A4:X{3 + len(col_B_desde_4)}"
            filas = worksheet.get(rango_fallback)

        aspirantes = []
        for i, fila in enumerate(filas):
            # normaliza ancho a 24 columnas (A..X)
            if len(fila) < 24:
                fila += [''] * (24 - len(fila))

            usuario = fila[1].strip()  # Columna B
            if not usuario:
                # si la fila no tiene usuario, saltamos
                continue

            aspirante = {
                "usuario": usuario,
                "telefono": fila[2].strip().replace(" ", "").replace("+", ""),  # C
                "disponibilidad": fila[3].strip(),                              # D
                "motivo_no_apto": fila[4].strip().upper(),                     # E
                "perfil": fila[5].strip(),                                     # F
                "contacto": fila[8].strip(),                                   # I
                "respuesta_creador": fila[9].strip(),                          # J
                "entrevista": fila[11].strip(),                                 # L
                "tipo_solicitud": fila[15].strip(),                             # P
                "email": fila[16].strip(),                                      # Q
                "nickname": fila[17].strip(),                                   # R
                "razon_no_contacto": fila[18].strip().upper(),                  # S
                # Métricas vienen del TXT (no del sheet)
                "seguidores": "",
                "videos": "",
                "likes": "",
                "Duracion_Emisiones": "",
                "Dias_Emisiones": "",
                "fila_excel": i + 4,  # porque empezamos en la fila 4
            }

            # Mezcla con datos del TXT por usuario
            datos_txt = info_por_usuario.get(usuario, {})
            if datos_txt:
                aspirante["seguidores"] = datos_txt.get("seguidores", "")
                aspirante["videos"] = datos_txt.get("videos", "")
                aspirante["likes"] = datos_txt.get("likes", "")
                aspirante["Duracion_Emisiones"] = datos_txt.get("Duracion_Emisiones", "")
                aspirante["Dias_Emisiones"] = datos_txt.get("Dias_Emisiones", "")
                aspirante["nombre"] = datos_txt.get("nombre", "")
                aspirante["caducidad_solicitud"] = datos_txt.get("caducidad_solicitud", "")
                aspirante["agente_recluta"] = datos_txt.get("agente_recluta", "")
                aspirante["fecha_solicitud"] = datos_txt.get("fecha_solicitud", "")
                aspirante["canal"] = datos_txt.get("canal", "")

            aspirantes.append(aspirante)

        return aspirantes

    except Exception as e:
        print(f"❌ Error leyendo hoja de cálculo: {e}")
        return []
from datetime import datetime

# --- helpers mínimos usados en la función ---
def get_text(v):
    return (v or "").strip()

def to_int_relaxed(v, default=0):
    if v is None:
        return default
    s = str(v).strip()
    # quita espacios internos en números como "47 237"
    s = s.replace(" ", "").replace(",", "")  # por si viene "1,234"
    try:
        return int(s)
    except:
        return default

def parse_duration_to_hours_int(s, default=0):
    """
    Convierte '3d 2h 54m 10s' (o cualquier subconjunto) a HORAS (int, redondeado).
    """
    if not s:
        return default
    txt = str(s).lower().strip()
    # normaliza: m/min, s/sec
    txt = txt.replace("min", "m").replace("mins", "m").replace("sec", "s").replace("secs", "s")
    # tokens por espacios
    parts = txt.split()
    days = hours = minutes = seconds = 0
    for p in parts:
        if p.endswith("d"):
            days = to_int_relaxed(p[:-1], 0)
        elif p.endswith("h"):
            hours = to_int_relaxed(p[:-1], 0)
        elif p.endswith("m"):
            minutes = to_int_relaxed(p[:-1], 0)
        elif p.endswith("s"):
            seconds = to_int_relaxed(p[:-1], 0)
    total_hours = days * 24 + hours + minutes / 60.0 + seconds / 3600.0
    try:
        return int(round(total_hours))
    except:
        return default

def parse_days_to_int(s, default=0):
    if not s:
        return default
    txt = str(s).lower().strip()
    if txt.endswith("d"):
        txt = txt[:-1]
    return to_int_relaxed(txt, default)

def _parse_fecha_solicitud(fecha_str):
    """
    Intenta parsear 'dd/mm/yyyy HH:MM:SS' (o sin segundos) a datetime (naive).
    Retorna None si no se puede.
    """
    if not fecha_str:
        return None
    s = str(fecha_str).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None
# --- fin helpers ---


def guardar_aspirantes(
    aspirantes, nombre_archivo=None, hoja_excel=None, lote_carga=None, procesado_por=None,
    observaciones=None
):
    """
    Guarda aspirantes mapeando los campos del TXT a los tipos de la tabla,
    y ACTUALIZA/INSERTA creadores.fecha_solicitud cuando venga en el TXT.
    """
    conn = get_connection()
    cur = conn.cursor()
    resultados = []
    filas_fallidas = []

    for c in aspirantes:
        try:
            usuario = get_text(c.get("usuario"))
            nickname = get_text(c.get("nickname"))
            email = get_text(c.get("email"))
            telefono = limpiar_telefono(get_text(c.get("telefono")))
            disponibilidad = get_text(c.get("disponibilidad"))
            perfil = get_text(c.get("perfil"))
            motivo_no_apto = get_text(c.get("motivo_no_apto"))
            contacto = get_text(c.get("contacto"))
            respuesta_creador = get_text(c.get("respuesta_creador"))
            entrevista = get_text(c.get("entrevista"))
            tipo_solicitud = get_text(c.get("tipo_solicitud"))
            razon_no_contacto = get_text(c.get("razon_no_contacto"))

            # NUEVO: fecha_solicitud (del TXT parseado)
            # tu parseador guarda c["fecha_solicitud"] o (a veces) "fecha_solcitud"; probamos ambos
            fecha_sol_txt = c.get("fecha_solicitud")
            fecha_solicitud_dt = _parse_fecha_solicitud(fecha_sol_txt)

            if fecha_sol_txt:
                logger.info(f"[{usuario}] fecha_solicitud TXT='{fecha_sol_txt}' -> parsed={fecha_solicitud_dt}")
            else:
                logger.info(f"[{usuario}] SIN fecha_solicitud en el aspirante")

            # === Conversión a tipos que exige la tabla ===
            seguidores = to_int_relaxed(c.get("seguidores"), default=0)
            cantidad_videos = to_int_relaxed(c.get("videos"), default=0)
            likes_totales = to_int_relaxed(c.get("likes"), default=0)
            duracion_emisiones = parse_duration_to_hours_int(c.get("Duracion_Emisiones"), default=0)
            dias_emisiones = parse_days_to_int(c.get("Dias_Emisiones"), default=0)

            fila_excel = c.get("fila_excel")
            apto = not bool(motivo_no_apto)

            # 1) creadores
            cur.execute("SELECT id FROM creadores WHERE usuario = %s", (usuario,))
            creador_row = cur.fetchone()
            if creador_row:
                creador_id = creador_row[0]
                # construimos UPDATE dinámico para no pisar fecha_solicitud con NULL
                set_cols = [
                    "nickname = %s",
                    "email = %s",
                    "telefono = %s",
                    "estado_id = 3",
                    "actualizado_en = NOW()",
                ]
                params = [
                    get_text(c.get("nombre")),  # nickname se actualiza con el "Nombre" del TXT
                    email,
                    telefono,
                ]
                if fecha_solicitud_dt is not None:
                    logger.info(f"[{usuario}] Actualizando/insertando creadores.fecha_solicitud={fecha_solicitud_dt}")

                    set_cols.insert(3, "fecha_solicitud = %s")  # antes de estado_id por orden
                    params.insert(3, fecha_solicitud_dt)

                sql_update = f"""
                    UPDATE creadores SET
                        {", ".join(set_cols)}
                    WHERE id = %s
                """
                params.append(creador_id)
                cur.execute(sql_update, tuple(params))

            else:
                # INSERT incluye fecha_solicitud si viene
                if fecha_solicitud_dt is not None:

                    logger.info(f"[{usuario}] Actualizando/insertando creadores.fecha_solicitud={fecha_solicitud_dt}")

                    cur.execute("""
                        INSERT INTO creadores (usuario, nickname, email, telefono, fecha_solicitud, estado_id, activo, creado_en, actualizado_en)
                        VALUES (%s, %s, %s, %s, %s, 3, TRUE, NOW(), NOW())
                        RETURNING id
                    """, (
                        usuario,
                        get_text(c.get("nombre")),
                        email,
                        telefono,
                        fecha_solicitud_dt
                    ))
                else:
                    cur.execute("""
                        INSERT INTO creadores (usuario, nickname, email, telefono, estado_id, activo, creado_en, actualizado_en)
                        VALUES (%s, %s, %s, %s, 3, TRUE, NOW(), NOW())
                        RETURNING id
                    """, (
                        usuario,
                        get_text(c.get("nombre")),
                        email,
                        telefono
                    ))
                creador_id = cur.fetchone()[0]

            # 2) perfil_creador
            cur.execute("SELECT id FROM perfil_creador WHERE creador_id = %s", (creador_id,))
            perfil_row = cur.fetchone()
            if perfil_row:
                cur.execute("""
                    UPDATE perfil_creador SET
                        usuario = %s,
                        seguidores = %s,
                        videos = %s,
                        likes = %s,
                        duracion_emisiones = %s,
                        dias_emisiones = %s,
                        nombre = %s,
                        actualizado_en = NOW()
                    WHERE creador_id = %s
                """, (
                    usuario,
                    seguidores, cantidad_videos, likes_totales,
                    duracion_emisiones, dias_emisiones,
                    get_text(c.get("nombre")),
                    creador_id
                ))
            else:
                cur.execute("""
                    INSERT INTO perfil_creador (
                        usuario, creador_id,
                        seguidores, videos, likes,
                        duracion_emisiones, dias_emisiones,
                        nombre, creado_en, actualizado_en
                    ) VALUES (
                        %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, NOW(), NOW()
                    )
                """, (
                    usuario, creador_id,
                    seguidores, cantidad_videos, likes_totales,
                    duracion_emisiones, dias_emisiones,
                    get_text(c.get("nombre"))
                ))

            # 3) cargue_creadores
            cur.execute("SELECT id FROM cargue_creadores WHERE usuario = %s AND hoja_excel = %s", (usuario, hoja_excel))
            cargue_row = cur.fetchone()
            if cargue_row:
                cargue_id = cargue_row[0]
                cur.execute("""
                    UPDATE cargue_creadores SET
                        nickname = %s,
                        email = %s,
                        telefono = %s,
                        disponibilidad = %s,
                        perfil = %s,
                        motivo_no_apto = %s,
                        contacto = %s,
                        respuesta_creador = %s,
                        entrevista = %s,
                        tipo_solicitud = %s,
                        razon_no_contacto = %s,
                        seguidores = %s,
                        cantidad_videos = %s,
                        likes_totales = %s,
                        duracion_emisiones = %s,
                        dias_emisiones = %s,
                        nombre_archivo = %s,
                        fila_excel = %s,
                        lote_carga = %s,
                        estado = %s,
                        procesado = %s,
                        procesado_por = %s,
                        creador_id = %s,
                        apto = %s,
                        observaciones = %s,
                        actualizado_en = NOW()
                    WHERE id = %s
                """, (
                    nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
                    contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
                    seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
                    nombre_archivo, fila_excel, lote_carga, "Procesando", False, procesado_por,
                    creador_id, apto, observaciones, cargue_id
                ))
            else:
                cur.execute("""
                    INSERT INTO cargue_creadores (
                        usuario, nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
                        contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
                        seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
                        nombre_archivo, hoja_excel, fila_excel, lote_carga,
                        estado, procesado, procesado_por, creador_id,
                        apto, observaciones, activo, creado_en, actualizado_en
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, TRUE, NOW(), NOW()
                    )
                """, (
                    usuario, nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
                    contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
                    seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
                    nombre_archivo, hoja_excel, fila_excel, lote_carga,
                    "Procesando", False, procesado_por, creador_id,
                    apto, observaciones
                ))

            resultados.append({
                "fila": fila_excel,
                "usuario": usuario,
                "creador_id": creador_id
            })

        except Exception as e:
            logger.error(
                f"Error al guardar aspirante (fila={c.get('fila_excel')}, usuario={c.get('usuario')}): {e}",
                exc_info=True
            )
            filas_fallidas.append({
                "fila": c.get("fila_excel"),
                "usuario": c.get("usuario"),
                "error": str(e),
            })
            # continúa con el siguiente

    conn.commit()
    cur.close()
    conn.close()

    print(f"✅ Contactos procesados. Filas exitosas: {len(resultados)}. Filas fallidas: {len(filas_fallidas)}")
    return {
        "status": "ok",
        "exitosos": resultados,
        "fallidos": filas_fallidas
    }


# def guardar_aspirantes(
#     aspirantes, nombre_archivo=None, hoja_excel=None, lote_carga=None, procesado_por=None,
#     observaciones=None
# ):
#     """
#     Guarda aspirantes mapeando los campos del TXT a los tipos de la tabla:
#
#       perfil_creador.seguidores            -> integer
#       perfil_creador.videos                -> integer
#       perfil_creador.likes                 -> bigint
#       perfil_creador.duracion_emisiones    -> integer (HORAS redondeadas)
#       perfil_creador.dias_emisiones        -> integer (quita 'd')
#
#     NOTA: Se asume que get_connection(), limpiar_telefono() y logger existen en tu entorno.
#     """
#     conn = get_connection()
#     cur = conn.cursor()
#     resultados = []
#     filas_fallidas = []
#
#     for c in aspirantes:
#         try:
#             usuario = get_text(c.get("usuario"))
#             nickname = get_text(c.get("nickname"))
#             email = get_text(c.get("email"))
#             telefono = limpiar_telefono(get_text(c.get("telefono")))
#             disponibilidad = get_text(c.get("disponibilidad"))
#             perfil = get_text(c.get("perfil"))
#             motivo_no_apto = get_text(c.get("motivo_no_apto"))
#             contacto = get_text(c.get("contacto"))
#             respuesta_creador = get_text(c.get("respuesta_creador"))
#             entrevista = get_text(c.get("entrevista"))
#             tipo_solicitud = get_text(c.get("tipo_solicitud"))
#             razon_no_contacto = get_text(c.get("razon_no_contacto"))
#
#             # === Conversión a tipos que exige la tabla ===
#             seguidores = to_int_relaxed(c.get("seguidores"), default=0)          # integer
#             cantidad_videos = to_int_relaxed(c.get("videos"), default=0)         # integer
#             likes_totales = to_int_relaxed(c.get("likes"), default=0)            # bigint ok
#             duracion_emisiones = parse_duration_to_hours_int(c.get("Duracion_Emisiones"), default=0)  # integer (horas)
#             dias_emisiones = parse_days_to_int(c.get("Dias_Emisiones"), default=0)                    # integer (días)
#
#             fila_excel = c.get("fila_excel")
#             apto = not bool(motivo_no_apto)
#
#             # 1) creadores
#             cur.execute("SELECT id FROM creadores WHERE usuario = %s", (usuario,))
#             creador_row = cur.fetchone()
#             if creador_row:
#                 creador_id = creador_row[0]
#                 cur.execute("""
#                     UPDATE creadores SET
#                         nickname = %s,
#                         email = %s,
#                         telefono = %s,
#                         estado_id = 3,
#                         actualizado_en = NOW()
#                     WHERE id = %s
#                 """, (
#                     get_text(c.get("nombre")),  # ⬅️ usamos el mismo nombre que en perfil_creador.nombre
#                     email,
#                     telefono,
#                     creador_id
#                 ))
#             else:
#                 cur.execute("""
#                     INSERT INTO creadores (usuario, nickname, email, telefono, estado_id, activo, creado_en, actualizado_en)
#                     VALUES (%s, %s, %s, %s, 3, TRUE, NOW(), NOW())
#                     RETURNING id
#                 """, (
#                     usuario,
#                     get_text(c.get("nombre")),  # ⬅️ nickname = nombre del TXT
#                     email,
#                     telefono
#                 ))
#                 creador_id = cur.fetchone()[0]
#
#             # 2) perfil_creador (tipos numéricos según tu esquema)
#             cur.execute("SELECT id FROM perfil_creador WHERE creador_id = %s", (creador_id,))
#             perfil_row = cur.fetchone()
#             if perfil_row:
#                 cur.execute("""
#                     UPDATE perfil_creador SET
#                         usuario = %s,
#                         seguidores = %s,
#                         videos = %s,
#                         likes = %s,
#                         duracion_emisiones = %s,
#                         dias_emisiones = %s,
#                         nombre = %s,
#                         actualizado_en = NOW()
#                     WHERE creador_id = %s
#                 """, (
#                     usuario,
#                     seguidores, cantidad_videos, likes_totales,
#                     duracion_emisiones, dias_emisiones,
#                     get_text(c.get("nombre")),  # opcional: guardar el "Nombre" del TXT
#                     creador_id
#                 ))
#             else:
#                 cur.execute("""
#                     INSERT INTO perfil_creador (
#                         usuario, creador_id,
#                         seguidores, videos, likes,
#                         duracion_emisiones, dias_emisiones,
#                         nombre, creado_en, actualizado_en
#                     ) VALUES (
#                         %s, %s,
#                         %s, %s, %s,
#                         %s, %s,
#                         %s, NOW(), NOW()
#                     )
#                 """, (
#                     usuario, creador_id,
#                     seguidores, cantidad_videos, likes_totales,
#                     duracion_emisiones, dias_emisiones,
#                     get_text(c.get("nombre"))
#                 ))
#
#             # 3) cargue_creadores (si tus columnas allí son texto/mixtas, se mantienen)
#             cur.execute("SELECT id FROM cargue_creadores WHERE usuario = %s AND hoja_excel = %s", (usuario, hoja_excel))
#             cargue_row = cur.fetchone()
#             if cargue_row:
#                 cargue_id = cargue_row[0]
#                 cur.execute("""
#                     UPDATE cargue_creadores SET
#                         nickname = %s,
#                         email = %s,
#                         telefono = %s,
#                         disponibilidad = %s,
#                         perfil = %s,
#                         motivo_no_apto = %s,
#                         contacto = %s,
#                         respuesta_creador = %s,
#                         entrevista = %s,
#                         tipo_solicitud = %s,
#                         razon_no_contacto = %s,
#                         seguidores = %s,
#                         cantidad_videos = %s,
#                         likes_totales = %s,
#                         duracion_emisiones = %s,
#                         dias_emisiones = %s,
#                         nombre_archivo = %s,
#                         fila_excel = %s,
#                         lote_carga = %s,
#                         estado = %s,
#                         procesado = %s,
#                         procesado_por = %s,
#                         creador_id = %s,
#                         apto = %s,
#                         observaciones = %s,
#                         actualizado_en = NOW()
#                     WHERE id = %s
#                 """, (
#                     nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
#                     contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
#                     seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
#                     nombre_archivo, fila_excel, lote_carga, "Procesando", False, procesado_por,
#                     creador_id, apto, observaciones, cargue_id
#                 ))
#             else:
#                 cur.execute("""
#                     INSERT INTO cargue_creadores (
#                         usuario, nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
#                         contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
#                         seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
#                         nombre_archivo, hoja_excel, fila_excel, lote_carga,
#                         estado, procesado, procesado_por, creador_id,
#                         apto, observaciones, activo, creado_en, actualizado_en
#                     ) VALUES (
#                         %s, %s, %s, %s, %s, %s, %s,
#                         %s, %s, %s, %s, %s,
#                         %s, %s, %s, %s, %s,
#                         %s, %s, %s, %s,
#                         %s, %s, %s, %s,
#                         %s, %s, TRUE, NOW(), NOW()
#                     )
#                 """, (
#                     usuario, nickname, email, telefono, disponibilidad, perfil, motivo_no_apto,
#                     contacto, respuesta_creador, entrevista, tipo_solicitud, razon_no_contacto,
#                     seguidores, cantidad_videos, likes_totales, duracion_emisiones, dias_emisiones,
#                     nombre_archivo, hoja_excel, fila_excel, lote_carga,
#                     "Procesando", False, procesado_por, creador_id,
#                     apto, observaciones
#                 ))
#
#             resultados.append({
#                 "fila": fila_excel,
#                 "usuario": usuario,
#                 "creador_id": creador_id
#             })
#
#         except Exception as e:
#             logger.error(
#                 f"Error al guardar aspirante (fila={c.get('fila_excel')}, usuario={c.get('usuario')}): {e}",
#                 exc_info=True
#             )
#             filas_fallidas.append({
#                 "fila": c.get("fila_excel"),
#                 "usuario": c.get("usuario"),
#                 "error": str(e),
#             })
#             # continuar con el siguiente registro
#
#     conn.commit()
#     cur.close()
#     conn.close()
#
#     print(f"✅ Contactos procesados. Filas exitosas: {len(resultados)}. Filas fallidas: {len(filas_fallidas)}")
#     return {
#         "status": "ok",
#         "exitosos": resultados,
#         "fallidos": filas_fallidas
#     }


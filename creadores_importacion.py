"""
Importación masiva de creadores desde Excel, CSV o TXT.
Independiente de reportes semanales / performance.
"""
from __future__ import annotations

import csv
import io
import json
import re
import traceback
import unicodedata
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from psycopg2.extras import Json, RealDictCursor
from psycopg2 import errors as pg_errors
from pydantic import BaseModel, EmailStr, TypeAdapter, ValidationError

from DataBase import get_connection_context
from main_auth import es_manager, obtener_usuario_actual
from creadores_manager import construir_indices_manager, resolver_manager_id_por_indices

router = APIRouter()

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

COLUMNAS_OFICIALES = [
    "nombre",
    "usuario_tiktok",
    "telefono",
    "email",
    "manager",
    "fecha_incorporacion",
    "estado",
]

ALIAS_ENCABEZADOS = {
    "nombre": "nombre",
    "usuario_tiktok": "usuario_tiktok",
    "usuario tiktok": "usuario_tiktok",
    "usuariotiktok": "usuario_tiktok",
    "telefono": "telefono",
    "teléfono": "telefono",
    "email": "email",
    "correo": "email",
    "manager": "manager",
    "fecha_incorporacion": "fecha_incorporacion",
    "fecha incorporacion": "fecha_incorporacion",
    "fecha incorporación": "fecha_incorporacion",
    "estado": "estado",
}

ESTADOS_TEXTO_PERMITIDOS = {
    "activo": "Activo",
    "inactivo": "Inactivo",
}

MAX_FILAS = 5000
MAX_FILE_BYTES = 5 * 1024 * 1024  # 5 MB
EXTENSIONES_PERMITIDAS = {".xlsx", ".xls", ".csv", ".txt"}

TIKTOK_URL_RE = re.compile(r"tiktok\.com/@([^/?#\s]+)", re.IGNORECASE)
EMAIL_ADAPTER = TypeAdapter(EmailStr)

FILA_VALIDA = "valida"
FILA_ADVERTENCIA = "advertencia"
FILA_DUPLICADO = "duplicado"
FILA_ERROR = "error"

SQL_CREAR_TABLA_HISTORIAL = """
CREATE TABLE IF NOT EXISTS creadores_importaciones (
    id SERIAL PRIMARY KEY,
    archivo_nombre VARCHAR(255),
    separador VARCHAR(8),
    total_filas INTEGER,
    validas INTEGER,
    advertencias INTEGER,
    duplicados INTEGER,
    errores INTEGER,
    importables INTEGER,
    creadores_creados INTEGER,
    omitidos INTEGER,
    errores_importacion INTEGER,
    estado VARCHAR(30) DEFAULT 'procesado',
    cargado_por INTEGER,
    fecha_carga TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW(),
    metadata_json JSONB DEFAULT '{}'::jsonb
)
"""

SQL_INDICE_HISTORIAL = """
CREATE INDEX IF NOT EXISTS idx_creadores_importaciones_fecha
    ON creadores_importaciones (fecha_carga DESC)
"""


# ---------------------------------------------------------------------------
# Permisos
# ---------------------------------------------------------------------------


def _require_permiso_importacion(usuario: dict) -> None:
    if not usuario:
        raise HTTPException(status_code=401, detail="No autenticado")
    if es_manager(usuario):
        raise HTTPException(
            status_code=403,
            detail="No tienes permiso para importar creadores",
        )


# ---------------------------------------------------------------------------
# Utilidades de normalización
# ---------------------------------------------------------------------------


def _normalizar_encabezado(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace(" ", "_")
    return ALIAS_ENCABEZADOS.get(s, s)


def _celda_texto(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _colapsar_espacios(texto: str) -> str:
    return re.sub(r"\s+", " ", texto).strip()


def _es_solo_simbolos(texto: str) -> bool:
    return not any(ch.isalnum() for ch in texto)


def normalizar_usuario_tiktok(raw: Any) -> Tuple[Optional[str], List[str]]:
    mensajes: List[str] = []
    if raw is None or _celda_texto(raw) == "":
        return None, ["Usuario TikTok vacío"]

    s = _celda_texto(raw)
    url_match = TIKTOK_URL_RE.search(s)
    if url_match:
        s = url_match.group(1)
    elif "tiktok.com" in s.lower():
        return None, ["URL de TikTok no reconocida"]

    s = s.lstrip("@").strip().lower()
    if " " in s:
        return None, ["El usuario TikTok no puede contener espacios"]
    if not s:
        return None, ["Usuario TikTok vacío"]
    if len(s) > 50:
        return None, ["Usuario TikTok supera 50 caracteres"]
    return s, mensajes


def normalizar_nombre(raw: Any) -> Tuple[Optional[str], List[str]]:
    if raw is None or _celda_texto(raw) == "":
        return None, ["Nombre vacío"]
    nombre = _colapsar_espacios(_celda_texto(raw))
    if not nombre:
        return None, ["Nombre vacío"]
    if _es_solo_simbolos(nombre):
        return None, ["Nombre inválido"]
    if len(nombre) > 100:
        return None, ["Nombre supera 100 caracteres"]
    return nombre, []


def normalizar_telefono(raw: Any) -> Tuple[Optional[str], List[str]]:
    if raw is None or _celda_texto(raw) == "":
        return None, []
    tel = _celda_texto(raw)
    tel = re.sub(r"[\s\-().]", "", tel)
    if not tel:
        return None, []
    if len(tel) > 50:
        return None, ["Teléfono supera 50 caracteres"]
    digitos = re.sub(r"\D", "", tel)
    if len(digitos) < 7:
        return tel, ["Teléfono con formato dudoso"]
    return tel, []


def normalizar_email(raw: Any) -> Tuple[Optional[str], List[str]]:
    if raw is None or _celda_texto(raw) == "":
        return None, []
    email = _celda_texto(raw).lower()
    try:
        EMAIL_ADAPTER.validate_python(email)
    except ValidationError:
        return None, ["Email con formato inválido"]
    if len(email) > 200:
        return None, ["Email supera 200 caracteres"]
    return email, []


def normalizar_fecha(raw: Any) -> Tuple[Optional[date], List[str]]:
    if raw is None or _celda_texto(raw) == "":
        return None, []

    if isinstance(raw, datetime):
        return raw.date(), []
    if isinstance(raw, date):
        return raw, []

    texto = _celda_texto(raw)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", texto):
        try:
            return datetime.strptime(texto, "%Y-%m-%d").date(), []
        except ValueError:
            return None, ["Fecha inválida; use YYYY-MM-DD"]
    return None, ["Fecha con formato ambiguo; use YYYY-MM-DD"]


def normalizar_estado_texto(raw: Any) -> Tuple[Optional[str], List[str]]:
    if raw is None or _celda_texto(raw) == "":
        return None, ["Estado vacío"]
    clave = _celda_texto(raw).lower()
    if clave not in ESTADOS_TEXTO_PERMITIDOS:
        return None, [f'Estado "{_celda_texto(raw)}" no válido; use Activo o Inactivo']
    return ESTADOS_TEXTO_PERMITIDOS[clave], []


def _sql_usuario_normalizado(campo: str = "usuario_tiktok") -> str:
    return f"LOWER(TRIM(BOTH '@' FROM TRIM({campo})))"


# ---------------------------------------------------------------------------
# Lectura de archivos
# ---------------------------------------------------------------------------


def _validar_archivo(file: UploadFile, content: bytes) -> str:
    nombre = (file.filename or "").lower()
    if not nombre:
        raise HTTPException(status_code=400, detail="Nombre de archivo requerido")
    ext = "." + nombre.rsplit(".", 1)[-1] if "." in nombre else ""
    if ext not in EXTENSIONES_PERMITIDAS:
        raise HTTPException(
            status_code=400,
            detail="Tipo de archivo no permitido. Use .xlsx, .csv o .txt",
        )
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo supera el límite de {MAX_FILE_BYTES // (1024 * 1024)} MB",
        )
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    return ext


def _detectar_separador(muestra: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(muestra, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        for sep in (";", ",", "\t"):
            if sep in muestra:
                return sep
        return ","


def _leer_dataframe(
    content: bytes,
    ext: str,
    separador: Optional[str] = None,
) -> pd.DataFrame:
    buffer = io.BytesIO(content)

    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(buffer, sheet_name=0)
    else:
        texto = content.decode("utf-8-sig", errors="replace")
        if not texto.strip():
            raise HTTPException(status_code=400, detail="El archivo está vacío")
        sep = separador or _detectar_separador(texto[:4096])
        df = pd.read_csv(io.StringIO(texto), sep=sep, dtype=str, keep_default_na=False)

    if df.empty:
        raise HTTPException(status_code=400, detail="El archivo no contiene filas de datos")

    # Normalizar encabezados
    nuevas_cols: Dict[str, str] = {}
    for col in df.columns:
        canon = _normalizar_encabezado(col)
        nuevas_cols[str(col)] = canon
    df = df.rename(columns=nuevas_cols)

    # Quitar filas completamente vacías
    df = df.dropna(how="all")
    df = df.fillna("")

    if len(df) > MAX_FILAS:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo supera el límite de {MAX_FILAS} filas",
        )
    return df


def _validar_encabezados(df: pd.DataFrame) -> Tuple[List[str], List[str]]:
    columnas_presentes = list(df.columns)
    faltantes = [c for c in COLUMNAS_OFICIALES if c not in columnas_presentes]
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}",
        )

    advertencias_globales: List[str] = []
    extras = [c for c in columnas_presentes if c not in COLUMNAS_OFICIALES]
    for col in extras:
        advertencias_globales.append(f'La columna "{col}" no será importada')
    return columnas_presentes, advertencias_globales


# ---------------------------------------------------------------------------
# Catálogos tenant (estados, managers)
# ---------------------------------------------------------------------------


def _cargar_estados_map(cur) -> Dict[str, int]:
    cur.execute(
        """
        SELECT id, nombre
        FROM creadores_estados
        WHERE COALESCE(activo, true) = true
        """
    )
    estados: Dict[str, int] = {}
    for row in cur.fetchall():
        estados[str(row["nombre"]).strip().lower()] = int(row["id"])
    return estados


def _cargar_managers_index(cur) -> Dict[str, Any]:
    cur.execute(
        """
        SELECT
            a.id,
            a.email,
            a.agente,
            a.nombre_completo,
            a.username
        FROM administradores a
        INNER JOIN administradores_roles ur ON ur.id = a.administradores_roles_id
        WHERE ur.nombre = 'Manager'
          AND COALESCE(a.activo, true) = true
        """
    )
    return construir_indices_manager(cur.fetchall())


def _resolver_manager(valor: Any, index: Dict[str, Any]) -> Tuple[Optional[int], List[str], str]:
    """Retorna (manager_id, mensajes, severidad_extra). severidad_extra: ok|warn|error"""
    return resolver_manager_id_por_indices(
        valor,
        index.get("by_agente") or {},
        index.get("by_email") or {},
    )


def _resolver_estado_id(estado_nombre: str, estados_map: Dict[str, int]) -> Optional[int]:
    return estados_map.get(estado_nombre.strip().lower())


def _obtener_zona_horaria_agencia(cur) -> Optional[str]:
    cur.execute(
        """
        SELECT valor
        FROM configuracion_agencia
        WHERE clave = 'zona_horaria'
        LIMIT 1
        """
    )
    row = cur.fetchone()
    if not row:
        return None
    raw = row.get("valor") or ""
    raw = str(raw).strip()
    return raw or None


def _buscar_duplicado_db(cur, usuario_norm: str) -> Optional[Dict[str, Any]]:
    cur.execute(
        f"""
        SELECT id, nombre, usuario_tiktok
        FROM creadores
        WHERE {_sql_usuario_normalizado()} = %s
        LIMIT 1
        """,
        (usuario_norm,),
    )
    row = cur.fetchone()
    if not row:
        return None
    return dict(row)


# ---------------------------------------------------------------------------
# Validación de filas
# ---------------------------------------------------------------------------


def _validar_fila(
    numero_fila: int,
    row: pd.Series,
    estados_map: Dict[str, int],
    managers_index: Dict[str, Any],
    usuarios_vistos: Dict[str, int],
    cur,
) -> Dict[str, Any]:
    datos_originales = {col: _celda_texto(row.get(col, "")) for col in COLUMNAS_OFICIALES}
    mensajes: List[str] = []
    errores_bloqueantes: List[str] = []

    nombre, msgs_nombre = normalizar_nombre(row.get("nombre"))
    mensajes.extend(msgs_nombre)
    if not nombre:
        errores_bloqueantes.extend(msgs_nombre or ["Nombre vacío"])

    usuario, msgs_user = normalizar_usuario_tiktok(row.get("usuario_tiktok"))
    mensajes.extend(msgs_user)
    if not usuario:
        errores_bloqueantes.extend(msgs_user or ["Usuario TikTok vacío"])

    telefono, msgs_tel = normalizar_telefono(row.get("telefono"))
    mensajes.extend(msgs_tel)

    email, msgs_email = normalizar_email(row.get("email"))
    if msgs_email and email is None:
        errores_bloqueantes.extend(msgs_email)
    else:
        mensajes.extend(msgs_email)

    fecha_inc, msgs_fecha = normalizar_fecha(row.get("fecha_incorporacion"))
    if msgs_fecha:
        errores_bloqueantes.extend(msgs_fecha)
    else:
        mensajes.extend(msgs_fecha)

    estado_nombre, msgs_estado = normalizar_estado_texto(row.get("estado"))
    mensajes.extend(msgs_estado)
    estado_id = None
    if not estado_nombre:
        errores_bloqueantes.extend(msgs_estado or ["Estado vacío"])
    else:
        estado_id = _resolver_estado_id(estado_nombre, estados_map)
        if estado_id is None:
            errores_bloqueantes.append(f'Estado "{estado_nombre}" no configurado en la agencia')

    manager_id, msgs_mgr, mgr_sev = _resolver_manager(row.get("manager"), managers_index)
    if mgr_sev == "error":
        errores_bloqueantes.extend(msgs_mgr)
    elif mgr_sev == "warn":
        mensajes.extend(msgs_mgr)

    datos_normalizados: Dict[str, Any] = {
        "nombre": nombre,
        "usuario_tiktok": usuario,
        "telefono": telefono,
        "email": email,
        "manager_id": manager_id,
        "fecha_incorporacion": fecha_inc.isoformat() if fecha_inc else None,
        "estado_id": estado_id,
        "estado_nombre": estado_nombre,
    }

    estado_fila = FILA_VALIDA
    referencia_duplicado = None

    if errores_bloqueantes:
        estado_fila = FILA_ERROR
        mensajes = list(dict.fromkeys(errores_bloqueantes + mensajes))
    elif usuario:
        if usuario in usuarios_vistos:
            estado_fila = FILA_DUPLICADO
            orig = usuarios_vistos[usuario]
            mensajes.append(f"Usuario repetido en las filas {orig} y {numero_fila}.")
        else:
            usuarios_vistos[usuario] = numero_fila
            dup_db = _buscar_duplicado_db(cur, usuario)
            if dup_db:
                estado_fila = FILA_DUPLICADO
                referencia_duplicado = dup_db
                mensajes.append(
                    f'Ya existe el creador "{dup_db.get("nombre")}" '
                    f'(@{dup_db.get("usuario_tiktok")}, id {dup_db.get("id")}).'
                )

    if estado_fila == FILA_VALIDA and mgr_sev == "warn":
        estado_fila = FILA_ADVERTENCIA
    if estado_fila == FILA_VALIDA and msgs_tel:
        estado_fila = FILA_ADVERTENCIA

    return {
        "numero_fila": numero_fila,
        "datos_originales": datos_originales,
        "datos_normalizados": datos_normalizados,
        "estado_fila": estado_fila,
        "mensajes": mensajes,
        "referencia_duplicado": referencia_duplicado,
    }


def _procesar_archivo(
    content: bytes,
    ext: str,
    separador: Optional[str],
) -> Dict[str, Any]:
    df = _leer_dataframe(content, ext, separador)
    _, advertencias_globales = _validar_encabezados(df)

    with get_connection_context() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            estados_map = _cargar_estados_map(cur)
            managers_index = _cargar_managers_index(cur)

            filas: List[Dict[str, Any]] = []
            usuarios_vistos: Dict[str, int] = {}

            for idx, row in df.iterrows():
                numero_fila = int(idx) + 2  # encabezado = fila 1
                fila = _validar_fila(
                    numero_fila,
                    row,
                    estados_map,
                    managers_index,
                    usuarios_vistos,
                    cur,
                )
                filas.append(fila)

    conteos = {
        "total_filas": len(filas),
        "validas": sum(1 for f in filas if f["estado_fila"] == FILA_VALIDA),
        "advertencias": sum(1 for f in filas if f["estado_fila"] == FILA_ADVERTENCIA),
        "duplicados": sum(1 for f in filas if f["estado_fila"] == FILA_DUPLICADO),
        "errores": sum(1 for f in filas if f["estado_fila"] == FILA_ERROR),
    }

    return {
        "ok": True,
        **conteos,
        "advertencias_globales": advertencias_globales,
        "filas": filas,
        "importables": conteos["validas"] + conteos["advertencias"],
    }


# ---------------------------------------------------------------------------
# Inserción
# ---------------------------------------------------------------------------


def _insertar_creador_importado(
    cur,
    datos: Dict[str, Any],
    zona_horaria: Optional[str],
) -> int:
    cur.execute(
        """
        INSERT INTO creadores (
            aspirante_id,
            nombre,
            usuario_tiktok,
            email,
            telefono,
            foto,
            creador_tiktok_id,
            estado_id,
            categoria_id,
            arquetipo_id,
            zona_horaria,
            created_at,
            updated_at
        )
        VALUES (
            NULL, %s, %s, %s, %s,
            NULL, %s, %s, NULL, NULL, %s,
            NOW(), NOW()
        )
        RETURNING id
        """,
        (
            datos["nombre"],
            datos["usuario_tiktok"],
            datos.get("email"),
            datos.get("telefono"),
            datos.get("creador_tiktok_id"),
            datos["estado_id"],
            zona_horaria,
        ),
    )
    row = cur.fetchone()
    creador_id = int(row["id"])

    fecha_inc = datos.get("fecha_incorporacion")
    cur.execute(
        """
        INSERT INTO creadores_detalle (
            creador_id,
            manager_id,
            horario_lives,
            tiempo_disponible,
            fecha_incorporacion,
            fecha_graduacion,
            seguidores,
            videos,
            me_gusta,
            diamantes,
            horas_live,
            numero_partidas,
            dias_emision,
            created_at,
            updated_at
        )
        VALUES (
            %s, %s, NULL, NULL, %s, NULL,
            0, 0, 0, 0, 0, 0, 0,
            NOW(), NOW()
        )
        """,
        (
            creador_id,
            datos.get("manager_id"),
            fecha_inc,
        ),
    )
    return creador_id


def _iso_ts(val: Any) -> Optional[str]:
    if val is None:
        return None
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _historial_item_desde_row(row: Dict[str, Any]) -> Dict[str, Any]:
    meta = row.get("metadata_json") or {}
    if isinstance(meta, str):
        try:
            meta = json.loads(meta)
        except (json.JSONDecodeError, TypeError):
            meta = {}
    return {
        "id": row.get("id"),
        "archivo_nombre": row.get("archivo_nombre"),
        "total_filas": row.get("total_filas") or 0,
        "validas": row.get("validas") or 0,
        "advertencias": row.get("advertencias") or 0,
        "duplicados": row.get("duplicados") or 0,
        "errores": row.get("errores") or 0,
        "importables": row.get("importables") or 0,
        "creadores_creados": row.get("creadores_creados") or 0,
        "omitidos": row.get("omitidos") or 0,
        "errores_importacion": row.get("errores_importacion") or 0,
        "estado": row.get("estado") or "procesado",
        "fecha_carga": _iso_ts(row.get("fecha_carga")),
        "cargado_por": row.get("cargado_por"),
        "cargado_por_nombre": row.get("cargado_por_nombre") or None,
        "errores_detalle": (meta or {}).get("errores_detalle") or [],
    }


def _asegurar_tabla_historial(cur) -> None:
    cur.execute(SQL_CREAR_TABLA_HISTORIAL)
    cur.execute(SQL_INDICE_HISTORIAL)


def _guardar_historial_importacion(
    resultado: Dict[str, Any],
    archivo_nombre: Optional[str],
    separador: Optional[str],
    cargado_por: Optional[int],
) -> None:
    preview = resultado.get("resumen") or {}
    errores_detalle = (resultado.get("errores_importacion") or [])[:50]
    try:
        with get_connection_context() as conn:
            with conn.cursor() as cur:
                _asegurar_tabla_historial(cur)
                cur.execute(
                    """
                    INSERT INTO creadores_importaciones (
                        archivo_nombre,
                        separador,
                        total_filas,
                        validas,
                        advertencias,
                        duplicados,
                        errores,
                        importables,
                        creadores_creados,
                        omitidos,
                        errores_importacion,
                        estado,
                        cargado_por,
                        metadata_json
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'procesado', %s, %s)
                    """,
                    (
                        (archivo_nombre or "")[:255] or None,
                        (separador or "")[:8] or None,
                        preview.get("total_filas") or resultado.get("total_procesadas") or 0,
                        preview.get("validas") or 0,
                        preview.get("advertencias") or 0,
                        preview.get("duplicados") or 0,
                        preview.get("errores") or 0,
                        preview.get("importables") or 0,
                        resultado.get("creadores_creados") or 0,
                        resultado.get("omitidos") or 0,
                        len(resultado.get("errores_importacion") or []),
                        cargado_por,
                        Json({"errores_detalle": errores_detalle}),
                    ),
                )
            conn.commit()
    except Exception as exc:
        print(f"⚠️ No se pudo guardar historial de importación de creadores: {exc}")
        traceback.print_exc()


def _listar_historial_importacion(limit: int = 100) -> List[Dict[str, Any]]:
    try:
        limit = max(1, min(int(limit or 100), 200))
    except (TypeError, ValueError):
        limit = 100
    try:
        with get_connection_context() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                try:
                    _asegurar_tabla_historial(cur)
                    conn.commit()
                except Exception:
                    conn.rollback()
                cur.execute(
                    """
                    SELECT
                        i.id,
                        i.archivo_nombre,
                        i.total_filas,
                        i.validas,
                        i.advertencias,
                        i.duplicados,
                        i.errores,
                        i.importables,
                        i.creadores_creados,
                        i.omitidos,
                        i.errores_importacion,
                        i.estado,
                        i.fecha_carga,
                        i.cargado_por,
                        i.metadata_json,
                        a.nombre_completo AS cargado_por_nombre
                    FROM creadores_importaciones i
                    LEFT JOIN administradores a ON a.id = i.cargado_por
                    ORDER BY i.fecha_carga DESC, i.id DESC
                    LIMIT %s
                    """,
                    (limit,),
                )
                return [_historial_item_desde_row(dict(row)) for row in cur.fetchall()]
    except pg_errors.UndefinedTable:
        return []
    except Exception as exc:
        print(f"❌ Error listando historial de importación de creadores: {exc}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al consultar el historial de cargues")


def _confirmar_filas(
    content: bytes,
    ext: str,
    separador: Optional[str],
    numeros_fila: List[int],
    archivo_nombre: Optional[str] = None,
    cargado_por: Optional[int] = None,
) -> Dict[str, Any]:
    preview = _procesar_archivo(content, ext, separador)
    numeros_set = set(numeros_fila)
    filas_objetivo = [f for f in preview["filas"] if f["numero_fila"] in numeros_set]

    if not filas_objetivo:
        raise HTTPException(status_code=400, detail="No hay filas seleccionadas para importar")

    creados = 0
    omitidos = 0
    errores_import: List[Dict[str, Any]] = []

    for fila in filas_objetivo:
        if fila["estado_fila"] not in (FILA_VALIDA, FILA_ADVERTENCIA):
            omitidos += 1
            continue

        try:
            with get_connection_context() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    estados_map = _cargar_estados_map(cur)
                    managers_index = _cargar_managers_index(cur)
                    zona_horaria = _obtener_zona_horaria_agencia(cur)

                    # Revalidación completa en confirmación
                    datos = fila["datos_normalizados"]
                    usuario = datos.get("usuario_tiktok")
                    if not usuario:
                        raise ValueError("Usuario TikTok inválido")

                    dup = _buscar_duplicado_db(cur, usuario)
                    if dup:
                        omitidos += 1
                        errores_import.append({
                            "numero_fila": fila["numero_fila"],
                            "mensaje": "Duplicado detectado al confirmar",
                        })
                        continue

                    estado_nombre = datos.get("estado_nombre")
                    estado_id = _resolver_estado_id(estado_nombre or "", estados_map)
                    if estado_id is None:
                        raise ValueError(f"Estado no válido: {estado_nombre}")

                    manager_id = None
                    mgr_raw = fila["datos_originales"].get("manager")
                    if mgr_raw:
                        mid, msgs, sev = _resolver_manager(mgr_raw, managers_index)
                        if sev == "error":
                            raise ValueError("; ".join(msgs))
                        manager_id = mid

                    payload = {
                        "nombre": datos["nombre"],
                        "usuario_tiktok": usuario,
                        "telefono": datos.get("telefono"),
                        "email": datos.get("email"),
                        "estado_id": estado_id,
                        "manager_id": manager_id,
                        "fecha_incorporacion": datos.get("fecha_incorporacion"),
                    }
                    _insertar_creador_importado(cur, payload, zona_horaria)
                conn.commit()
                creados += 1
        except Exception as exc:
            omitidos += 1
            errores_import.append({
                "numero_fila": fila["numero_fila"],
                "mensaje": str(exc),
            })
            print(f"❌ Error importando fila {fila['numero_fila']}: {exc}")
            traceback.print_exc()

    resultado = {
        "ok": True,
        "mensaje": "Importación terminada",
        "total_procesadas": len(filas_objetivo),
        "creadores_creados": creados,
        "omitidos": omitidos,
        "errores_importacion": errores_import,
        "resumen": preview,
    }
    _guardar_historial_importacion(resultado, archivo_nombre, separador, cargado_por)
    return resultado


# ---------------------------------------------------------------------------
# Plantilla
# ---------------------------------------------------------------------------


def _generar_plantilla_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Creadores"

    headers = COLUMNAS_OFICIALES
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ejemplos = [
        ["María López", "maria_live", "+573001234567", "maria@email.com", "alejandra@agencia.com", "2026-05-10", "Activo"],
        ["Carlos Pérez", "carlosperez", "", "", "daniela@agencia.com", "2025-11-15", "Inactivo"],
    ]
    for row in ejemplos:
        ws.append(row)

    # Teléfono como texto
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "@"
        ws.cell(row=row_idx, column=6).number_format = "yyyy-mm-dd"

    dv = DataValidation(type="list", formula1='"Activo,Inactivo"', allow_blank=False)
    dv.add(f"G2:G{MAX_FILAS}")
    ws.add_data_validation(dv)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generar_plantilla_csv() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(COLUMNAS_OFICIALES)
    writer.writerow(["María López", "maria_live", "+573001234567", "maria@email.com", "alejandra@agencia.com", "2026-05-10", "Activo"])
    writer.writerow(["Carlos Pérez", "carlosperez", "", "", "daniela@agencia.com", "2025-11-15", "Inactivo"])
    return buffer.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/api/creadores/importacion/plantilla")
def descargar_plantilla_importacion(
    formato: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    usuario: dict = Depends(obtener_usuario_actual),
):
    _require_permiso_importacion(usuario)

    if formato == "csv":
        contenido = _generar_plantilla_csv()
        return Response(
            content=contenido,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="plantilla_importar_creadores.csv"'},
        )

    contenido = _generar_plantilla_xlsx()
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_importar_creadores.xlsx"'},
    )


@router.post("/api/creadores/importacion/validar")
async def validar_importacion_creadores(
    file: UploadFile = File(...),
    separador: Optional[str] = Form(None),
    usuario: dict = Depends(obtener_usuario_actual),
):
    _require_permiso_importacion(usuario)

    content = await file.read()
    ext = _validar_archivo(file, content)

    try:
        return _procesar_archivo(content, ext, separador or None)
    except HTTPException:
        raise
    except Exception as exc:
        print("❌ Error validando importación de creadores:", exc)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al validar el archivo")


class ConfirmarImportacionIn(BaseModel):
    numeros_fila: List[int]


@router.post("/api/creadores/importacion/confirmar")
async def confirmar_importacion_creadores(
    file: UploadFile = File(...),
    numeros_fila: str = Form(...),
    separador: Optional[str] = Form(None),
    usuario: dict = Depends(obtener_usuario_actual),
):
    _require_permiso_importacion(usuario)

    try:
        nums = json.loads(numeros_fila)
        if not isinstance(nums, list):
            raise ValueError("numeros_fila debe ser una lista")
        numeros = [int(n) for n in nums]
    except (json.JSONDecodeError, ValueError, TypeError) as exc:
        raise HTTPException(status_code=400, detail=f"numeros_fila inválido: {exc}") from exc

    content = await file.read()
    ext = _validar_archivo(file, content)

    try:
        return _confirmar_filas(
            content,
            ext,
            separador or None,
            numeros,
            archivo_nombre=file.filename,
            cargado_por=usuario.get("id") if usuario else None,
        )
    except HTTPException:
        raise
    except Exception as exc:
        print("❌ Error confirmando importación de creadores:", exc)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al importar creadores")


@router.get("/api/creadores/importacion/historial")
def listar_historial_importacion_creadores(
    limit: int = Query(100, ge=1, le=200),
    usuario: dict = Depends(obtener_usuario_actual),
):
    _require_permiso_importacion(usuario)
    return {"ok": True, "items": _listar_historial_importacion(limit)}

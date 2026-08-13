"""
Carga sencilla de información del asistente conversacional.

Una sola experiencia: pegar textos por categoría o importar Excel → analizar
(dry-run) → revisar → guardar. Reutiliza las tablas existentes; no escribe en
análisis; no desactiva registros omitidos al guardar.
"""
from __future__ import annotations

import io
import json
import logging
import re
import unicodedata
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import database_chatbot_conversacional as db
from chatbot_conversacional_exceptions import ConversacionalError

logger = logging.getLogger("uvicorn.error")

CODIGO_FLUJO_CONVERSION = "conversion_base"
CODIGO_RECURSO_SOLICITUD = "solicitud_principal"

# chk_regla_escalamiento_prioridad: varchar enum (NO es entero 0-100 como FAQ).
PRIORIDADES_REGLA_ESCALAMIENTO = frozenset({"baja", "normal", "alta", "urgente"})
# Semántica: 'urgente' > 'alta' > 'normal' > 'baja' (severidad de atención humana).
# El orden de resolución entre reglas usa `orden`, no este campo.
EVENTOS_CONTACTO_HUMANO = frozenset(
    {"solicitud_humano", "transferir_humano", "escalar"}
)
# Alineado con la regla plantilla: «El usuario solicita expresamente hablar…»
PRIORIDAD_CONTACTO_HUMANO = "alta"

_URL_RE = re.compile(r"https?://[^\s<>\"']+", re.I)
_PREGUNTA_RE = re.compile(
    r"(?:pregunta|q)\s*[:\-–]\s*(.+?)(?:\n|\r|$)", re.I
)
_RESPUESTA_RE = re.compile(
    r"(?:respuesta|r)\s*[:\-–]\s*(.+?)(?=(?:\n|\r)\s*(?:pregunta|q)\s*[:\-–]|\Z)",
    re.I | re.S,
)


def _slug(valor: Any, prefijo: str = "item") -> str:
    return db._slug_codigo(valor, prefijo=prefijo)


def _codigo_unico(
    base: Any,
    usados: set,
    *,
    prefijo: str = "item",
    indice: int = 1,
    max_len: int = 80,
) -> str:
    """Genera codigo único ≤ max_len (varchar(80) en BD)."""
    codigo = str(base or "").strip() or _slug(prefijo, prefijo)
    codigo = _slug(codigo, prefijo)[:max_len]
    if codigo not in usados:
        return codigo
    suf = f"_{indice}"
    recorte = max_len - len(suf)
    if recorte < 1:
        recorte = 1
        suf = suf[: max_len - 1]
    candidato = f"{codigo[:recorte]}{suf}"
    n = indice
    while candidato in usados:
        n += 1
        suf = f"_{n}"
        recorte = max(1, max_len - len(suf))
        candidato = f"{codigo[:recorte]}{suf}"
    return candidato[:max_len]


def normalizar_prioridad_regla_escalamiento(valor: Any) -> str:
    """
    Normaliza prioridad al enum de BD:
    baja | normal | alta | urgente.

    Acepta el enum textual o enteros legacy (confundidos con prioridad FAQ 0-100)
    y nunca devuelve un valor incompatible con chk_regla_escalamiento_prioridad.
    """
    if isinstance(valor, str):
        v = valor.strip().lower()
        if v in PRIORIDADES_REGLA_ESCALAMIENTO:
            return v
        # Alias frecuentes desde IA / textos libres
        aliases = {
            "low": "baja",
            "medium": "normal",
            "med": "normal",
            "high": "alta",
            "critical": "urgente",
            "critica": "urgente",
            "crítica": "urgente",
        }
        if v in aliases:
            return aliases[v]
    if isinstance(valor, bool):
        return "normal"
    if isinstance(valor, (int, float)):
        # Legacy erróneo: enteros estilo FAQ. Mapear a severidad razonable.
        n = int(valor)
        if n >= 80:
            return "urgente"
        if n >= 40:
            return "alta"
        if n >= 10:
            return "normal"
        if n > 0:
            return "baja"
        return "normal"
    return "normal"


def _campos_regla_escalamiento_seguros(campos: Dict[str, Any]) -> Dict[str, Any]:
    """Garantiza prioridad válida antes de crear/actualizar la regla."""
    out = dict(campos or {})
    out["prioridad"] = normalizar_prioridad_regla_escalamiento(out.get("prioridad"))
    return out


def _es_regla_contacto_humano_equivalente(regla: Optional[Dict[str, Any]]) -> bool:
    """
    Detecta reglas ya existentes que cubren «hablar con una persona».

    Incluye la plantilla:
    «El usuario solicita expresamente hablar con una persona, asesor,
    reclutador o manager.»
    """
    if not isinstance(regla, dict):
        return False
    evento = str(regla.get("evento") or "").strip().lower()
    if evento in EVENTOS_CONTACTO_HUMANO:
        return True
    desc = _norm_nombre(
        " ".join(
            [
                str(regla.get("descripcion") or ""),
                str(regla.get("nombre") or ""),
                str(regla.get("mensaje_usuario") or ""),
            ]
        )
    )
    if "contacto humano" in desc:
        return True
    if "solicita expresamente" in desc and "persona" in desc:
        return True
    if "hablar con una persona" in desc or "hablar con un asesor" in desc:
        return True
    return False


def _prioridad_contacto_humano(valor: Any) -> str:
    """
    Prioridad para solicitud_humano.

    Si el valor ya es un enum válido, se respeta; enteros legacy (p.ej. 10)
    u otros inválidos → 'alta' (misma semántica que la regla plantilla).
    """
    if isinstance(valor, str):
        v = valor.strip().lower()
        if v in PRIORIDADES_REGLA_ESCALAMIENTO:
            return v
    return PRIORIDAD_CONTACTO_HUMANO


def _norm_nombre(valor: Any) -> str:
    texto = str(valor or "").strip().lower()
    # Quitar viñetas/guiones repetidos para dedupe ("- - edad" == "edad")
    prev = None
    while prev != texto:
        prev = texto
        texto = re.sub(r"^[\-\*\u2022●•]+[\.\)\-]?\s*", "", texto).strip()
        texto = re.sub(r"^\d+[\.\)\-]\s*", "", texto).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^\w\s]", " ", texto, flags=re.UNICODE)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def _lineas_bullet(texto: str) -> List[str]:
    out: List[str] = []
    for raw in (texto or "").splitlines():
        linea = raw.strip()
        if not linea:
            continue
        # Quitar prefijos de lista repetidos: "- - - Nombre"
        prev = None
        while prev != linea:
            prev = linea
            linea = re.sub(r"^[\-\*\u2022●•]+[\.\)\-]?\s*", "", linea).strip()
            linea = re.sub(r"^\d+[\.\)\-]\s*", "", linea).strip()
        if linea:
            out.append(linea)
    return out


def _parece_dump_lista(texto: str) -> bool:
    t = str(texto or "")
    guiones = t.count(" - ") + t.count("\n-") + t.count("- -")
    pasos_num = len(re.findall(r"(?:^|\s)\d+[\.\)]\s+\S", t))
    # Frases que se repiten muchas veces = dump fractal
    repeticiones = 0
    for frag in (
        "mayoría de edad",
        "disponibilidad",
        "teléfono",
        "ser mayor",
        "disposición",
    ):
        repeticiones += t.lower().count(frag)
    return (
        guiones >= 2
        or pasos_num >= 2
        or repeticiones >= 3
        or len(t) > 800
    )


def _texto_corto_limpio(texto: Optional[str], *, max_len: int = 500) -> str:
    """Corta dumps anidados y deja un texto usable."""
    d = str(texto or "").strip()
    if not d:
        return ""
    if _parece_dump_lista(d):
        m = re.search(
            r"\s+-\s+-|\n\s*[-*]|\s+—\s+\d+[\.\)]|\s+\d+[\.\)]\s+\S",
            d,
        )
        if m:
            d = d[: m.start()].strip()
        if _parece_dump_lista(d) or len(d) > max_len:
            d = re.split(r"[.\n]", d, maxsplit=1)[0].strip()
    return d[:max_len].strip()


def _descripcion_requisito_limpia(nombre: str, desc: Optional[str]) -> str:
    """
    Evita reinyectar dumps anidados de listas previas en la descripción.
    """
    d = _texto_corto_limpio(desc, max_len=500)
    if not d or d in {"-", "—", "–", ".", "…"}:
        return ""
    n = _nombre_item_limpio(str(nombre or "").strip())
    if n and d.lower().startswith(n.lower()):
        d = d[len(n) :].lstrip(" .:-–—")
    if not d or d in {"-", "—", "–", ".", "…"}:
        return ""
    if _norm_nombre(d) == _norm_nombre(n):
        return ""
    # Descripción que es solo otro dump de lista
    if _parece_dump_lista(d) or d.startswith("-"):
        d = _texto_corto_limpio(d, max_len=240)
    if not d or d in {"-", "—", "–", ".", "…"}:
        return ""
    return d.strip()


def _tema_requisito_carga(texto: str) -> Optional[str]:
    """
    Agrupa variantes equivalentes (p. ej. 'Mayoría de edad' ≈ 'Ser mayor de edad')
    para no crear filas nuevas en cada guardado.
    """
    n = _norm_nombre(_nombre_item_limpio(texto))
    if not n:
        return None
    if any(k in n for k in ("18", "edad", "mayor de edad", "mayoria")):
        return "edad"
    if ("hora" in n or "horas" in n) and (
        "dia" in n or "diaria" in n or "diario" in n
    ):
        return "horas_dia"
    if "semana" in n or "semanal" in n:
        return "semana"
    if any(k in n for k in ("telefono", "celular", "conexion", "internet")):
        return "equipo"
    if any(
        k in n
        for k in ("hablador", "convers", "interact", "comunicativ", "audiencia")
    ):
        return "comunicacion"
    if any(k in n for k in ("energia", "camara", "entretenid")):
        return "energia"
    if any(
        k in n
        for k in ("responsable", "responsabilidad", "constancia", "compromiso")
    ):
        return "compromiso"
    if "batalla" in n:
        return "batallas"
    if any(k in n for k in ("acompanamiento", "capacitacion", "orientacion")):
        return "acompanamiento"
    if any(k in n for k in ("interes", "crecer", "aprender", "mejorar")):
        return "crecimiento"
    if any(k in n for k in ("disponibilidad", "disponible", "transmit")):
        return "disponibilidad"
    return None


def _clave_dedupe_requisito(nombre: str) -> str:
    limpio = _nombre_item_limpio(nombre)
    tema = _tema_requisito_carga(limpio)
    if tema:
        return f"tema:{tema}"
    return f"n:{_norm_nombre(limpio)}"


_ETIQUETAS_TEMA_BENEFICIO = {
    "acompanamiento": "Acompañamiento de manager",
    "capacitacion": "Capacitaciones para creadores LIVE",
    "comunidad": "Comunidad y participación en batallas",
    "bono_meta": "Bono por cumplimiento de meta",
    "bono_incorporacion": "Bono de incorporación",
}


def _familia_tipo_beneficio(tipo: str) -> str:
    t = str(tipo or "beneficio").lower().strip()
    if t in {"bono", "incentivo"}:
        return "bono"
    return "beneficio"


def _tema_beneficio_carga(nombre: str, desc: Optional[str] = None) -> Optional[str]:
    """Agrupa variantes de beneficios/bonos (título vs descripción como nombre)."""
    n = _norm_nombre(
        f"{_nombre_item_limpio(nombre)} {_nombre_item_limpio(desc or '', max_len=240)}"
    )
    if not n:
        return None
    if re.fullmatch(r"\d+([.,]\d+)?\s*(usd|us\$|eur|cop)?", n) or n in {
        "valor",
        "00 usd",
        "0 usd",
    }:
        return "junk_money"
    if any(
        k in n
        for k in (
            "incorporacion",
            "nuevos creadores",
            "nuevo creador",
            "bienvenida",
            "periodo de vigencia",
        )
    ):
        return "bono_incorporacion"
    if any(
        k in n
        for k in (
            "cumplimiento de meta",
            "bono por cumplimiento",
            "incentivo sujeto",
            "meta durante",
        )
    ):
        return "bono_meta"
    if any(k in n for k in ("acompanamiento", "manager", "orientacion")):
        return "acompanamiento"
    if any(
        k in n
        for k in (
            "capacitacion",
            "aprendizaje",
            "habilidades de transmision",
            "clases",
        )
    ):
        return "capacitacion"
    if any(
        k in n
        for k in (
            "batalla",
            "comunidad",
            "actividades",
            "oportunidades de participar",
        )
    ):
        return "comunidad"
    return None


def _clave_dedupe_beneficio(
    tipo: str, nombre: str, desc: Optional[str] = None
) -> str:
    tema = _tema_beneficio_carga(nombre, desc)
    if tema == "junk_money":
        return "tema:junk_money"
    if tema:
        return f"tema:{tema}"
    fam = _familia_tipo_beneficio(tipo)
    return f"{fam}:{_norm_nombre(_nombre_item_limpio(nombre))}"


def _nombre_beneficio_parece_descripcion(nombre: str) -> bool:
    n = _nombre_item_limpio(nombre)
    if not n:
        return True
    if len(n) > 80:
        return True
    low = n.lower()
    if low.startswith(
        (
            "la agencia",
            "espacios de",
            "orientaci",
            "incentivo sujeto",
            "bono dirigido",
            "bono para",
            "participaci",
            "valor:",
            "valor ",
        )
    ):
        return True
    if re.match(r"^[\d.,]+\s*(usd|us\$|\$)?$", low):
        return True
    return False


def _desc_beneficio_de_row(row: Dict[str, Any]) -> str:
    return str(
        row.get("descripcion_corta")
        or row.get("texto_autorizado")
        or row.get("descripcion_completa")
        or row.get("descripcion")
        or ""
    ).strip()


def _score_beneficio_keeper(row: Dict[str, Any]) -> Tuple[int, int]:
    nombre = str(row.get("nombre") or "")
    desc = _desc_beneficio_de_row(row)
    score = 0
    if row.get("activo") is not False:
        score += 1000
    if not _nombre_beneficio_parece_descripcion(nombre):
        score += 400
    if not _nombre_parece_fractal(nombre):
        score += 200
    nlimpio = _nombre_item_limpio(nombre)
    score += max(0, 100 - len(nlimpio))
    d = _descripcion_requisito_limpia(nlimpio, desc)
    if d and _norm_nombre(d) != _norm_nombre(nlimpio):
        score += min(len(d), 120)
    tipo = str(row.get("tipo") or "").lower()
    tema = _tema_beneficio_carga(nombre, desc)
    if tema and tema.startswith("bono_") and tipo in {"bono", "incentivo"}:
        score += 80
    if tema and not tema.startswith("bono_") and tipo not in {"bono", "incentivo"}:
        score += 40
    if row.get("valor") not in (None, "", 0, 0.0):
        score += 60
    rid = int(row.get("id") or 0)
    return (score, -rid)


def _nombre_parece_fractal(nombre: str) -> bool:
    n = str(nombre or "").strip()
    if not n:
        return True
    if n.startswith("-") or " - -" in n or n.count(" - ") >= 2:
        return True
    return _parece_dump_lista(n)


def _score_requisito_keeper(row: Dict[str, Any]) -> Tuple[int, int]:
    """Mayor score = mejor fila a conservar. Desempate: id más bajo."""
    nombre = str(row.get("nombre") or "")
    desc = str(row.get("descripcion") or "")
    score = 0
    if row.get("activo") is not False:
        score += 1000
    if not _nombre_parece_fractal(nombre):
        score += 300
    nlimpio = _nombre_item_limpio(nombre)
    # Prefiere títulos cortos frente a frases largas usadas como nombre.
    score += max(0, 120 - len(nlimpio))
    d = _descripcion_requisito_limpia(nlimpio, desc)
    if d:
        score += min(len(d), 120)
    if _parece_dump_lista(desc) or _nombre_parece_fractal(desc):
        score -= 200
    rid = int(row.get("id") or 0)
    return (score, -rid)


def _nombre_item_limpio(nombre: str, *, max_len: int = 160) -> str:
    n = str(nombre or "").strip()
    prev = None
    while prev != n:
        prev = n
        n = re.sub(r"^[\-\*\u2022●•]+[\.\)\-]?\s*", "", n).strip()
        n = re.sub(r"^\d+[\.\)\-]\s*", "", n).strip()
    # Cortar dumps "Nombre - - Nombre..."
    if " - -" in n or n.count(" - ") >= 2:
        n = re.split(r"\s+-\s+-|\s+-\s+", n, maxsplit=1)[0].strip()
    if _parece_dump_lista(n) or len(n) > max_len:
        n = n.split(".")[0].split("—")[0].split(" - ")[0].strip()[:max_len]
    return n[:max_len].strip()


def _linea_requisito_legible(nombre: str, desc: Optional[str]) -> Optional[str]:
    n = _nombre_item_limpio(nombre)
    if not n:
        return None
    d = _descripcion_requisito_limpia(n, desc)
    if d:
        return f"- {n}. {d}"
    return f"- {n}"


def _linea_beneficio_legible(
    nombre: str,
    desc: Optional[str],
    *,
    valor: Any = None,
    moneda: Optional[str] = None,
) -> Optional[str]:
    n = _nombre_item_limpio(nombre)
    if not n:
        return None
    # No listar basura monetaria como ficha.
    if re.match(
        r"^(valor\s*:?\s*)?[\d.,]+\s*(usd|us\$|\$|eur|cop)?$",
        n.strip(),
        re.I,
    ):
        return None
    d = _descripcion_requisito_limpia(n, desc)
    # Evitar meter "Valor: …" dentro de la descripción.
    if d:
        d = re.sub(
            r"(?im)^\s*valor\s*:\s*[\d.,]+\s*[A-Za-z$€]*\s*$",
            "",
            d,
        ).strip()
    lineas = [f"- {n}"]
    if d and _norm_nombre(d) != _norm_nombre(n):
        lineas.append(f"  {d}")
    if valor not in (None, ""):
        try:
            vnum = float(valor)
        except (TypeError, ValueError):
            vnum = None
        if vnum is not None and vnum != 0:
            mon = (moneda or "").strip()
            if vnum == int(vnum):
                valor_txt = str(int(vnum))
            else:
                valor_txt = f"{vnum:.2f}"
            lineas.append(f"  Valor: {valor_txt} {mon}".rstrip())
    return "\n".join(lineas)


def _linea_proceso_legible(
    orden: int, nombre: str, mensaje: Optional[str]
) -> Optional[str]:
    n = _nombre_item_limpio(nombre)
    if not n:
        return None
    msg = _texto_corto_limpio(mensaje, max_len=300)
    # Evitar mensaje = nombre o dump del proceso completo
    if msg and _norm_nombre(msg) == _norm_nombre(n):
        msg = ""
    if msg and (
        _parece_dump_lista(msg)
        or msg.count(".") > 4
        or re.search(r"\d+[\.\)]\s+\S", msg)
    ):
        # Si quedó resto con otro paso numerado, quedarse solo con la 1ª frase
        primera = re.split(r"[.\n]", msg, maxsplit=1)[0].strip()
        msg = primera if primera and not re.search(r"\d+[\.\)]\s+\S", primera) else ""
    if msg:
        return f"{orden}. {n} — {msg}"
    return f"{orden}. {n}"


def _url_real(url: Optional[str]) -> Optional[str]:
    texto = str(url or "").strip()
    if not texto:
        return None
    bajos = texto.lower()
    if bajos in {"#", "about:blank"} or "example.com" in bajos or "localhost" in bajos:
        return None
    parsed = urlparse(texto if "://" in texto else f"https://{texto}")
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return None
    return texto if "://" in texto else f"https://{texto}"


# ---------------------------------------------------------------------------
# Textos legibles desde BD
# ---------------------------------------------------------------------------


def obtener_textos_carga(
    agencia_id: int, chatbot_configuracion_id: int, *, cur=None
) -> Dict[str, Any]:
    with db._cursor(cur) as c:
        db._exige_configuracion(agencia_id, chatbot_configuracion_id, cur=c)
        asistente = db.obtener_asistente_por_config(
            agencia_id, chatbot_configuracion_id, cur=c
        ) or {}
        requisitos = db.listar_requisitos(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            incluir_globales=False,
            solo_activos=True,
            cur=c,
        )
        beneficios = db.listar_beneficios(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            incluir_globales=False,
            solo_activos=True,
            cur=c,
        )
        faqs = db.listar_faqs(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            incluir_globales=False,
            solo_activos=True,
            cur=c,
        )
        recursos = db.listar_recursos(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            incluir_globales=False,
            solo_activos=True,
            cur=c,
        )
        reglas = db.listar_reglas_escalamiento(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            solo_activas=True,
            cur=c,
        )
        flujos = db.listar_flujos(
            agencia_id,
            chatbot_configuracion_id=chatbot_configuracion_id,
            tipo_flujo="conversion",
            solo_activos=True,
            cur=c,
        )

        req_txt = []
        vistos_req: set = set()
        for r in requisitos:
            nombre = _nombre_item_limpio((r.get("nombre") or "").strip())
            clave = _clave_dedupe_requisito(nombre)
            if not clave or clave in vistos_req:
                continue
            # Saltar filas basura cuyo nombre limpio quedó vacío o es un dump
            if not nombre or _parece_dump_lista(nombre):
                continue
            linea = _linea_requisito_legible(nombre, r.get("descripcion"))
            if not linea:
                continue
            vistos_req.add(clave)
            req_txt.append(linea)

        ben_txt, bon_txt = [], []
        vistos_ben: set = set()
        vistos_bon: set = set()
        for b in beneficios:
            tipo = str(b.get("tipo") or "beneficio").lower()
            nombre = _nombre_item_limpio((b.get("nombre") or "").strip())
            if not nombre or _parece_dump_lista(nombre):
                continue
            desc = _desc_beneficio_de_row(b)
            clave = _clave_dedupe_beneficio(tipo, nombre, desc)
            if not clave or clave == "tema:junk_money":
                continue
            if tipo in {"bono", "incentivo"}:
                if clave in vistos_bon:
                    continue
                linea = _linea_beneficio_legible(
                    nombre,
                    desc,
                    valor=b.get("valor"),
                    moneda=b.get("moneda"),
                )
                if not linea:
                    continue
                vistos_bon.add(clave)
                bon_txt.append(linea)
            else:
                if clave in vistos_ben:
                    continue
                # Si el nombre era una descripción, preferir etiqueta de tema
                tema = _tema_beneficio_carga(nombre, desc)
                if tema and _nombre_beneficio_parece_descripcion(nombre):
                    if not desc:
                        desc = nombre
                    nombre = _ETIQUETAS_TEMA_BENEFICIO.get(tema, nombre)
                linea = _linea_beneficio_legible(nombre, desc)
                if not linea:
                    continue
                vistos_ben.add(clave)
                ben_txt.append(linea)

        faq_txt = []
        vistos_faq: set = set()
        for f in faqs:
            preg = (f.get("pregunta") or "").strip()
            resp = (f.get("respuesta_completa") or f.get("respuesta_corta") or "").strip()
            clave = _norm_nombre(preg)
            if not preg or not resp or not clave or clave in vistos_faq:
                continue
            if _parece_dump_lista(preg) or _parece_dump_lista(resp):
                preg = _nombre_item_limpio(preg, max_len=200)
                resp = _texto_corto_limpio(resp, max_len=800)
            if not preg or not resp:
                continue
            vistos_faq.add(clave)
            faq_txt.append(f"Pregunta: {preg}\nRespuesta: {resp}")

        proceso_txt = []
        vistos_paso: set = set()
        if flujos:
            pasos = db.listar_flujo_pasos(
                agencia_id, int(flujos[0]["id"]), solo_activos=True, cur=c
            )
            orden_vis = 0
            for p in pasos:
                nombre = (p.get("nombre") or p.get("codigo") or "").strip()
                clave = _norm_nombre(nombre)
                if not clave or clave in vistos_paso:
                    continue
                msg = (p.get("mensaje_instrucciones") or "").strip()
                orden_vis += 1
                linea = _linea_proceso_legible(orden_vis, nombre, msg)
                if not linea:
                    continue
                vistos_paso.add(clave)
                proceso_txt.append(linea)

        enlaces_txt = []
        for r in recursos:
            nombre = (r.get("nombre") or r.get("tipo") or "Enlace").strip()
            url = (r.get("url_template") or "").strip()
            boton = (r.get("texto_boton") or "").strip()
            bloque = f"{nombre}:\n{url}"
            if boton:
                bloque += f"\nTexto del botón: {boton}"
            enlaces_txt.append(bloque)
        for regla in reglas[:1]:
            equipo = (regla.get("equipo_destino") or "").strip()
            if equipo:
                enlaces_txt.append(f"Equipo o contacto:\n{equipo}")

        tono = asistente.get("tono") or "cercano"
        if tono not in {"profesional", "cercano", "juvenil"}:
            tono = "cercano"

        general = {
            "nombre_asistente": asistente.get("nombre_asistente") or "",
            "presentacion_inicial": asistente.get("presentacion_inicial") or "",
            "presentacion_informativo": asistente.get("presentacion_informativo") or "",
            "presentacion_inteligente": asistente.get("presentacion_inteligente") or "",
            "tono": tono,
        }
        _log_presentacion_get(
            agencia_id=agencia_id,
            config_id=chatbot_configuracion_id,
            fuente="obtener_textos_carga",
            asistente=asistente,
            general=general,
        )
        try:
            raw = db.leer_presentaciones_raw(
                agencia_id, chatbot_configuracion_id, cur=c
            )
            logger.info(
                "[CARGA_INFO_PRESENTACION_GET] fuente=obtener_textos_carga_raw "
                "agencia_id=%s config_id=%s columnas_bd=%s "
                "raw_inicial_chars=%s raw_informativo_chars=%s "
                "raw_inicial_inicio=%r raw_informativo_inicio=%r",
                agencia_id,
                chatbot_configuracion_id,
                raw.get("columnas_presentacion_en_bd"),
                len(str(raw.get("presentacion_inicial") or "")),
                len(str(raw.get("presentacion_informativo") or ""))
                if raw.get("presentacion_informativo") != "__COLUMNA_AUSENTE__"
                else -1,
                _inicio_texto(raw.get("presentacion_inicial")),
                _inicio_texto(raw.get("presentacion_informativo"))
                if raw.get("presentacion_informativo") != "__COLUMNA_AUSENTE__"
                else "__COLUMNA_AUSENTE__",
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "[CARGA_INFO_PRESENTACION_GET] raw falló: %s", exc
            )

        return {
            "general": general,
            "requisitos_texto": "\n".join(req_txt),
            # Separar fichas con línea en blanco para que el parse no parta
            # descripciones/Valor: como ítems nuevos.
            "beneficios_texto": "\n\n".join(ben_txt),
            "bonos_texto": "\n\n".join(bon_txt),
            "faq_texto": "\n\n".join(faq_txt),
            "proceso_ingreso_texto": "\n".join(proceso_txt),
            "enlaces_contacto_texto": "\n\n".join(enlaces_txt),
            "asistente_existe": bool(asistente.get("id")),
        }


# ---------------------------------------------------------------------------
# Análisis (dry-run): heurística + OpenAI opcional
# ---------------------------------------------------------------------------


def _parse_requisitos(texto: str) -> List[Dict[str, Any]]:
    items = []
    vistos: set = set()
    for i, linea in enumerate(_lineas_bullet(texto), start=1):
        linea = _nombre_item_limpio(linea, max_len=500)
        if not linea:
            continue
        if _parece_dump_lista(linea) and linea.count(".") > 3:
            # Línea basura: intentar rescatar solo el primer nombre
            nombre = _nombre_item_limpio(linea.split(".")[0].split(" - ")[0])
            if not nombre:
                continue
            desc = ""
        else:
            if "." in linea:
                nombre_raw, resto = linea.split(".", 1)
                nombre = _nombre_item_limpio(nombre_raw)[:160]
                desc = _descripcion_requisito_limpia(nombre, resto.strip())
            else:
                # "Nombre - dump" o solo nombre
                nombre = _nombre_item_limpio(linea.split(" - ")[0])[:160]
                desc = ""
        clave = _clave_dedupe_requisito(nombre)
        if not clave or clave == "n:" or clave in vistos:
            continue
        # Descartar restos que son solo eco del nombre
        if desc and _norm_nombre(desc) == _norm_nombre(nombre):
            desc = ""
        vistos.add(clave)
        items.append(
            {
                "nombre": nombre,
                "descripcion": desc or None,
                "categoria": "obligatorio",
                "bloquea_proceso": True,
                "mensaje_si_no_cumple": None,
                "orden": len(items) + 1,
            }
        )
    return items


def _parse_beneficios(texto: str, tipo: str = "beneficio") -> List[Dict[str, Any]]:
    """
    Parsea beneficios/bonos en bloques.

    Formato esperado (una ficha por item):
      - Nombre del bono
        Descripcion…
        Valor: 35.00 USD

    No trata cada linea suelta como un item nuevo (eso fragmentaba
    Valor: y descripciones largas).
    """
    items: List[Dict[str, Any]] = []
    vistos: set = set()

    def _es_inicio_item(linea: str) -> bool:
        return bool(re.match(r"^[\-\*•]\s+\S", linea.strip()))

    def _es_linea_valor(linea: str) -> bool:
        s = linea.strip()
        if re.match(r"^valor\s*:\s*[\d.,]+", s, re.I):
            return True
        if re.match(r"^[\d.,]+\s*(usd|us\$|\$|eur|cop)?$", s, re.I):
            return True
        return False

    # Armar bloques: nueva vineta "- …" o linea en blanco = nuevo item.
    bloques: List[str] = []
    actual: List[str] = []
    for raw in (texto or "").splitlines():
        stripped = raw.strip()
        if not stripped:
            if actual:
                bloques.append("\n".join(actual))
                actual = []
            continue
        if _es_inicio_item(stripped) and actual:
            bloques.append("\n".join(actual))
            actual = [stripped]
        else:
            actual.append(stripped)
    if actual:
        bloques.append("\n".join(actual))

    for bloque in bloques:
        lineas = [x.strip() for x in str(bloque).splitlines() if x.strip()]
        if not lineas:
            continue
        if _es_linea_valor(lineas[0]):
            continue
        nombre = _nombre_item_limpio(re.sub(r"^[\-\*•]\s*", "", lineas[0]))
        if not nombre:
            continue
        if _tema_beneficio_carga(nombre) == "junk_money":
            continue
        if re.match(r"^valor\s*:\s*", nombre, re.I):
            continue

        cuerpo: List[str] = []
        for ln in lineas[1:]:
            if _es_linea_valor(ln):
                continue
            cuerpo.append(ln)
        desc_raw = "\n".join(cuerpo).strip()

        if (not desc_raw) and _nombre_beneficio_parece_descripcion(nombre):
            tema = _tema_beneficio_carga(nombre)
            if tema and tema in _ETIQUETAS_TEMA_BENEFICIO:
                desc_raw = nombre
                nombre = _ETIQUETAS_TEMA_BENEFICIO[tema]

        desc = _descripcion_requisito_limpia(nombre, desc_raw)
        if desc and _norm_nombre(desc) == _norm_nombre(nombre):
            desc = ""
        if desc:
            desc = re.sub(
                r"(?im)^\s*valor\s*:\s*[\d.,]+\s*[A-Za-z$€]*\s*$",
                "",
                desc,
            ).strip()
            desc = re.sub(
                r"(?i)\s*valor\s*:\s*[\d.,]+\s*[A-Za-z$€]*\s*$",
                "",
                desc,
            ).strip()

        tema = _tema_beneficio_carga(nombre, desc)
        if tema == "junk_money":
            continue
        if tema and _nombre_beneficio_parece_descripcion(nombre):
            if not desc:
                desc = _descripcion_requisito_limpia(
                    _ETIQUETAS_TEMA_BENEFICIO.get(tema, nombre), nombre
                )
            nombre = _ETIQUETAS_TEMA_BENEFICIO.get(
                tema, _nombre_item_limpio(nombre)[:80]
            )

        clave = _clave_dedupe_beneficio(tipo, nombre, desc)
        if not clave or clave.endswith(":") or clave in vistos:
            continue

        valor = None
        moneda = None
        m = re.search(r"valor\s*:\s*([\d.,]+)\s*([A-Za-z$€]*)", bloque, re.I)
        if not m:
            m = re.search(
                r"(?m)^\s*([\d.,]+)\s*(usd|us\$|\$|eur|cop)\s*$",
                bloque,
                re.I,
            )
        if m:
            try:
                valor = float(m.group(1).replace(",", ""))
            except ValueError:
                valor = None
            moneda = (m.group(2) or "").strip() or None
            if moneda == "$":
                moneda = "USD"

        vistos.add(clave)
        items.append(
            {
                "nombre": nombre[:160],
                "descripcion": desc or None,
                "tipo": tipo,
                "valor": valor,
                "moneda": moneda,
                "requiere_validacion_humana": tipo == "bono",
                "orden": len(items) + 1,
            }
        )
    return items



def _inferir_meta_faq(pregunta: str) -> Dict[str, Any]:
    """Heurística ligera de categoria/intencion/keywords para carga masiva."""
    n = _norm_nombre(pregunta)
    meta: Dict[str, Any] = {"categoria": "general", "intencion": None}
    if any(x in n for x in ("monetiz", "ganar dinero", "ingresos", "ganancia")):
        meta.update(
            {
                "categoria": "monetizacion",
                "intencion": "como_monetizar",
                "palabras_clave": [
                    "monetizar",
                    "monetizacion",
                    "ganar dinero",
                    "ingresos",
                    "ganancias",
                ],
            }
        )
    elif any(x in n for x in ("regalo", "gift")):
        meta.update(
            {
                "categoria": "monetizacion",
                "intencion": "que_son_regalos",
                "palabras_clave": ["regalos", "regalo", "gifts", "gift"],
            }
        )
    elif "diamant" in n or "diamond" in n:
        meta.update(
            {
                "categoria": "monetizacion",
                "intencion": "que_son_diamantes",
                "palabras_clave": ["diamantes", "diamante"],
            }
        )
    elif any(x in n for x in ("iniciar live", "no puedo", "no deja", "error live")):
        meta.update(
            {
                "categoria": "soporte",
                "intencion": "problema_iniciar_live",
                "palabras_clave": ["iniciar", "live", "error", "bloqueo"],
            }
        )
    elif any(x in n for x in ("cobra", "cobran", "costo", "gratis")):
        meta.update(
            {
                "categoria": "agencia",
                "intencion": "cobro_agencia",
                "palabras_clave": ["cobra", "cobran", "costo", "gratis"],
            }
        )
    elif "experiencia" in n:
        meta.update(
            {
                "categoria": "requisitos",
                "intencion": "experiencia",
                "palabras_clave": ["experiencia", "principiante"],
            }
        )
    return meta


def _parse_faq(texto: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    raw = (texto or "").strip()
    if not raw:
        return items

    # Bloques Pregunta:/Respuesta:
    bloques = re.split(r"(?=(?:Pregunta|Q)\s*[:\-–])", raw, flags=re.I)
    for bloque in bloques:
        bloque = bloque.strip()
        if not bloque:
            continue
        pm = re.search(r"(?:Pregunta|Q)\s*[:\-–]\s*(.+)", bloque, re.I)
        rm = re.search(r"(?:Respuesta|R)\s*[:\-–]\s*(.+)", bloque, re.I | re.S)
        if pm and rm:
            pregunta = pm.group(1).splitlines()[0].strip()
            respuesta = rm.group(1).strip()
            if pregunta and respuesta:
                item = {
                    "pregunta": pregunta[:2000],
                    "respuesta": respuesta[:8000],
                }
                item.update(_inferir_meta_faq(pregunta))
                items.append(item)
    if items:
        return items

    # Fallback: líneas "pregunta? respuesta"
    for linea in _lineas_bullet(raw):
        if "?" in linea:
            preg, _, resp = linea.partition("?")
            pregunta = (preg + "?").strip()
            respuesta = resp.strip()
            if pregunta and respuesta:
                item = {
                    "pregunta": pregunta[:2000],
                    "respuesta": respuesta[:8000],
                }
                item.update(_inferir_meta_faq(pregunta))
                items.append(item)
    return items


def _parse_proceso(texto: str) -> List[Dict[str, Any]]:
    items = []
    vistos: set = set()
    for linea in _lineas_bullet(texto):
        while linea.startswith("-"):
            linea = linea.lstrip("- ").strip()
        cruda = linea
        nombre = cruda
        mensaje = ""
        if "—" in cruda:
            nombre, mensaje = cruda.split("—", 1)
        elif " - " in cruda and not cruda.strip().startswith("-"):
            # Solo partir en " - " si no es dump anidado
            if not _parece_dump_lista(cruda):
                nombre, mensaje = cruda.split(" - ", 1)
        nombre = _nombre_item_limpio(nombre)
        mensaje = _texto_corto_limpio(mensaje, max_len=300)
        if mensaje and _norm_nombre(mensaje) == _norm_nombre(nombre):
            mensaje = ""
        if not nombre:
            continue
        if any(_nombres_paso_compatibles(nombre, v) for v in vistos):
            continue
        vistos.add(nombre)
        texto_paso = f"{nombre}" + (f" — {mensaje}" if mensaje else "")
        accion = "informar"
        baja = texto_paso.lower()
        if "solicitud" in baja or "enlace" in baja:
            accion = "enviar_enlace"
        elif "evidencia" in baja:
            accion = "solicitar_evidencias"
        elif "manager" in baja or "humano" in baja or "equipo" in baja or "revis" in baja:
            accion = "transferir_humano"
        elif "interes" in baja or "confirm" in baja:
            accion = "confirmar_interes"
        items.append(
            {
                "orden": len(items) + 1,
                "nombre": nombre[:160],
                "descripcion": mensaje or nombre,
                "accion": accion,
                "mensaje": mensaje or nombre,
                "requiere_humano": accion == "transferir_humano",
            }
        )
    return items


def _parse_enlaces(texto: str) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    recursos: List[Dict[str, Any]] = []
    equipo: Optional[str] = None
    raw = texto or ""
    # Equipo
    m_eq = re.search(
        r"(?:equipo|contacto)\s*(?:o\s*contacto)?\s*[:\-–]\s*(.+)",
        raw,
        re.I,
    )
    if m_eq:
        equipo = m_eq.group(1).splitlines()[0].strip()[:120]

    boton = "Completar solicitud"
    m_btn = re.search(r"texto\s+del\s+bot[oó]n\s*[:\-–]\s*(.+)", raw, re.I)
    if m_btn:
        boton = m_btn.group(1).splitlines()[0].strip()[:120]

    urls = _URL_RE.findall(raw)
    for i, url in enumerate(urls):
        u = _url_real(url)
        if not u:
            continue
        tipo = "solicitud" if i == 0 or "solicitud" in raw.lower() else "otro"
        recursos.append(
            {
                "tipo": tipo,
                "nombre": "Solicitud principal" if tipo == "solicitud" else f"Enlace {i + 1}",
                "url": u,
                "texto_boton": boton if tipo == "solicitud" else "Abrir enlace",
            }
        )
    return recursos, equipo


def _analizar_heuristico(payload: Dict[str, Any]) -> Dict[str, Any]:
    advertencias: List[str] = []
    requisitos = _parse_requisitos(payload.get("requisitos_texto") or "")
    beneficios = _parse_beneficios(payload.get("beneficios_texto") or "", "beneficio")
    bonos = _parse_beneficios(payload.get("bonos_texto") or "", "bono")
    faq = _parse_faq(payload.get("faq_texto") or "")
    proceso = _parse_proceso(payload.get("proceso_ingreso_texto") or "")
    recursos, equipo = _parse_enlaces(payload.get("enlaces_contacto_texto") or "")

    # Orden solicitud antes de evidencias salvo texto explícito contrario
    if proceso:
        idxs = {p["accion"]: i for i, p in enumerate(proceso)}
        if "enviar_enlace" in idxs and "solicitar_evidencias" in idxs:
            if idxs["enviar_enlace"] > idxs["solicitar_evidencias"]:
                if not re.search(
                    r"evidenc.*antes.*solicitud|primero.*evidenc",
                    payload.get("proceso_ingreso_texto") or "",
                    re.I,
                ):
                    # reordenar
                    proceso.sort(
                        key=lambda p: {
                            "confirmar_interes": 0,
                            "enviar_enlace": 1,
                            "solicitar_evidencias": 2,
                            "transferir_humano": 3,
                        }.get(p["accion"], 9)
                    )
                    for i, p in enumerate(proceso, start=1):
                        p["orden"] = i
                    advertencias.append(
                        "Se mantuvo el orden predeterminado: solicitud antes de evidencias."
                    )

    if (payload.get("bonos_texto") or "").strip() and not bonos:
        advertencias.append("No se pudieron organizar bonos; revisa el formato.")
    if (payload.get("faq_texto") or "").strip() and not faq:
        advertencias.append(
            "No se detectaron pares Pregunta/Respuesta en las FAQ."
        )

    general_out: Dict[str, Any] = {
        "nombre_asistente": str(payload.get("nombre_asistente") or "").strip(),
        "tono": payload.get("tono")
        if payload.get("tono") in {"profesional", "cercano", "juvenil"}
        else "cercano",
    }
    # Solo incluir presentaciones si vinieron en el payload (no inventar vacíos).
    for clave in (
        "presentacion_inicial",
        "presentacion_informativo",
        "presentacion_inteligente",
    ):
        if clave in payload:
            general_out[clave] = str(payload.get(clave) or "").strip()

    return {
        "general": general_out,
        "requisitos": requisitos,
        "beneficios": beneficios,
        "bonos": bonos,
        "faq": faq,
        "proceso_ingreso": proceso,
        "recursos": recursos,
        "contacto_humano": {"equipo_destino": equipo or ""},
        "advertencias": advertencias,
        "origen_analisis": "heuristico",
    }


def _analizar_con_openai(payload: Dict[str, Any], base: Dict[str, Any]) -> Dict[str, Any]:
    """Refina la propuesta con OpenAI; si falla, devuelve la base heurística."""
    try:
        from chatbot_conversacional_agent_factory import openai_api_key, openai_configurado
    except Exception:
        return base
    if not openai_configurado():
        return base

    try:
        from openai import OpenAI

        client = OpenAI(api_key=openai_api_key())
        system = (
            "Organizas información de una agencia para un chatbot de captación. "
            "NO inventes montos, bonos, requisitos ni enlaces. "
            "Si falta un dato, omítelo. Conserva cifras y condiciones. "
            "Separa beneficios de bonos. "
            "En faq: cada ítem debe ser una pregunta real de aspirantes con su "
            "respuesta; incluye categoria, intencion y palabras_clave relevantes "
            "(sin mezclar temas distintos, p. ej. monetizar ≠ no poder iniciar LIVE). "
            "NO conviertas el proceso de ingreso en FAQs. "
            "Responde SOLO JSON válido con las claves: "
            "requisitos, beneficios, bonos, faq, proceso_ingreso, recursos, "
            "contacto_humano, advertencias."
        )
        user = json.dumps(
            {
                "textos": {
                    "requisitos_texto": payload.get("requisitos_texto"),
                    "beneficios_texto": payload.get("beneficios_texto"),
                    "bonos_texto": payload.get("bonos_texto"),
                    "faq_texto": payload.get("faq_texto"),
                    "proceso_ingreso_texto": payload.get("proceso_ingreso_texto"),
                    "enlaces_contacto_texto": payload.get("enlaces_contacto_texto"),
                },
                "propuesta_base": base,
            },
            ensure_ascii=False,
        )[:12000]
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            max_tokens=2500,
        )
        content = resp.choices[0].message.content or "{}"
        data = json.loads(content)
        # Fusionar conservadoramente: solo reemplazar listas no vacías del modelo
        out = dict(base)
        for key in (
            "requisitos",
            "beneficios",
            "bonos",
            "faq",
            "proceso_ingreso",
            "recursos",
        ):
            if isinstance(data.get(key), list) and data[key]:
                out[key] = data[key]
        if isinstance(data.get("contacto_humano"), dict):
            out["contacto_humano"] = {
                **(out.get("contacto_humano") or {}),
                **data["contacto_humano"],
            }
        if isinstance(data.get("advertencias"), list):
            out["advertencias"] = list(
                dict.fromkeys((out.get("advertencias") or []) + data["advertencias"])
            )
        out["origen_analisis"] = "openai"
        return out
    except Exception as exc:  # noqa: BLE001
        logger.warning("[CARGA_INFO] OpenAI falló, se usa heurística: %s", exc)
        base = dict(base)
        base.setdefault("advertencias", []).append(
            "La organización automática usó el analizador local."
        )
        return base


def analizar_informacion(
    agencia_id: int,
    chatbot_configuracion_id: int,
    payload: Dict[str, Any],
    *,
    cur=None,
    usar_ia: Optional[bool] = None,
) -> Dict[str, Any]:
    """Dry-run: no escribe en BD.

    usar_ia=False evita OpenAI (segundos en lugar de ~1 min).
    Si no se pasa, respeta payload['usar_ia'] (default True).
    """
    with db._cursor(cur) as c:
        db._exige_configuracion(agencia_id, chatbot_configuracion_id, cur=c)

    if usar_ia is None:
        usar_ia = bool((payload or {}).get("usar_ia", True))

    base = _analizar_heuristico(payload)
    propuesta = _analizar_con_openai(payload, base) if usar_ia else base
    if not usar_ia:
        propuesta = dict(propuesta or {})
        propuesta["origen_analisis"] = "heuristico"

    # Adjuntar IDs existentes por coincidencia (sin guardar)
    with db._cursor(cur) as c:
        propuesta = _anotar_ids_existentes(
            agencia_id, chatbot_configuracion_id, propuesta, cur=c
        )

    logger.info(
        "[CARGA_INFO] agencia_id=%s chatbot_configuracion_id=%s accion=analizar "
        "requisitos=%s beneficios=%s bonos=%s faq=%s pasos=%s recursos=%s "
        "origen=%s resultado=ok",
        agencia_id,
        chatbot_configuracion_id,
        len(propuesta.get("requisitos") or []),
        len(propuesta.get("beneficios") or []),
        len(propuesta.get("bonos") or []),
        len(propuesta.get("faq") or []),
        len(propuesta.get("proceso_ingreso") or []),
        len(propuesta.get("recursos") or []),
        propuesta.get("origen_analisis"),
    )
    return propuesta


def _anotar_ids_existentes(
    agencia_id: int,
    cfg_id: int,
    propuesta: Dict[str, Any],
    *,
    cur,
) -> Dict[str, Any]:
    requisitos = db.listar_requisitos(
        agencia_id, chatbot_configuracion_id=cfg_id, incluir_globales=False, solo_activos=False, cur=cur
    )
    beneficios = db.listar_beneficios(
        agencia_id, chatbot_configuracion_id=cfg_id, incluir_globales=False, solo_activos=False, cur=cur
    )
    faqs = db.listar_faqs(
        agencia_id, chatbot_configuracion_id=cfg_id, incluir_globales=False, solo_activos=False, cur=cur
    )
    recursos = db.listar_recursos(
        agencia_id, chatbot_configuracion_id=cfg_id, incluir_globales=False, solo_activos=False, cur=cur
    )

    req_map = {_norm_nombre(r.get("nombre")): r for r in requisitos}
    for item in propuesta.get("requisitos") or []:
        hit = req_map.get(_norm_nombre(item.get("nombre")))
        if hit:
            item["id"] = hit.get("id")
            item["codigo"] = hit.get("codigo")

    ben_map = {
        (str(b.get("tipo") or "").lower(), _norm_nombre(b.get("nombre"))): b
        for b in beneficios
    }
    for tipo_key, lista_key in (("beneficio", "beneficios"), ("bono", "bonos")):
        for item in propuesta.get(lista_key) or []:
            tipo = str(item.get("tipo") or tipo_key).lower()
            hit = ben_map.get((tipo, _norm_nombre(item.get("nombre"))))
            if hit:
                item["id"] = hit.get("id")
                item["codigo"] = hit.get("codigo")

    faq_map = {_norm_nombre(f.get("pregunta")): f for f in faqs}
    for item in propuesta.get("faq") or []:
        hit = faq_map.get(_norm_nombre(item.get("pregunta")))
        if hit:
            item["id"] = hit.get("id")
            item["codigo"] = hit.get("codigo")

    rec_map = {
        (str(r.get("tipo") or "").lower(), _norm_nombre(r.get("url_template"))): r
        for r in recursos
    }
    for item in propuesta.get("recursos") or []:
        hit = rec_map.get(
            (str(item.get("tipo") or "").lower(), _norm_nombre(item.get("url")))
        )
        if hit:
            item["id"] = hit.get("id")
            item["codigo"] = hit.get("codigo")

    # Pasos del flujo de conversión: anotar id/codigo para guardar idempotente.
    flujos = db.listar_flujos(
        agencia_id,
        chatbot_configuracion_id=cfg_id,
        tipo_flujo="conversion",
        solo_activos=False,
        cur=cur,
    )
    flujo = next(
        (f for f in flujos if f.get("codigo") == CODIGO_FLUJO_CONVERSION),
        flujos[0] if flujos else None,
    )
    if flujo and flujo.get("id"):
        pasos_db = db.listar_flujo_pasos(
            agencia_id, int(flujo["id"]), solo_activos=False, cur=cur
        )
        ids_usados: set = set()
        antes = len(propuesta.get("proceso_ingreso") or [])
        for item in propuesta.get("proceso_ingreso") or []:
            if not isinstance(item, dict):
                continue
            hit, _via = _resolver_paso_existente(
                item, pasos_db, ids_ya_usados=ids_usados
            )
            if hit:
                item["id"] = hit.get("id")
                item["codigo"] = hit.get("codigo")
                # Conservar orden real del flujo (no el del textarea/IA).
                if hit.get("orden") is not None:
                    item["orden"] = hit.get("orden")
                ids_usados.add(int(hit["id"]))
        propuesta["proceso_ingreso"] = _dedupe_proceso_ingreso(
            propuesta.get("proceso_ingreso") or []
        )
        despues = len(propuesta.get("proceso_ingreso") or [])
        ids_out = [
            p.get("id")
            for p in (propuesta.get("proceso_ingreso") or [])
            if isinstance(p, dict) and p.get("id")
        ]
        codigos_out = [
            p.get("codigo")
            for p in (propuesta.get("proceso_ingreso") or [])
            if isinstance(p, dict) and p.get("codigo")
        ]
        logger.info(
            "[CARGA_INFO_ANALISIS_PASOS] antes=%s despues=%s ids=%s "
            "codigos=%s duplicados_eliminados=%s",
            antes,
            despues,
            ids_out,
            codigos_out,
            max(0, antes - despues),
        )

    return propuesta


def _dedupe_proceso_ingreso(pasos: List[Any]) -> List[Dict[str, Any]]:
    """Una sola representación por paso (id → codigo → nombre compatible)."""
    out: List[Dict[str, Any]] = []
    vistos_id: set = set()
    vistos_cod: set = set()
    nombres_keep: List[str] = []

    for item in pasos or []:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        if not nombre:
            continue

        iid = item.get("id")
        if iid is not None:
            try:
                iid_i = int(iid)
            except (TypeError, ValueError):
                iid_i = None
            if iid_i is not None:
                if iid_i in vistos_id:
                    continue
                vistos_id.add(iid_i)
                out.append(item)
                nombres_keep.append(nombre)
                cod = str(item.get("codigo") or "").strip().lower()
                if cod:
                    vistos_cod.add(cod)
                continue

        cod = str(item.get("codigo") or "").strip().lower()
        if cod:
            if cod in vistos_cod:
                continue
            vistos_cod.add(cod)
            out.append(item)
            nombres_keep.append(nombre)
            continue

        if any(_nombres_paso_compatibles(nombre, n) for n in nombres_keep):
            continue
        nombres_keep.append(nombre)
        out.append(item)

    return out


def _nombres_paso_compatibles(a: Any, b: Any) -> bool:
    na = _norm_nombre(_nombre_item_limpio(str(a or "")))
    nb = _norm_nombre(_nombre_item_limpio(str(b or "")))
    if not na or not nb:
        return False
    if na == nb:
        return True
    # Prefijo significativo: "explicar beneficios" ≈ "explicar beneficios y bonos…"
    corto, largo = (na, nb) if len(na) <= len(nb) else (nb, na)
    if len(corto) >= 12 and largo.startswith(corto):
        return True
    return False


def _resolver_paso_existente(
    item: Dict[str, Any],
    pasos_db: List[Dict[str, Any]],
    *,
    ids_ya_usados: Optional[set] = None,
) -> Tuple[Optional[Dict[str, Any]], str]:
    """
    Identifica el paso existente. Prioridad:
    1) id (mismo flujo)
    2) codigo
    3) nombre (exacto o prefijo significativo)
    4) orden (protección final anti-colisión)
    """
    usados = set(ids_ya_usados or set())
    candidatos = [p for p in pasos_db if p.get("id") and int(p["id"]) not in usados]

    item_id = item.get("id")
    if item_id:
        for p in candidatos:
            if int(p["id"]) == int(item_id):
                return p, "id"

    codigo = str(item.get("codigo") or "").strip().lower()
    if codigo:
        for p in candidatos:
            if str(p.get("codigo") or "").strip().lower() == codigo:
                return p, "codigo"

    nombre = str(item.get("nombre") or "").strip()
    if nombre:
        for p in candidatos:
            if _nombres_paso_compatibles(nombre, p.get("nombre")):
                return p, "nombre"

    orden = item.get("orden")
    if orden is not None and str(orden).strip() != "":
        try:
            orden_i = int(orden)
        except (TypeError, ValueError):
            orden_i = None
        if orden_i is not None:
            for p in candidatos:
                if int(p.get("orden") or -1) == orden_i:
                    return p, "orden"

    return None, ""


def _validar_ordenes_proceso(pasos: List[Any]) -> None:
    vistos: set = set()
    for item in pasos:
        if not isinstance(item, dict):
            continue
        if item.get("orden") is None or str(item.get("orden")).strip() == "":
            continue
        try:
            orden = int(item.get("orden"))
        except (TypeError, ValueError):
            continue
        if orden in vistos:
            raise ConversacionalError(
                f"El proceso de ingreso trae dos pasos con el mismo orden ({orden})."
            )
        vistos.add(orden)


# ---------------------------------------------------------------------------
# Guardado transaccional
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Persistencia datos generales (bienvenida / tono)
# ---------------------------------------------------------------------------


def _inicio_texto(valor: Any, n: int = 80) -> str:
    t = str(valor or "")
    if not t:
        return ""
    return t[:n].replace("\n", "\\n")


def _log_presentacion_save(
    *,
    agencia_id: int,
    config_id: int,
    campos_escritos: Dict[str, Any],
    raw_post: Dict[str, Any],
    origen: str,
) -> None:
    for campo in (
        "presentacion_inicial",
        "presentacion_informativo",
        "presentacion_inteligente",
    ):
        if campo not in campos_escritos:
            continue
        val = campos_escritos.get(campo)
        logger.info(
            "[CARGA_INFO_PRESENTACION_SAVE] origen=%s agencia_id=%s config_id=%s "
            "campo=%s chars=%s valor_inicio=%r",
            origen,
            agencia_id,
            config_id,
            campo,
            len(str(val or "")),
            _inicio_texto(val),
        )
    logger.info(
        "[CARGA_INFO_PRESENTACION_SAVE] origen=%s agencia_id=%s config_id=%s "
        "asistente_id=%s columnas_bd=%s "
        "raw_inicial_chars=%s raw_informativo_chars=%s raw_inteligente_chars=%s "
        "raw_inicial_inicio=%r raw_informativo_inicio=%r raw_inteligente=%r",
        origen,
        agencia_id,
        config_id,
        raw_post.get("asistente_id"),
        raw_post.get("columnas_presentacion_en_bd"),
        len(str(raw_post.get("presentacion_inicial") or "")),
        len(str(raw_post.get("presentacion_informativo") or ""))
        if raw_post.get("presentacion_informativo") != "__COLUMNA_AUSENTE__"
        else -1,
        len(str(raw_post.get("presentacion_inteligente") or ""))
        if raw_post.get("presentacion_inteligente") != "__COLUMNA_AUSENTE__"
        else -1,
        _inicio_texto(raw_post.get("presentacion_inicial")),
        _inicio_texto(raw_post.get("presentacion_informativo"))
        if raw_post.get("presentacion_informativo") != "__COLUMNA_AUSENTE__"
        else "__COLUMNA_AUSENTE__",
        _inicio_texto(raw_post.get("presentacion_inteligente"))
        if raw_post.get("presentacion_inteligente") != "__COLUMNA_AUSENTE__"
        else "__COLUMNA_AUSENTE__",
    )


def _log_presentacion_get(
    *,
    agencia_id: int,
    config_id: int,
    fuente: str,
    asistente: Optional[Dict[str, Any]],
    general: Optional[Dict[str, Any]] = None,
) -> None:
    a = asistente or {}
    g = general or {}
    pi = a.get("presentacion_inicial")
    pf = a.get("presentacion_informativo")
    pt = a.get("presentacion_inteligente")
    logger.info(
        "[CARGA_INFO_PRESENTACION_GET] fuente=%s agencia_id=%s config_id=%s "
        "asistente_id=%s presentacion_inicial_chars=%s "
        "presentacion_informativo_chars=%s presentacion_inteligente_chars=%s "
        "presentacion_inicial_inicio=%r presentacion_informativo_inicio=%r "
        "general_informativo_chars=%s general_inicial_chars=%s "
        "general_informativo_inicio=%r",
        fuente,
        agencia_id,
        config_id,
        a.get("id"),
        len(str(pi or "")),
        len(str(pf or "")),
        len(str(pt or "")),
        _inicio_texto(pi),
        _inicio_texto(pf),
        len(str(g.get("presentacion_informativo") or "")),
        len(str(g.get("presentacion_inicial") or "")),
        _inicio_texto(g.get("presentacion_informativo") or g.get("presentacion_inicial")),
    )


def persistir_datos_generales_asistente(
    agencia_id: int,
    chatbot_configuracion_id: int,
    general: Optional[Dict[str, Any]],
    *,
    origen: str = "persistir_datos_generales",
) -> Dict[str, Any]:
    """
    Guarda nombre / presentaciones / tono en transacción propia.

    No limpia presentaciones si llegan vacías: así un guardado parcial
    de catálogos no borra la bienvenida ya configurada.
    """
    general = dict(general or {})
    campos_a: Dict[str, Any] = {}
    if general.get("nombre_asistente"):
        campos_a["nombre_asistente"] = str(general["nombre_asistente"]).strip()[:120]

    def _tomar_si_enviada(clave: str) -> Optional[str]:
        """Solo el campo pedido; no usa otros como alias de escritura."""
        if clave not in general:
            return None
        nueva = str(general.get(clave) or "").strip()
        return nueva[:4000] if nueva else None

    # Independientes: no copiar un valor sobre los otros.
    info = _tomar_si_enviada("presentacion_informativo")
    if info:
        campos_a["presentacion_informativo"] = info
    intel = _tomar_si_enviada("presentacion_inteligente")
    if intel:
        campos_a["presentacion_inteligente"] = intel
    legacy = _tomar_si_enviada("presentacion_inicial")
    if legacy:
        campos_a["presentacion_inicial"] = legacy

    if (
        "presentacion_informativo" in general
        and not info
        and legacy
        and origen.startswith("PUT_")
    ):
        logger.warning(
            "[CARGA_INFO_PRESENTACION_SAVE] origen=%s agencia_id=%s config_id=%s "
            "ALERTA: presentacion_informativo vacío pero presentacion_inicial "
            "tiene %s chars. El cliente está mandando el texto al campo legacy. "
            "No se copia automáticamente a presentacion_informativo.",
            origen,
            agencia_id,
            chatbot_configuracion_id,
            len(legacy),
        )
    if "descripcion_agencia" in general:
        desc = str(general.get("descripcion_agencia") or "").strip()
        if desc:
            campos_a["descripcion_agencia"] = desc[:4000]

    if general.get("formato_respuestas_informativas") in {
        "lista",
        "texto_breve",
        "automatico",
    }:
        campos_a["formato_respuestas_informativas"] = general[
            "formato_respuestas_informativas"
        ]

    if general.get("tono") in {"profesional", "cercano", "juvenil"}:
        campos_a["tono"] = general["tono"]

    logger.info(
        "[CARGA_INFO_PRESENTACION_SAVE] origen=%s fase=entrada agencia_id=%s "
        "config_id=%s keys_general=%s info_chars=%s intel_chars=%s legacy_chars=%s "
        "campos_a=%s",
        origen,
        agencia_id,
        chatbot_configuracion_id,
        sorted(general.keys()),
        len(info or ""),
        len(intel or ""),
        len(legacy or ""),
        sorted(campos_a.keys()),
    )

    def _chars(asistente_row: Optional[Dict[str, Any]]) -> int:
        a = asistente_row or {}
        return max(
            len(str(a.get("presentacion_informativo") or "")),
            len(str(a.get("presentacion_inteligente") or "")),
            len(str(a.get("presentacion_inicial") or "")),
        )

    if not campos_a:
        asistente = db.obtener_asistente_por_config(agencia_id, chatbot_configuracion_id)
        raw = db.leer_presentaciones_raw(agencia_id, chatbot_configuracion_id)
        _log_presentacion_save(
            agencia_id=agencia_id,
            config_id=chatbot_configuracion_id,
            campos_escritos={},
            raw_post=raw,
            origen=f"{origen}:sin_cambios",
        )
        return {
            "ok": True,
            "asistente": asistente,
            "actualizado": False,
            "presentacion_chars": _chars(asistente),
            "raw_presentacion": raw,
        }

    asistente = db.obtener_asistente_por_config(agencia_id, chatbot_configuracion_id)
    if asistente:
        actualizado = db.actualizar_asistente(
            agencia_id, int(asistente["id"]), campos_a
        )
    else:
        campos_a.setdefault(
            "nombre_asistente",
            str(general.get("nombre_asistente") or "Asistente virtual")[:120],
        )
        campos_a.update(
            {
                "activo": True,
                "declarar_asistente_virtual": True,
                "modo_informativo_activo": True,
                "modo_conversion_activo": True,
                "modo_predeterminado": "conversion",
            }
        )
        actualizado = db.upsert_asistente(
            agencia_id, chatbot_configuracion_id, campos_a
        )

    # Relectura en conexión nueva (post-commit) para validar persistencia real.
    raw = db.leer_presentaciones_raw(agencia_id, chatbot_configuracion_id)
    _log_presentacion_save(
        agencia_id=agencia_id,
        config_id=chatbot_configuracion_id,
        campos_escritos={
            k: campos_a[k]
            for k in (
                "presentacion_inicial",
                "presentacion_informativo",
                "presentacion_inteligente",
            )
            if k in campos_a
        },
        raw_post=raw,
        origen=origen,
    )

    logger.info(
        "[CARGA_INFO] generales agencia_id=%s config_id=%s "
        "presentacion_chars=%s campos=%s",
        agencia_id,
        chatbot_configuracion_id,
        _chars(actualizado),
        sorted(campos_a.keys()),
    )
    return {
        "ok": True,
        "asistente": actualizado,
        "actualizado": True,
        "presentacion_chars": _chars(actualizado),
        "raw_presentacion": raw,
    }


def _scrub_requisitos_duplicados(
    agencia_id: int,
    chatbot_configuracion_id: int,
    existentes: Dict[int, Dict[str, Any]],
    *,
    ids_conservar: Optional[set] = None,
    cur=None,
) -> Dict[str, int]:
    """
    Colapsa requisitos fractal/duplicados por tema o nombre limpio.

    Conserva la mejor fila de cada grupo, limpia nombre/descripcion y
    desactiva el resto (incluidas filas cuyo nombre es solo basura).
    """
    ids_conservar = set(ids_conservar or set())
    grupos: Dict[str, List[Dict[str, Any]]] = {}
    basura: List[Dict[str, Any]] = []

    for row in existentes.values():
        nombre = str(row.get("nombre") or "")
        clave = _clave_dedupe_requisito(nombre)
        limpio = _nombre_item_limpio(nombre)
        if not limpio or clave == "n:":
            basura.append(row)
            continue
        grupos.setdefault(clave, []).append(row)

    desactivados = 0
    limpiados = 0

    for row in basura:
        rid = int(row["id"])
        if rid in ids_conservar:
            continue
        if row.get("activo") is not False:
            db.actualizar_requisito(agencia_id, rid, {"activo": False}, cur=cur)
            desactivados += 1

    for _clave, filas in grupos.items():
        if not filas:
            continue
        # Preferir ids tocados en este guardado si están en el grupo.
        keep = None
        for f in filas:
            if int(f["id"]) in ids_conservar:
                keep = f
                break
        if keep is None:
            keep = max(filas, key=_score_requisito_keeper)
        keep_id = int(keep["id"])

        nombre_limpio = _nombre_item_limpio(str(keep.get("nombre") or ""))
        desc_limpia = _descripcion_requisito_limpia(
            nombre_limpio, keep.get("descripcion")
        )
        patch: Dict[str, Any] = {}
        if nombre_limpio and nombre_limpio != str(keep.get("nombre") or "").strip():
            patch["nombre"] = nombre_limpio[:160]
        if (desc_limpia or None) != (str(keep.get("descripcion") or "").strip() or None):
            patch["descripcion"] = desc_limpia or None
        if keep.get("activo") is False:
            patch["activo"] = True
        if patch:
            db.actualizar_requisito(agencia_id, keep_id, patch, cur=cur)
            limpiados += 1
            existentes[keep_id] = {**keep, **patch}

        for f in filas:
            rid = int(f["id"])
            if rid == keep_id:
                continue
            if f.get("activo") is not False:
                db.actualizar_requisito(
                    agencia_id, rid, {"activo": False}, cur=cur
                )
                desactivados += 1

    if desactivados or limpiados:
        logger.info(
            "[CARGA_INFO_REQUISITOS_SCRUB] agencia_id=%s config_id=%s "
            "desactivados=%s limpiados=%s grupos=%s",
            agencia_id,
            chatbot_configuracion_id,
            desactivados,
            limpiados,
            len(grupos),
        )
    return {"desactivados": desactivados, "limpiados": limpiados}


def _scrub_beneficios_duplicados(
    agencia_id: int,
    chatbot_configuracion_id: int,
    existentes: Dict[int, Dict[str, Any]],
    *,
    ids_conservar: Optional[set] = None,
    cur=None,
) -> Dict[str, int]:
    """Colapsa beneficios/bonos duplicados por tema o nombre limpio."""
    ids_conservar = set(ids_conservar or set())
    grupos: Dict[str, List[Dict[str, Any]]] = {}
    basura: List[Dict[str, Any]] = []

    for row in existentes.values():
        nombre = str(row.get("nombre") or "")
        desc = _desc_beneficio_de_row(row)
        tipo = str(row.get("tipo") or "beneficio")
        clave = _clave_dedupe_beneficio(tipo, nombre, desc)
        limpio = _nombre_item_limpio(nombre)
        if (
            not limpio
            or clave.endswith(":")
            or clave == "tema:junk_money"
            or _parece_dump_lista(nombre)
        ):
            basura.append(row)
            continue
        grupos.setdefault(clave, []).append(row)

    desactivados = 0
    limpiados = 0

    for row in basura:
        rid = int(row["id"])
        if rid in ids_conservar:
            continue
        if row.get("activo") is not False:
            db.actualizar_beneficio(agencia_id, rid, {"activo": False}, cur=cur)
            desactivados += 1

    for clave, filas in grupos.items():
        if not filas:
            continue
        keep = None
        for f in filas:
            if int(f["id"]) in ids_conservar:
                keep = f
                break
        if keep is None:
            keep = max(filas, key=_score_beneficio_keeper)
        keep_id = int(keep["id"])

        nombre_raw = str(keep.get("nombre") or "")
        desc_raw = _desc_beneficio_de_row(keep)
        tema = _tema_beneficio_carga(nombre_raw, desc_raw)
        nombre_limpio = _nombre_item_limpio(nombre_raw)
        if tema and (
            _nombre_beneficio_parece_descripcion(nombre_limpio)
            or tema in _ETIQUETAS_TEMA_BENEFICIO
        ):
            etiqueta = _ETIQUETAS_TEMA_BENEFICIO.get(tema)
            if etiqueta and (
                _nombre_beneficio_parece_descripcion(nombre_limpio)
                or _norm_nombre(nombre_limpio) != _norm_nombre(etiqueta)
                and len(nombre_limpio) > len(etiqueta) + 10
            ):
                if not desc_raw or _norm_nombre(desc_raw) == _norm_nombre(nombre_limpio):
                    desc_raw = nombre_limpio if _nombre_beneficio_parece_descripcion(
                        nombre_limpio
                    ) else desc_raw
                nombre_limpio = etiqueta
        desc_limpia = _descripcion_requisito_limpia(nombre_limpio, desc_raw)
        if desc_limpia and _norm_nombre(desc_limpia) == _norm_nombre(nombre_limpio):
            desc_limpia = ""

        patch: Dict[str, Any] = {}
        if nombre_limpio and nombre_limpio != str(keep.get("nombre") or "").strip():
            patch["nombre"] = nombre_limpio[:160]
        if (desc_limpia or None) != (desc_raw or None):
            patch["descripcion_corta"] = desc_limpia or None
            patch["texto_autorizado"] = desc_limpia or None
        # Normalizar tipo canónico por tema
        tipo_keep = str(keep.get("tipo") or "beneficio").lower()
        if tema and tema.startswith("bono_") and tipo_keep not in {"bono", "incentivo"}:
            patch["tipo"] = "bono"
            patch["requiere_validacion_humana"] = True
        elif tema and not tema.startswith("bono_") and tipo_keep in {"bono", "incentivo"}:
            # Si no hay valor monetario, conviene beneficio narrativo
            if keep.get("valor") in (None, "", 0, 0.0):
                patch["tipo"] = "beneficio"
        if keep.get("activo") is False:
            patch["activo"] = True
        if patch:
            db.actualizar_beneficio(agencia_id, keep_id, patch, cur=cur)
            limpiados += 1
            existentes[keep_id] = {**keep, **patch}

        for f in filas:
            rid = int(f["id"])
            if rid == keep_id:
                continue
            if f.get("activo") is not False:
                db.actualizar_beneficio(
                    agencia_id, rid, {"activo": False}, cur=cur
                )
                desactivados += 1

    if desactivados or limpiados:
        logger.info(
            "[CARGA_INFO_BENEFICIOS_SCRUB] agencia_id=%s config_id=%s "
            "desactivados=%s limpiados=%s grupos=%s",
            agencia_id,
            chatbot_configuracion_id,
            desactivados,
            limpiados,
            len(grupos),
        )
    return {"desactivados": desactivados, "limpiados": limpiados}


def guardar_informacion_organizada(
    agencia_id: int,
    chatbot_configuracion_id: int,
    propuesta: Dict[str, Any],
    *,
    cur=None,
) -> Dict[str, Any]:
    datos = dict(propuesta or {})
    general = dict(datos.get("general") or {})
    creados = actualizados = 0
    advertencias: List[str] = []

    # 1) Bienvenida / tono / nombre: commit independiente (no se revierte
    #    si falla después el guardado de requisitos/flujo/enlaces).
    logger.info(
        "[CARGA_INFO_PRESENTACION_SAVE] origen=guardar_informacion_organizada "
        "fase=antes_persistir agencia_id=%s config_id=%s general_keys=%s "
        "general_info_chars=%s general_inicial_chars=%s general_info_inicio=%r",
        agencia_id,
        chatbot_configuracion_id,
        sorted(general.keys()),
        len(str(general.get("presentacion_informativo") or "")),
        len(str(general.get("presentacion_inicial") or "")),
        _inicio_texto(
            general.get("presentacion_informativo")
            or general.get("presentacion_inicial")
        ),
    )
    resumen_general = persistir_datos_generales_asistente(
        agencia_id,
        chatbot_configuracion_id,
        general,
        origen="guardar_informacion_organizada",
    )

    with db._cursor(cur) as c:
        db._exige_configuracion(agencia_id, chatbot_configuracion_id, cur=c)

        asistente = db.obtener_asistente_por_config(
            agencia_id, chatbot_configuracion_id, cur=c
        )
        # No volver a tocar presentacion_inicial aquí (ya quedó en paso 1).
        if not asistente:
            # Defensa: crear cascarón si el paso 1 no pudo (sin presentación vacía).
            db.upsert_asistente(
                agencia_id,
                chatbot_configuracion_id,
                {
                    "nombre_asistente": str(
                        general.get("nombre_asistente") or "Asistente virtual"
                    )[:120],
                    "activo": True,
                    "declarar_asistente_virtual": True,
                    "modo_informativo_activo": True,
                    "modo_conversion_activo": True,
                    "modo_predeterminado": "conversion",
                },
                cur=c,
            )
            creados += 1

        # Requisitos
        existentes_req = {
            int(r["id"]): r
            for r in db.listar_requisitos(
                agencia_id,
                chatbot_configuracion_id=chatbot_configuracion_id,
                incluir_globales=False,
                solo_activos=False,
                cur=c,
            )
            if r.get("id")
        }
        usados_cod = {
            str(r.get("codigo"))
            for r in existentes_req.values()
            if r.get("codigo")
        }
        # Preferir la mejor fila por tema/nombre (evita ecos fractal).
        por_clave: Dict[str, Dict[str, Any]] = {}
        for r in existentes_req.values():
            clave = _clave_dedupe_requisito(str(r.get("nombre") or ""))
            if not clave or clave == "n:":
                continue
            prev = por_clave.get(clave)
            if prev is None or _score_requisito_keeper(r) > _score_requisito_keeper(
                prev
            ):
                por_clave[clave] = r
        ids_tocados: set = set()

        for i, item in enumerate(datos.get("requisitos") or [], start=1):
            if not isinstance(item, dict):
                continue
            nombre = _nombre_item_limpio(str(item.get("nombre") or "").strip())
            if not nombre:
                continue
            campos = {
                "nombre": nombre[:160],
                "descripcion": _descripcion_requisito_limpia(
                    nombre, item.get("descripcion")
                )
                or None,
                "categoria": (
                    "obligatorio"
                    if str(item.get("categoria") or "obligatorio") == "obligatorio"
                    else "deseable"
                ),
                "bloquea_proceso": bool(item.get("bloquea_proceso", True)),
                "mensaje_si_no_cumple": item.get("mensaje_si_no_cumple"),
                "orden": int(item.get("orden") or i),
                "activo": True,
                "permitir_mencion_automatica": True,
            }
            item_id = item.get("id")
            clave = _clave_dedupe_requisito(nombre)
            hit = existentes_req.get(int(item_id)) if item_id else por_clave.get(clave)
            if hit:
                db.actualizar_requisito(agencia_id, int(hit["id"]), campos, cur=c)
                ids_tocados.add(int(hit["id"]))
                merged = {**hit, **campos, "id": hit["id"]}
                por_clave[clave] = merged
                existentes_req[int(hit["id"])] = merged
                actualizados += 1
            else:
                codigo = _codigo_unico(
                    item.get("codigo") or nombre,
                    usados_cod,
                    prefijo="requisito",
                    indice=i,
                )
                usados_cod.add(codigo)
                campos.update(
                    {
                        "codigo": codigo,
                        "chatbot_configuracion_id": chatbot_configuracion_id,
                        "tipo_dato": "texto",
                    }
                )
                creado = db.crear_requisito(agencia_id, campos, cur=c)
                if creado and creado.get("id"):
                    cid = int(creado["id"])
                    ids_tocados.add(cid)
                    por_clave[clave] = creado
                    existentes_req[cid] = creado
                creados += 1

        # Scrub: desactivar duplicados por tema/nombre y filas fractal.
        scrub = _scrub_requisitos_duplicados(
            agencia_id,
            chatbot_configuracion_id,
            existentes_req,
            ids_conservar=ids_tocados,
            cur=c,
        )
        actualizados += int(scrub.get("desactivados") or 0) + int(
            scrub.get("limpiados") or 0
        )
        # Beneficios / bonos
        existentes_ben = {
            int(b["id"]): b
            for b in db.listar_beneficios(
                agencia_id,
                chatbot_configuracion_id=chatbot_configuracion_id,
                incluir_globales=False,
                solo_activos=False,
                cur=c,
            )
            if b.get("id")
        }
        usados_b = {str(b.get("codigo")) for b in existentes_ben.values() if b.get("codigo")}
        por_clave_ben: Dict[str, Dict[str, Any]] = {}
        for b in existentes_ben.values():
            clave_b = _clave_dedupe_beneficio(
                str(b.get("tipo") or "beneficio"),
                str(b.get("nombre") or ""),
                _desc_beneficio_de_row(b),
            )
            if not clave_b or clave_b.endswith(":") or clave_b == "tema:junk_money":
                continue
            prev = por_clave_ben.get(clave_b)
            if prev is None or _score_beneficio_keeper(b) > _score_beneficio_keeper(
                prev
            ):
                por_clave_ben[clave_b] = b
        ids_ben_tocados: set = set()

        for tipo_default, clave in (("beneficio", "beneficios"), ("bono", "bonos")):
            for i, item in enumerate(datos.get(clave) or [], start=1):
                if not isinstance(item, dict):
                    continue
                nombre = _nombre_item_limpio(str(item.get("nombre") or "").strip())
                if not nombre:
                    continue
                tipo = str(item.get("tipo") or tipo_default).lower()
                desc_limpia = (
                    _descripcion_requisito_limpia(nombre, item.get("descripcion"))
                    or None
                )
                tema = _tema_beneficio_carga(nombre, desc_limpia)
                if tema == "junk_money":
                    continue
                if tema and _nombre_beneficio_parece_descripcion(nombre):
                    if not desc_limpia:
                        desc_limpia = _descripcion_requisito_limpia(
                            _ETIQUETAS_TEMA_BENEFICIO.get(tema, nombre), nombre
                        ) or None
                    nombre = _ETIQUETAS_TEMA_BENEFICIO.get(tema, nombre[:80])
                if tema and tema.startswith("bono_"):
                    tipo = "bono"
                elif tema and tipo in {"bono", "incentivo"} and item.get("valor") in (
                    None,
                    "",
                ):
                    # Narrativo tipificado como bono sin monto → beneficio
                    if tipo_default == "beneficio":
                        tipo = "beneficio"
                campos = {
                    "nombre": nombre[:160],
                    "descripcion_corta": desc_limpia,
                    "texto_autorizado": desc_limpia,
                    "tipo": tipo,
                    "activo": True,
                    "permitir_mencion_automatica": True,
                }
                if item.get("valor") not in (None, ""):
                    campos["valor"] = item.get("valor")
                if item.get("moneda"):
                    campos["moneda"] = str(item.get("moneda"))[:10]
                if tipo == "bono" or "requiere_validacion_humana" in item or "requiere_confirmacion_humana" in item:
                    campos["requiere_validacion_humana"] = bool(
                        item.get("requiere_validacion_humana")
                        or item.get("requiere_confirmacion_humana")
                        or tipo == "bono"
                    )
                item_id = item.get("id")
                clave_b = _clave_dedupe_beneficio(tipo, nombre, desc_limpia)
                hit = (
                    existentes_ben.get(int(item_id))
                    if item_id
                    else por_clave_ben.get(clave_b)
                )
                if hit:
                    db.actualizar_beneficio(agencia_id, int(hit["id"]), campos, cur=c)
                    ids_ben_tocados.add(int(hit["id"]))
                    merged = {**hit, **campos, "id": hit["id"]}
                    por_clave_ben[clave_b] = merged
                    existentes_ben[int(hit["id"])] = merged
                    actualizados += 1
                else:
                    codigo = _codigo_unico(
                        item.get("codigo") or nombre,
                        usados_b,
                        prefijo=tipo,
                        indice=i,
                    )
                    usados_b.add(codigo)
                    campos.update(
                        {
                            "codigo": codigo,
                            "chatbot_configuracion_id": chatbot_configuracion_id,
                        }
                    )
                    creado = db.crear_beneficio(agencia_id, campos, cur=c)
                    if creado and creado.get("id"):
                        cid = int(creado["id"])
                        ids_ben_tocados.add(cid)
                        por_clave_ben[clave_b] = creado
                        existentes_ben[cid] = creado
                    creados += 1

        scrub_ben = _scrub_beneficios_duplicados(
            agencia_id,
            chatbot_configuracion_id,
            existentes_ben,
            ids_conservar=ids_ben_tocados,
            cur=c,
        )
        actualizados += int(scrub_ben.get("desactivados") or 0) + int(
            scrub_ben.get("limpiados") or 0
        )

        # FAQ
        existentes_faq = {
            int(f["id"]): f
            for f in db.listar_faqs(
                agencia_id,
                chatbot_configuracion_id=chatbot_configuracion_id,
                incluir_globales=False,
                solo_activos=False,
                cur=c,
            )
            if f.get("id")
        }
        usados_f = {str(f.get("codigo")) for f in existentes_faq.values() if f.get("codigo")}
        por_preg = {_norm_nombre(f.get("pregunta")): f for f in existentes_faq.values()}

        for i, item in enumerate(datos.get("faq") or [], start=1):
            if not isinstance(item, dict):
                continue
            pregunta = str(item.get("pregunta") or "").strip()
            respuesta = str(item.get("respuesta") or "").strip()
            if not pregunta or not respuesta:
                continue
            campos = {
                "pregunta": pregunta[:2000],
                "respuesta_corta": str(
                    item.get("respuesta_corta") or respuesta
                )[:300],
                "respuesta_completa": str(
                    item.get("respuesta_completa") or respuesta
                )[:8000],
                "activo": True,
                "categoria": str(item.get("categoria") or "general")[:100],
                "prioridad": max(0, 100 - i),
            }
            if item.get("intencion"):
                campos["intencion"] = str(item.get("intencion"))[:100]
            # Palabras clave derivadas (sin stopwords / ruido).
            if not item.get("palabras_clave"):
                stop = {
                    "como",
                    "puedo",
                    "puede",
                    "que",
                    "cual",
                    "para",
                    "hacer",
                    "haciendo",
                    "live",
                    "lives",
                    "tiktok",
                }
                toks = [
                    t
                    for t in _norm_nombre(pregunta).split()
                    if len(t) > 3 and t not in stop
                ][:8]
                if toks:
                    campos["palabras_clave"] = toks
            elif item.get("palabras_clave") is not None:
                campos["palabras_clave"] = item.get("palabras_clave")
            item_id = item.get("id")
            hit = existentes_faq.get(int(item_id)) if item_id else por_preg.get(
                _norm_nombre(pregunta)
            )
            if hit:
                db.actualizar_faq(agencia_id, int(hit["id"]), campos, cur=c)
                actualizados += 1
            else:
                codigo = _codigo_unico(
                    item.get("codigo") or pregunta,
                    usados_f,
                    prefijo="faq",
                    indice=i,
                )
                usados_f.add(codigo)
                campos.update(
                    {
                        "codigo": codigo,
                        "chatbot_configuracion_id": chatbot_configuracion_id,
                        "fuente": "carga_informacion",
                    }
                )
                db.crear_faq(agencia_id, campos, cur=c)
                creados += 1

        # Recursos
        for item in datos.get("recursos") or []:
            if not isinstance(item, dict):
                continue
            url = _url_real(item.get("url") or item.get("url_template"))
            if not url:
                continue
            tipo = str(item.get("tipo") or "solicitud").lower()
            nombre = str(item.get("nombre") or "Enlace").strip()[:160]
            texto_boton = str(item.get("texto_boton") or "Abrir enlace").strip()[:120]
            codigo = (
                CODIGO_RECURSO_SOLICITUD
                if tipo == "solicitud"
                else _codigo_unico(nombre, set(), prefijo="recurso", indice=1)
            )
            existente = db.obtener_recurso_por_codigo(
                agencia_id,
                codigo,
                chatbot_configuracion_id=chatbot_configuracion_id,
                cur=c,
            )
            tipo = tipo if tipo in {
                "solicitud",
                "agendamiento",
                "privacidad",
                "terminos",
                "soporte",
                "whatsapp",
                "red_social",
                "instructivo",
                "otro",
            } else "otro"
            campos = {
                "nombre": nombre,
                "tipo": tipo,
                "url_template": url,
                "texto_boton": texto_boton,
                "activo": True,
                "abrir_externo": True,
                "chatbot_configuracion_id": chatbot_configuracion_id,
            }
            if existente:
                db.actualizar_recurso(agencia_id, int(existente["id"]), campos, cur=c)
                actualizados += 1
            else:
                campos["codigo"] = codigo
                try:
                    db.crear_recurso(agencia_id, campos, cur=c)
                    creados += 1
                except Exception as exc:  # noqa: BLE001
                    advertencias.append(f"No se pudo crear el enlace '{nombre}': {exc}")

        # Contacto humano (regla)
        # Origen histórico del bug: prioridad=10 (entero FAQ) → CheckViolation.
        # Semántica alineada a la plantilla existente (prioridad='alta').
        contacto = dict(datos.get("contacto_humano") or {})
        if contacto.get("equipo_destino"):
            reglas = db.listar_reglas_escalamiento(
                agencia_id,
                chatbot_configuracion_id=chatbot_configuracion_id,
                solo_activas=False,
                cur=c,
            )
            regla = next(
                (r for r in reglas if _es_regla_contacto_humano_equivalente(r)),
                None,
            )
            equipo = str(contacto.get("equipo_destino") or "").strip()[:120]
            prioridad = _prioridad_contacto_humano(contacto.get("prioridad"))
            if regla:
                # Reutilizar regla equivalente: no crear duplicado.
                campos_upd: Dict[str, Any] = {}
                if equipo and str(regla.get("equipo_destino") or "").strip() != equipo:
                    campos_upd["equipo_destino"] = equipo
                pri_actual = str(regla.get("prioridad") or "").strip().lower()
                if pri_actual not in PRIORIDADES_REGLA_ESCALAMIENTO:
                    campos_upd["prioridad"] = prioridad
                if not regla.get("activo", True):
                    campos_upd["activo"] = True
                if campos_upd:
                    db.actualizar_regla_escalamiento(
                        agencia_id,
                        int(regla["id"]),
                        _campos_regla_escalamiento_seguros(campos_upd),
                        cur=c,
                    )
                    actualizados += 1
                else:
                    advertencias.append(
                        "Se reutilizó la regla de contacto humano existente "
                        f"(id={regla.get('id')}); no se creó duplicado."
                    )
                logger.info(
                    "[CARGA_INFO] regla_contacto reutilizada id=%s prioridad=%s "
                    "actualizada=%s",
                    regla.get("id"),
                    pri_actual or prioridad,
                    bool(campos_upd),
                )
            else:
                campos_r = _campos_regla_escalamiento_seguros(
                    {
                        "chatbot_configuracion_id": chatbot_configuracion_id,
                        "evento": "solicitud_humano",
                        "descripcion": "Contacto humano (carga de información)",
                        "prioridad": prioridad,
                        "equipo_destino": equipo,
                        "canal_destino": "panel",
                        "mensaje_usuario": (
                            "Te conectaré con una persona del equipo para continuar."
                        ),
                        "estado_destino": "escalado_humano",
                        "activo": True,
                        "orden": 1,
                    }
                )
                assert campos_r["prioridad"] in PRIORIDADES_REGLA_ESCALAMIENTO
                db.crear_regla_escalamiento(agencia_id, campos_r, cur=c)
                creados += 1
                logger.info(
                    "[CARGA_INFO] regla_contacto creada prioridad=%s",
                    campos_r["prioridad"],
                )

        # Proceso → flujo conversión (crear faltantes, no borrar)
        pasos = datos.get("proceso_ingreso") or []
        if pasos:
            _validar_ordenes_proceso(pasos)
            flujos = db.listar_flujos(
                agencia_id,
                chatbot_configuracion_id=chatbot_configuracion_id,
                tipo_flujo="conversion",
                solo_activos=False,
                cur=c,
            )
            flujo = next(
                (f for f in flujos if f.get("codigo") == CODIGO_FLUJO_CONVERSION),
                flujos[0] if flujos else None,
            )
            if not flujo:
                flujo = db.crear_flujo(
                    agencia_id,
                    {
                        "chatbot_configuracion_id": chatbot_configuracion_id,
                        "codigo": CODIGO_FLUJO_CONVERSION,
                        "nombre": "Ingreso de aspirante",
                        "tipo_flujo": "conversion",
                        "descripcion": "Generado desde carga de información",
                        "estado_inicial": "inicio",
                        "estado_final": "escalado_humano",
                        "activo": True,
                    },
                    cur=c,
                )
                creados += 1
            flujo_id = int(flujo["id"])
            pasos_db = db.listar_flujo_pasos(
                agencia_id, flujo_id, solo_activos=False, cur=c
            )
            usados_cod_pasos = {
                str(p.get("codigo")) for p in pasos_db if p.get("codigo")
            }
            ids_pasos_tocados: set = set()

            for item in pasos:
                if not isinstance(item, dict):
                    continue
                nombre = str(item.get("nombre") or "").strip()
                if not nombre:
                    continue
                accion = str(item.get("accion") or "informar")
                # Mapear a tipo_accion válido
                validas = {
                    "informar",
                    "hacer_pregunta",
                    "confirmar_interes",
                    "explicar_requisitos",
                    "explicar_beneficios",
                    "explicar_bonos",
                    "enviar_enlace",
                    "agendar_live",
                    "solicitar_live",
                    "solicitar_evidencias",
                    "confirmar_evidencias",
                    "transferir_humano",
                    "esperar_respuesta",
                    "finalizar",
                }
                if accion not in validas:
                    accion = "informar"

                hit, via = _resolver_paso_existente(
                    item, pasos_db, ids_ya_usados=ids_pasos_tocados
                )
                msg_limpio = _texto_corto_limpio(
                    item.get("mensaje") or item.get("descripcion") or "",
                    max_len=500,
                )
                nombre_limpio = _nombre_item_limpio(nombre)
                if msg_limpio and _norm_nombre(msg_limpio) == _norm_nombre(nombre_limpio):
                    msg_limpio = ""
                # No tocar orden ni siguiente_paso_id: el motor avanza por orden.
                campos_p = {
                    "nombre": (nombre_limpio or nombre)[:160],
                    "tipo_accion": accion,
                    "mensaje_instrucciones": msg_limpio or None,
                    "requiere_humano": bool(item.get("requiere_humano")),
                    "activo": True,
                    "configuracion": (
                        {"codigo_recurso": CODIGO_RECURSO_SOLICITUD}
                        if accion == "enviar_enlace"
                        else {}
                    ),
                }
                logger.info(
                    "[CARGA_INFO_MATCH_PASO] entrada_nombre=%r entrada_codigo=%r "
                    "entrada_orden=%s entrada_id=%s match_por_id=%s match_por_codigo=%s "
                    "match_por_nombre=%s match_por_orden=%s paso_existente_id=%s "
                    "accion_final=%s",
                    nombre,
                    item.get("codigo"),
                    item.get("orden"),
                    item.get("id"),
                    via == "id",
                    via == "codigo",
                    via == "nombre",
                    via == "orden",
                    hit.get("id") if hit else None,
                    "actualizar" if hit else "crear",
                )
                logger.info(
                    "[CARGA_INFO_PASO] id_entrada=%s codigo=%s orden=%s "
                    "id_existente=%s accion=%s flujo_id=%s nombre=%r",
                    item.get("id"),
                    (hit or {}).get("codigo") or item.get("codigo"),
                    (hit or {}).get("orden", item.get("orden")),
                    hit.get("id") if hit else None,
                    "actualizar" if hit else "crear",
                    flujo_id,
                    nombre_limpio or nombre,
                )
                if hit:
                    db.actualizar_flujo_paso(
                        agencia_id, int(hit["id"]), campos_p, cur=c
                    )
                    ids_pasos_tocados.add(int(hit["id"]))
                    merged = {**hit, **campos_p}
                    for i, p in enumerate(pasos_db):
                        if int(p.get("id") or 0) == int(hit["id"]):
                            pasos_db[i] = merged
                            break
                    actualizados += 1
                else:
                    codigo = _codigo_unico(
                        item.get("codigo") or nombre,
                        usados_cod_pasos,
                        prefijo=f"paso_{item.get('orden') or 0}",
                        indice=int(item.get("orden") or 0) or 1,
                    )
                    usados_cod_pasos.add(codigo)
                    campos_p.update(
                        {
                            "flujo_id": flujo_id,
                            "codigo": codigo,
                            "obligatorio": True,
                            # orden=None → DB asigna MAX+1 (evita uq_flujo_paso_orden)
                            "orden": None,
                        }
                    )
                    creado = db.crear_flujo_paso(agencia_id, campos_p, cur=c)
                    if creado and creado.get("id"):
                        ids_pasos_tocados.add(int(creado["id"]))
                        pasos_db.append(creado)
                    creados += 1

        textos = obtener_textos_carga(agencia_id, chatbot_configuracion_id, cur=c)

    logger.info(
        "[CARGA_INFO] agencia_id=%s chatbot_configuracion_id=%s accion=guardar "
        "creados=%s actualizados=%s advertencias=%s resultado=ok",
        agencia_id,
        chatbot_configuracion_id,
        creados,
        actualizados,
        len(advertencias),
    )
    return {
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "advertencias": advertencias,
        "textos": textos,
        "general": resumen_general,
    }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def generar_plantilla_excel_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos generales"
    ws.append(
        [
            "nombre_asistente",
            "presentacion_informativo",
            "presentacion_inteligente",
            "tono",
        ]
    )
    ws.append(
        [
            "Asistente virtual",
            "¡Hola! Bienvenido(a) al menú informativo…",
            "¡Hola! Soy el asistente conversacional…",
            "cercano",
        ]
    )

    ws2 = wb.create_sheet("Requisitos")
    ws2.append(["nombre", "descripcion", "tipo", "mensaje_si_no_cumple", "orden"])
    ws2.append(["Mayoría de edad", "Ser mayor de 18 años", "obligatorio", "", 1])

    ws3 = wb.create_sheet("Beneficios y bonos")
    ws3.append(
        [
            "tipo",
            "nombre",
            "descripcion",
            "valor",
            "moneda",
            "condiciones",
            "requiere_validacion_humana",
        ]
    )
    ws3.append(
        ["beneficio", "Acompañamiento de manager", "Seguimiento personalizado", "", "", "", "false"]
    )
    ws3.append(
        ["bono", "Bono de bienvenida", "Solo si está confirmado", "", "", "", "true"]
    )

    ws4 = wb.create_sheet("Preguntas frecuentes")
    ws4.append(
        ["categoria", "pregunta", "respuesta", "palabras_clave", "requiere_humano"]
    )
    ws4.append(
        ["general", "¿La agencia cobra por ingresar?", "No. El ingreso no tiene costo.", "costo,pago", "false"]
    )

    ws5 = wb.create_sheet("Proceso de ingreso")
    ws5.append(
        ["orden", "nombre_paso", "descripcion", "accion", "mensaje", "requiere_humano"]
    )
    ws5.append([1, "Confirmar interés", "", "confirmar_interes", "", "false"])
    ws5.append([2, "Enviar solicitud", "", "enviar_enlace", "", "false"])
    ws5.append([3, "Solicitar evidencias", "", "solicitar_evidencias", "", "false"])
    ws5.append([4, "Revisión humana", "", "transferir_humano", "", "true"])

    ws6 = wb.create_sheet("Enlaces")
    ws6.append(["tipo", "nombre", "url", "texto_boton"])
    ws6.append(["solicitud", "Solicitud principal", "https://", "Completar solicitud"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _celda(row: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    return ""


def importar_excel_a_propuesta(
    agencia_id: int,
    chatbot_configuracion_id: int,
    contenido: bytes,
    *,
    cur=None,
) -> Dict[str, Any]:
    """Valida Excel → propuesta (sin guardar)."""
    from openpyxl import load_workbook

    with db._cursor(cur) as c:
        db._exige_configuracion(agencia_id, chatbot_configuracion_id, cur=c)

    errores: List[str] = []
    try:
        wb = load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ConversacionalError(f"No se pudo leer el Excel: {exc}") from exc

    def filas(nombre: str) -> List[Dict[str, Any]]:
        if nombre not in wb.sheetnames:
            return []
        ws = wb[nombre]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        out = []
        for i, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            out.append({headers[j]: row[j] for j in range(len(headers)) if j < len(row)})
            out[-1]["_fila"] = i
        return out

    general_rows = filas("Datos generales")
    general = {
        "nombre_asistente": "",
        "presentacion_inicial": "",
        "presentacion_informativo": "",
        "presentacion_inteligente": "",
        "tono": "cercano",
    }
    if general_rows:
        g = general_rows[0]
        general["nombre_asistente"] = _celda(g, "nombre_asistente")
        general["presentacion_inicial"] = _celda(g, "presentacion_inicial")
        general["presentacion_informativo"] = _celda(g, "presentacion_informativo")
        general["presentacion_inteligente"] = _celda(g, "presentacion_inteligente")
        tono = _celda(g, "tono").lower() or "cercano"
        general["tono"] = tono if tono in {"profesional", "cercano", "juvenil"} else "cercano"

    requisitos = []
    for row in filas("Requisitos"):
        nombre = _celda(row, "nombre")
        if not nombre:
            errores.append(f"Requisitos, fila {row.get('_fila')}: falta el nombre.")
            continue
        tipo = _celda(row, "tipo").lower() or "obligatorio"
        if tipo not in {"obligatorio", "recomendado", "deseable"}:
            tipo = "obligatorio"
        requisitos.append(
            {
                "nombre": nombre,
                "descripcion": _celda(row, "descripcion") or nombre,
                "categoria": "obligatorio" if tipo == "obligatorio" else "deseable",
                "bloquea_proceso": tipo == "obligatorio",
                "mensaje_si_no_cumple": _celda(row, "mensaje_si_no_cumple") or None,
                "orden": int(row.get("orden") or len(requisitos) + 1),
            }
        )

    beneficios, bonos = [], []
    for row in filas("Beneficios y bonos"):
        nombre = _celda(row, "nombre")
        if not nombre:
            errores.append(
                f"Beneficios y bonos, fila {row.get('_fila')}: falta el nombre."
            )
            continue
        tipo = _celda(row, "tipo").lower() or "beneficio"
        item = {
            "nombre": nombre,
            "descripcion": _celda(row, "descripcion") or nombre,
            "tipo": tipo,
            "valor": row.get("valor"),
            "moneda": _celda(row, "moneda") or None,
            "requiere_validacion_humana": str(
                row.get("requiere_validacion_humana") or ""
            ).lower()
            in {"1", "true", "si", "sí", "yes"},
        }
        if tipo == "bono":
            bonos.append(item)
        else:
            beneficios.append(item)

    faq = []
    for row in filas("Preguntas frecuentes"):
        pregunta = _celda(row, "pregunta")
        respuesta = _celda(row, "respuesta")
        if not pregunta or not respuesta:
            errores.append(
                f"Preguntas frecuentes, fila {row.get('_fila')}: falta la "
                f"{'pregunta' if not pregunta else 'respuesta'}."
            )
            continue
        faq.append(
            {
                "pregunta": pregunta,
                "respuesta": respuesta,
                "categoria": _celda(row, "categoria") or "general",
            }
        )

    proceso = []
    for row in filas("Proceso de ingreso"):
        nombre = _celda(row, "nombre_paso", "nombre")
        if not nombre:
            errores.append(
                f"Proceso de ingreso, fila {row.get('_fila')}: falta el nombre del paso."
            )
            continue
        proceso.append(
            {
                "orden": int(row.get("orden") or len(proceso) + 1),
                "nombre": nombre,
                "descripcion": _celda(row, "descripcion"),
                "accion": _celda(row, "accion") or "informar",
                "mensaje": _celda(row, "mensaje") or _celda(row, "descripcion"),
                "requiere_humano": str(row.get("requiere_humano") or "").lower()
                in {"1", "true", "si", "sí", "yes"},
            }
        )

    recursos = []
    for row in filas("Enlaces"):
        url = _url_real(_celda(row, "url"))
        if not url:
            if _celda(row, "url"):
                errores.append(
                    f"Enlaces, fila {row.get('_fila')}: la URL no es válida."
                )
            continue
        recursos.append(
            {
                "tipo": _celda(row, "tipo") or "solicitud",
                "nombre": _celda(row, "nombre") or "Enlace",
                "url": url,
                "texto_boton": _celda(row, "texto_boton") or "Abrir enlace",
            }
        )

    propuesta = {
        "general": general,
        "requisitos": requisitos,
        "beneficios": beneficios,
        "bonos": bonos,
        "faq": faq,
        "proceso_ingreso": proceso,
        "recursos": recursos,
        "contacto_humano": {"equipo_destino": ""},
        "advertencias": errores,
        "errores": errores,
        "origen_analisis": "excel",
    }
    with db._cursor(cur) as c:
        propuesta = _anotar_ids_existentes(
            agencia_id, chatbot_configuracion_id, propuesta, cur=c
        )
    return propuesta
